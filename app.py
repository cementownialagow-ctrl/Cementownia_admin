# -*- coding: utf-8 -*-
import os
import io
import csv
import base64
import re
import json
import hashlib
import glob
import sqlite3
import socket
import time
import threading
import uuid
import secrets
import hmac
import urllib.parse
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from flask import (
    Flask, request, redirect, url_for, jsonify, session, g,
    send_file, abort
)
from flask import render_template_string
from jinja2 import DictLoader
from werkzeug.security import check_password_hash, generate_password_hash

from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from ksef_module import build_ksef_draft_xml, validate_fa3_xml, validate_ksef_invoice, xml_filename
from cash_flow_module import register_cash_flow
from beton_logistics_module import register_beton_logistics, create_wz_from_order
from dispatch_module import register_dispatch
from operations_module import register_operations
try:
    from ksef_api import ksef_config_summary, send_invoice_to_ksef
except Exception:
    send_invoice_to_ksef = None

    def ksef_config_summary():
        return {"configured": False, "missing": ["ksef_api.py"], "env": "", "base_url": ""}

_EMAIL_IMPORT_ERROR = ""
# Wersja dla Beton Łagów nie wysyła automatycznych wiadomości do klientów.
# Adres e-mail pozostaje jedynie polem kontaktowym na dokumencie.
EMAIL_NOTIFICATIONS_ENABLED = False
try:
    from email_module import (
        email_config_summary,
        send_email,
        send_order_confirmation,
        send_invoice_available,
        send_payment_reminder,
    )
except Exception as exc:
    _EMAIL_IMPORT_ERROR = str(exc)
    send_email = None
    send_order_confirmation = None
    send_invoice_available = None
    send_payment_reminder = None

    def email_config_summary():
        return {
            "configured": False,
            "missing": ["email_module.py"],
            "enabled": False,
            "import_error": _EMAIL_IMPORT_ERROR,
        }


# =========================
# KONFIG
# =========================

# TWOJE IP (z ipconfig -> IPv4)
BASE_URL = (os.environ.get("PUBLIC_BASE_URL") or "http://127.0.0.1:5000").strip().rstrip("/")

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(APP_DIR, "data"))
DB_PATH = os.path.join(DATA_DIR, "app.db")

os.makedirs(DATA_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20 MB
app.config["JSON_AS_ASCII"] = False
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)

@app.get("/brand-logo.png")
def brand_logo():
    """Serves the logo placed next to app.py (the deployment layout used on GitHub)."""
    path = find_logo_path()
    if not path:
        abort(404)
    return send_file(path, mimetype="image/png", conditional=True)


_MOJIBAKE_REPLACEMENTS = {
    "Ä…": "ą", "Ä‡": "ć", "Ä™": "ę", "Ĺ‚": "ł", "Ĺ„": "ń",
    "Ăł": "ó", "Ĺ›": "ś", "Ĺş": "ź", "ĹĽ": "ż",
    "Ä„": "Ą", "Ä†": "Ć", "Ä": "Ę", "Ĺ": "Ł", "Ĺƒ": "Ń",
    "Ă“": "Ó", "Ĺš": "Ś", "Ĺą": "Ź", "Ĺ»": "Ż",
    "Ã³": "ó", "Å‚": "ł", "Å„": "ń", "Å›": "ś", "Åº": "ź",
    "Å¼": "ż", "Å": "Ł", "Åƒ": "Ń", "Åš": "Ś", "Å¹": "Ź",
    "Å»": "Ż", "Ä": "ą", "Ä": "ć", "Ä": "ę",
    "â€˘": "•", "â€¢": "•", "â€“": "–", "â€”": "—",
    "â€ž": "„", "â€ť": "”", "â€ś": "“", "â€™": "’",
    "â†": "←", "â†’": "→",
}


def fix_polish_mojibake(text: str) -> str:
    if not text or not any(marker in text for marker in ("Ä", "Ĺ", "Ă", "Å", "Ã", "â")):
        return text
    for bad, good in _MOJIBAKE_REPLACEMENTS.items():
        text = text.replace(bad, good)
    return text


@app.after_request
def force_utf8_html(response):
    if response.direct_passthrough:
        return response
    if response.mimetype in {"text/html", "text/plain", "application/json"}:
        response.headers["Content-Type"] = f"{response.mimetype}; charset=utf-8"
    if response.mimetype == "text/html":
        body = response.get_data(as_text=True)
        fixed = fix_polish_mojibake(body)
        if fixed != body:
            response.set_data(fixed)
    return response


def _detect_lan_base_url(port: int) -> str:
    try:
        sck = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sck.connect(("8.8.8.8", 80))
        ip = sck.getsockname()[0]
        sck.close()
        if ip and not ip.startswith("127."):
            return f"http://{ip}:{port}"
    except Exception:
        pass
    return ""


def build_public_url(path: str) -> str:
    # Dla QR preferuj adres LAN; jeĹ›li aplikacja jest otwarta lokalnie,
    # sprĂłbuj wykryÄ‡ LAN IP automatycznie (bardziej niezawodne niĹĽ staĹ‚y BASE_URL).
    base_cfg = (BASE_URL or "").rstrip("/")
    try:
        host = (request.host or "").split(":")[0].lower()
        req_base = (request.host_url or "").rstrip("/")
        req_port = to_int((request.host or "").split(":")[1] if ":" in (request.host or "") else 5000, 5000)
    except RuntimeError:
        host = ""
        req_base = ""
        req_port = 5000

    if host in {"localhost", "127.0.0.1", "::1", "0.0.0.0"}:
        base = _detect_lan_base_url(req_port) or base_cfg or req_base
    else:
        base = req_base or base_cfg or _detect_lan_base_url(req_port)

    return f"{base}{path}"


# =========================
# DB
# =========================

def conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    c = conn()
    cur = c.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sku TEXT UNIQUE NOT NULL,
        model TEXT,
        ean TEXT,
        name TEXT,
        unit TEXT NOT NULL DEFAULT 'm3',
        unit_net_price REAL NOT NULL DEFAULT 0,
        unit_gross_price REAL NOT NULL DEFAULT 0,
        unit_material_cost REAL NOT NULL DEFAULT 0,
        unit_production_cost REAL NOT NULL DEFAULT 0,
        unit_transport_cost REAL NOT NULL DEFAULT 0,
        unit_other_cost REAL NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS stock(
        product_id INTEGER PRIMARY KEY,
        qty INTEGER NOT NULL DEFAULT 0,
        FOREIGN KEY(product_id) REFERENCES products(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS customers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        address TEXT,
        phone TEXT,
        email TEXT,
        nip TEXT,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_no TEXT UNIQUE NOT NULL,
        customer_id INTEGER,
        customer_name TEXT NOT NULL,
        customer_address TEXT,
        customer_phone TEXT,
        customer_email TEXT,
        delivery_date TEXT,
        delivery_time TEXT,
        status TEXT NOT NULL DEFAULT 'new', -- new/packed/shipped/cancelled
        note TEXT,
        created_at TEXT NOT NULL,
        warehouse_issued INTEGER NOT NULL DEFAULT 0,
        FOREIGN KEY(customer_id) REFERENCES customers(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS order_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        product_id INTEGER NOT NULL,
        sku TEXT NOT NULL,
        qty REAL NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(order_id) REFERENCES orders(id),
        FOREIGN KEY(product_id) REFERENCES products(id)
    )
    """)

    # Paczki z materiałów (prosty moduĹ‚ na start)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS material_orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        package_no TEXT UNIQUE NOT NULL,
        status TEXT NOT NULL DEFAULT 'planned', -- planned/ordered/shipped/arrived
        tracking TEXT,
        note TEXT,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS material_order_items(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        package_id INTEGER NOT NULL,
        product_id INTEGER NOT NULL,
        sku TEXT NOT NULL,
        qty INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(package_id) REFERENCES material_orders(id),
        FOREIGN KEY(product_id) REFERENCES products(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS pricing(
        model TEXT PRIMARY KEY,
        net_price REAL NOT NULL DEFAULT 0,
        gross_price REAL NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS company_profile(
        id INTEGER PRIMARY KEY CHECK(id=1),
        company_name TEXT,
        address TEXT,
        nip TEXT,
        phone TEXT,
        email TEXT,
        bank_account TEXT,
        updated_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS invoices(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        invoice_no TEXT NOT NULL,
        issue_date TEXT NOT NULL,
        sell_date TEXT NOT NULL,
        payment_type TEXT NOT NULL,
        payment_to TEXT,
        buyer_name TEXT,
        buyer_tax_no TEXT,
        buyer_street TEXT,
        buyer_post_code TEXT,
        buyer_city TEXT,
        buyer_country TEXT,
        buyer_email TEXT,
        buyer_phone TEXT,
        total_net REAL NOT NULL DEFAULT 0,
        total_gross REAL NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL,
        UNIQUE(invoice_no),
        FOREIGN KEY(order_id) REFERENCES orders(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS invoice_meta(
        invoice_id INTEGER PRIMARY KEY,
        pdf_path TEXT,
        invoice_items_json TEXT,
        sent_to_client INTEGER NOT NULL DEFAULT 0,
        seen_by_client INTEGER NOT NULL DEFAULT 0,
        payment_reminder INTEGER NOT NULL DEFAULT 0,
        paid INTEGER NOT NULL DEFAULT 0,
        paid_at TEXT,
        seen_at TEXT,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(invoice_id) REFERENCES invoices(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS ksef_documents(
        invoice_id INTEGER PRIMARY KEY,
        status TEXT NOT NULL DEFAULT 'draft',
        ksef_number TEXT,
        xml_path TEXT,
        last_error TEXT,
        validated_at TEXT,
        sent_at TEXT,
        updated_at TEXT NOT NULL,
        FOREIGN KEY(invoice_id) REFERENCES invoices(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS cash_flow_settings(
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL DEFAULT '',
        updated_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS invoice_allocations(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id INTEGER NOT NULL,
        order_id INTEGER NOT NULL,
        order_item_id INTEGER NOT NULL,
        product_id INTEGER,
        sku TEXT,
        qty INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(invoice_id) REFERENCES invoices(id),
        FOREIGN KEY(order_id) REFERENCES orders(id),
        FOREIGN KEY(order_item_id) REFERENCES order_items(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS client_search_logs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_email TEXT,
        customer_name TEXT,
        query TEXT NOT NULL,
        product_sku TEXT,
        product_model TEXT,
        product_name TEXT,
        results_count INTEGER NOT NULL DEFAULT 0,
        source TEXT,
        created_at TEXT NOT NULL
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS email_events(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_key TEXT UNIQUE,
        event_type TEXT NOT NULL,
        ref_id TEXT,
        recipient TEXT,
        ok INTEGER NOT NULL DEFAULT 0,
        result_json TEXT,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("PRAGMA table_info(client_search_logs)")
    search_cols = {r["name"] for r in cur.fetchall()}
    if "product_sku" not in search_cols:
        cur.execute("ALTER TABLE client_search_logs ADD COLUMN product_sku TEXT")
    if "product_model" not in search_cols:
        cur.execute("ALTER TABLE client_search_logs ADD COLUMN product_model TEXT")
    if "product_name" not in search_cols:
        cur.execute("ALTER TABLE client_search_logs ADD COLUMN product_name TEXT")

    cur.execute("PRAGMA table_info(invoice_meta)")
    invoice_meta_cols = {r[1] for r in cur.fetchall()}
    if "seen_by_client" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN seen_by_client INTEGER NOT NULL DEFAULT 0")
    if "seen_at" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN seen_at TEXT")
    if "payment_reminder" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN payment_reminder INTEGER NOT NULL DEFAULT 0")
    if "paid" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN paid INTEGER NOT NULL DEFAULT 0")
    if "paid_at" not in invoice_meta_cols:
        cur.execute("ALTER TABLE invoice_meta ADD COLUMN paid_at TEXT")

    # migracja: starsze bazy mogÄ… nie mieÄ‡ kolumny NIP u klientĂłw
    cur.execute("PRAGMA table_info(customers)")
    customer_cols = {r[1] for r in cur.fetchall()}
    if "nip" not in customer_cols:
        cur.execute("ALTER TABLE customers ADD COLUMN nip TEXT")

    cur.execute("PRAGMA table_info(orders)")
    order_cols = {r[1] for r in cur.fetchall()}
    if "warehouse_issued" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN warehouse_issued INTEGER NOT NULL DEFAULT 0")
    if "idempotency_key" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN idempotency_key TEXT")
    if "delivery_date" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN delivery_date TEXT")
    if "delivery_time" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN delivery_time TEXT")

    product_cols = {r[1] for r in cur.execute("PRAGMA table_info(products)").fetchall()}
    for col, definition in {
        "unit": "TEXT NOT NULL DEFAULT 'm3'",
        "unit_net_price": "REAL NOT NULL DEFAULT 0",
        "unit_gross_price": "REAL NOT NULL DEFAULT 0",
        "unit_material_cost": "REAL NOT NULL DEFAULT 0",
        "unit_production_cost": "REAL NOT NULL DEFAULT 0",
        "unit_transport_cost": "REAL NOT NULL DEFAULT 0",
        "unit_other_cost": "REAL NOT NULL DEFAULT 0",
    }.items():
        if col not in product_cols:
            cur.execute(f"ALTER TABLE products ADD COLUMN {col} {definition}")

    # Migracja nazewnictwa z odziedziczonego modułu „China” do zamówień materiałów.
    old_material_tables = {r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    if {"china_packages", "china_items"}.issubset(old_material_tables):
        if cur.execute("SELECT COUNT(*) FROM material_orders").fetchone()[0] == 0:
            cur.execute("""INSERT OR IGNORE INTO material_orders(id,package_no,status,tracking,note,created_at)
                           SELECT id,package_no,status,tracking,note,created_at FROM china_packages""")
            cur.execute("""INSERT OR IGNORE INTO material_order_items(id,package_id,product_id,sku,qty,created_at)
                           SELECT id,package_id,product_id,sku,qty,created_at FROM china_items""")

    # UĹ‚atwia agregowanie "w dostawie" po statusach paczek
    cur.execute("CREATE INDEX IF NOT EXISTS idx_material_order_items_package_id ON material_order_items(package_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_material_order_items_product_id ON material_order_items(product_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_client_search_logs_created ON client_search_logs(created_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_client_search_logs_email_query ON client_search_logs(customer_email, query)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_client_search_logs_model ON client_search_logs(product_model)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_email_events_key ON email_events(event_key)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_email_events_type_created ON email_events(event_type, created_at)")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_orders_idempotency_key ON orders(idempotency_key) WHERE idempotency_key IS NOT NULL")

    cur.executescript("""
    CREATE TABLE IF NOT EXISTS app_users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE COLLATE NOCASE,
        display_name TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'admin',
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        last_login_at TEXT,
        deleted_at TEXT
    );
    CREATE TABLE IF NOT EXISTS audit_events(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        request_id TEXT NOT NULL,
        actor_type TEXT NOT NULL,
        actor_id TEXT,
        actor_username TEXT NOT NULL,
        actor_display_name TEXT,
        actor_role TEXT,
        action TEXT NOT NULL,
        method TEXT NOT NULL,
        path TEXT NOT NULL,
        entity_type TEXT,
        entity_id TEXT,
        payload_json TEXT,
        response_status INTEGER NOT NULL,
        ip_address TEXT,
        user_agent TEXT,
        created_at TEXT NOT NULL
    );
    CREATE INDEX IF NOT EXISTS idx_audit_events_created ON audit_events(created_at DESC);
    CREATE INDEX IF NOT EXISTS idx_audit_events_actor ON audit_events(actor_username,created_at DESC);
    CREATE INDEX IF NOT EXISTS idx_audit_events_path ON audit_events(path,created_at DESC);
    """)

    c.commit()
    c.close()

init_db()


# =========================
# UTILS
# =========================

APP_TZ = ZoneInfo("Europe/Warsaw") if ZoneInfo else None

def app_now():
    return datetime.now(APP_TZ) if APP_TZ else datetime.now()

def now_iso():
    return app_now().strftime("%Y-%m-%d %H:%M:%S")

SHORT_ORDER_NO_RE = re.compile(r"^ZAM-(\d{6})(\d+)$", re.I)


def order_date_code(created_at: str | None = "") -> str:
    created = norm(created_at)
    if len(created) >= 10 and created[4:5] == "-" and created[7:8] == "-":
        return created[2:4] + created[5:7] + created[8:10]
    return app_now().strftime("%y%m%d")


def is_short_order_no(value: str | None) -> bool:
    return bool(SHORT_ORDER_NO_RE.match(norm(value)))


def make_order_no(order_id: int | None = None, created_at: str | None = "") -> str:
    # Format: ZAM-2607141 = ZAM- + YYMMDD + kolejny numer w danym dniu.
    date_code = order_date_code(created_at)
    day = norm(created_at)[:10] if norm(created_at) else app_now().strftime("%Y-%m-%d")
    seq = 1
    try:
        c = conn()
        cur = c.cursor()
        if order_id:
            cur.execute("SELECT order_no FROM orders WHERE substr(created_at,1,10)=? AND id<>?", (day, int(order_id)))
        else:
            cur.execute("SELECT order_no FROM orders WHERE substr(created_at,1,10)=?", (day,))
        for r in cur.fetchall():
            raw = norm(r["order_no"])
            m = SHORT_ORDER_NO_RE.match(raw)
            if m and m.group(1) == date_code:
                seq = max(seq, int(m.group(2)) + 1)
            elif raw and raw.upper() != "TEMP":
                seq += 1
        c.close()
    except Exception:
        oid = int(order_id or 1)
        seq = max(1, oid)
    return f"ZAM-{date_code}{seq}"


def canonical_order_no(order_id: int | None, created_at: str | None = "", raw_order_no: str | None = "") -> str:
    raw = norm(raw_order_no)
    if raw and raw.upper() != "TEMP":
        if raw.startswith("ORD-"):
            return "ZAM-" + raw[4:]
        return raw

    return make_order_no(order_id, created_at)


def order_display_no(order_id: int | None, created_at: str | None = "", raw_order_no: str | None = "", note: str | None = "") -> str:
    base = canonical_order_no(order_id, created_at, raw_order_no)
    note_text = norm(note)
    return f"{base} {note_text}" if note_text else base


def normalize_temp_order_numbers():
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id, order_no, created_at FROM orders ORDER BY created_at, id")
    rows = cur.fetchall()
    changed = []
    used_seq_by_date = {}
    for r in rows:
        raw = norm(r["order_no"])
        m = SHORT_ORDER_NO_RE.match(raw)
        if m:
            used_seq_by_date[m.group(1)] = max(used_seq_by_date.get(m.group(1), 0), int(m.group(2)))

    for r in rows:
        raw = norm(r["order_no"])
        if is_short_order_no(raw):
            continue
        date_code = order_date_code(r["created_at"])
        used_seq_by_date[date_code] = used_seq_by_date.get(date_code, 0) + 1
        new_no = f"ZAM-{date_code}{used_seq_by_date[date_code]}"
        if new_no != (r["order_no"] or ""):
            cur.execute("UPDATE orders SET order_no=? WHERE id=?", (new_no, r["id"]))
            changed.append((int(r["id"]), new_no))
    c.commit()
    c.close()

    if supabase_enabled():
        for oid, ono in changed:
            try:
                supabase_update_rows("orders", {"order_no": ono}, {"id": oid})
            except Exception:
                pass
    return len(changed)

def _email_key(value: str) -> str:
    return norm(value).strip().lower()

def _order_name_is_fallback(order_name: str, email_value: str) -> bool:
    email_key = _email_key(email_value)
    if not email_key:
        return False
    local_part = email_key.split("@")[0]
    current = norm(order_name).strip().lower()
    return current in {"", email_key, local_part}

def link_orders_to_customers_by_email(sync_remote: bool = True):
    c = conn()
    cur = c.cursor()

    cur.execute("""
      SELECT id, name, address, phone, email
      FROM customers
      WHERE TRIM(COALESCE(email, '')) <> ''
      ORDER BY id
    """)
    customer_rows = [dict(r) for r in cur.fetchall()]
    customers_by_email = {_email_key(r["email"]): r for r in customer_rows if _email_key(r.get("email"))}

    cur.execute("""
      SELECT id, customer_id, customer_name, customer_address, customer_phone, customer_email
      FROM orders
      WHERE TRIM(COALESCE(customer_email, '')) <> ''
      ORDER BY id
    """)
    order_rows = [dict(r) for r in cur.fetchall()]

    changed = []
    for order_row in order_rows:
        email_key = _email_key(order_row.get("customer_email"))
        customer = customers_by_email.get(email_key)
        if not customer:
            continue

        updates = {}
        if int(order_row.get("customer_id") or 0) != int(customer["id"]):
            updates["customer_id"] = int(customer["id"])

        if _order_name_is_fallback(order_row.get("customer_name"), order_row.get("customer_email")) and norm(customer.get("name")):
            updates["customer_name"] = norm(customer.get("name"))

        if not norm(order_row.get("customer_address")) and norm(customer.get("address")):
            updates["customer_address"] = norm(customer.get("address"))

        if not norm(order_row.get("customer_phone")) and norm(customer.get("phone")):
            updates["customer_phone"] = norm(customer.get("phone"))

        if not norm(order_row.get("customer_email")) and norm(customer.get("email")):
            updates["customer_email"] = norm(customer.get("email"))

        if not updates:
            continue

        sets = ", ".join([f"{k}=?" for k in updates.keys()])
        values = list(updates.values()) + [int(order_row["id"])]
        cur.execute(f"UPDATE orders SET {sets} WHERE id=?", values)
        changed.append((int(order_row["id"]), updates))

    c.commit()
    c.close()

    if sync_remote and supabase_enabled():
        for order_id, updates in changed:
            try:
                supabase_update_rows("orders", updates, {"id": order_id})
            except Exception:
                pass

    return len(changed)


def next_invoice_no(issue_date: str) -> str:
    dt = datetime.strptime(issue_date, "%Y-%m-%d")
    mm = dt.strftime("%m")
    yyyy = dt.strftime("%Y")
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT COUNT(*) AS n FROM invoices WHERE substr(issue_date,1,7)=?", (f"{yyyy}-{mm}",))
    n = int(cur.fetchone()["n"] or 0) + 1
    c.close()
    return f"FVAT {n}/{mm}/{yyyy}"


def invoice_no_exists(invoice_no: str, exclude_invoice_id: int = 0) -> int:
    invoice_no = norm(invoice_no)
    if not invoice_no:
        return 0
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT id
      FROM invoices
      WHERE lower(trim(invoice_no)) = lower(trim(?))
        AND id <> ?
      LIMIT 1
    """, (invoice_no, int(exclude_invoice_id or 0)))
    row = cur.fetchone()
    c.close()
    return int(row["id"]) if row else 0


def split_address(addr: str):
    raw = (addr or "").strip()
    if not raw:
        return "", "", ""

    # wspieraj adres w wielu liniach oraz jednoliniowy (np. "ul. X 1, 00-001 Warszawa")
    parts = [x.strip() for x in raw.splitlines() if x.strip()]
    if len(parts) == 1 and "," in raw:
        comma_parts = [x.strip() for x in raw.split(",") if x.strip()]
        if len(comma_parts) >= 2:
            parts = [comma_parts[0], " ".join(comma_parts[1:])]

    street = parts[0] if parts else ""
    post_code = ""
    city = ""
    if len(parts) > 1:
        line2 = parts[1].strip()
        m = re.match(r"^(\d{2}-\d{3})\s*(.*)$", line2)
        if m:
            post_code = m.group(1).strip()
            city = m.group(2).strip()
        else:
            pc = line2.split(" ", 1)
            post_code = pc[0].strip() if pc else ""
            city = pc[1].strip() if len(pc) > 1 else ""
    return street, post_code, city


def payment_type_pl(x: str) -> str:
    v = norm(x).lower()
    mapping = {
        "cash": "gotĂłwka",
        "gotowka": "gotĂłwka",
        "transfer": "przelew",
        "card": "karta",
        "karta": "karta",
    }
    return mapping.get(v, v or "-")


VAT_23 = Decimal("0.23")
MONEY_Q = Decimal("0.01")
CURRENT_ORDER_STATUSES = {"new", "pending", "unconfirmed", "confirmed", "packed", "in_delivery", "shipped"}


def money_dec(value) -> Decimal:
    try:
        return Decimal(str(value or "0")).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0.00")


def money_float(value) -> float:
    return float(money_dec(value))


def vat23_from_net(net_value) -> Decimal:
    return (money_dec(net_value) * VAT_23).quantize(MONEY_Q, rounding=ROUND_HALF_UP)


def gross_from_net_23(net_value) -> Decimal:
    net = money_dec(net_value)
    return (net + vat23_from_net(net)).quantize(MONEY_Q, rounding=ROUND_HALF_UP)


def find_logo_path() -> str:
    search_dirs = [
        APP_DIR,
        os.path.join(APP_DIR, "static"),
        DATA_DIR,
    ]
    for folder in search_dirs:
        for fn in ("logo.png", "logo.jpg", "logo.jpeg", "logo.webp"):
            pth = os.path.join(folder, fn)
            if os.path.exists(pth):
                return pth
    return ""


def to_int(x, default=0):
    try:
        return int(str(x).strip())
    except:
        return default

def to_float(x, default=0.0):
    try:
        return float(str(x).strip().replace(" ", "").replace(",", "."))
    except:
        return default

def norm(s):
    if s is None:
        return ""
    return str(s).strip()

def order_status_label(status: str) -> str:
    v = norm(status).lower()
    mapping = {
        "new": "Niepotwierdzone",
        "pending": "Niepotwierdzone",
        "unconfirmed": "Niepotwierdzone",
        "confirmed": "Potwierdzone",
        "packed": "W dostawie",
        "in_delivery": "W dostawie",
        "issued": "Zrealizowane",
    }
    return mapping.get(v, status or "-")

def order_status_css(status: str) -> str:
    v = norm(status).lower()
    mapping = {
        "new": "st-unconfirmed",
        "pending": "st-unconfirmed",
        "unconfirmed": "st-unconfirmed",
        "confirmed": "st-confirmed",
        "packed": "st-delivery",
        "in_delivery": "st-delivery",
        "issued": "st-issued",
    }
    return mapping.get(v, "")

def guess_col(headers, candidates):
    h = [x.strip().lower() for x in headers]
    for cand in candidates:
        cand = cand.lower()
        if cand in h:
            return h.index(cand)
    # luĹşne dopasowanie: np. "model" w "Model uchwytu"
    for i, col in enumerate(h):
        for cand in candidates:
            if cand.lower() in col:
                return i
    return None

def ensure_stock_row(product_id):
    c = conn()
    cur = c.cursor()
    cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (product_id,))
    c.commit()
    c.close()


# =========================
# SUPABASE (cloud sync)
# =========================

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "").strip()
CLIENT_ALLOWED_ORIGINS = {
    value.strip().rstrip("/")
    for value in os.environ.get("CLIENT_ALLOWED_ORIGINS", "").split(",")
    if value.strip()
}
# Panel kierowcy jest osobną stroną Netlify. Ten bezpieczny, znany adres musi
# zawsze przejść przez CORS, aby kierowca mógł wysłać login do API na Render.
# Dodatkowe własne domeny nadal można dopisać w CLIENT_ALLOWED_ORIGINS.
CLIENT_ALLOWED_ORIGINS.add("https://panel-dostawy.netlify.app")
ADMIN_ACTION_TOKEN = os.environ.get("ADMIN_ACTION_TOKEN", "").strip()
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin").strip()
ADMIN_PASSWORD_HASH = os.environ.get("ADMIN_PASSWORD_HASH", "").strip()
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "").strip()
LEGACY_CLIENT_API_ENABLED = os.environ.get("ENABLE_LEGACY_CLIENT_API", "0").strip().lower() in ("1", "true", "yes", "on")

def bootstrap_admin_user():
    """Jednorazowo przenosi konto z konfiguracji środowiska do imiennej bazy użytkowników."""
    c = conn()
    try:
        if c.execute("SELECT COUNT(*) FROM app_users WHERE deleted_at IS NULL").fetchone()[0]:
            return
        password_hash = ADMIN_PASSWORD_HASH
        if not password_hash and ADMIN_PASSWORD:
            password_hash = generate_password_hash(ADMIN_PASSWORD)
        if password_hash:
            stamp = now_iso()
            c.execute("""INSERT INTO app_users(username,display_name,password_hash,role,created_at,updated_at)
                         VALUES(?,?,?,?,?,?)""",
                      (ADMIN_USERNAME or "admin", "Administrator główny", password_hash, "admin", stamp, stamp))
            c.commit()
    finally:
        c.close()

bootstrap_admin_user()
if not SUPABASE_SERVICE_ROLE_KEY:
    app.logger.error("Brak SUPABASE_SERVICE_ROLE_KEY; funkcje synchronizacji i zamówień klienta będą niedostępne.")
if not CLIENT_ALLOWED_ORIGINS:
    app.logger.warning("CLIENT_ALLOWED_ORIGINS jest puste; przeglądarkowe żądania tworzenia zamówień będą odrzucane.")
SUPABASE_STORAGE_BUCKET = (os.environ.get("SUPABASE_STORAGE_BUCKET") or "invoice-pdfs").strip()
SUPABASE_AUTO_SYNC_ON_WRITE = (os.environ.get("SUPABASE_AUTO_SYNC_ON_WRITE") or "1").strip().lower() in ("1", "true", "yes", "on")
# Na darmowym Renderze plik SQLite jest tymczasowy. Domyślnie każda zmiana
# czeka na potwierdzenie zapisu w Supabase zanim odpowiedź opuści serwer.
SUPABASE_SYNC_BEFORE_RESPONSE = (os.environ.get("SUPABASE_SYNC_BEFORE_RESPONSE") or "1").strip().lower() in ("1", "true", "yes", "on")
INVENTORY_AUTOMATION_ENABLED = (os.environ.get("INVENTORY_AUTOMATION_ENABLED") or "0").strip().lower() in ("1", "true", "yes", "on")
SUPABASE_MIN_SYNC_INTERVAL_SEC = float((os.environ.get("SUPABASE_MIN_SYNC_INTERVAL_SEC") or "2").strip())
SUPABASE_MIN_PULL_INTERVAL_SEC = float((os.environ.get("SUPABASE_MIN_PULL_INTERVAL_SEC") or "2").strip())

SUPABASE_SYNC_TABLES = [
    ("products", "id"),
    # Nie sterujemy automatycznie stanem magazynowym, ale jeżeli zostanie
    # wpisany ręcznie, jego kopia również musi trafić do chmury.
    ("stock", "product_id"),
    ("customers", "id"),
    ("orders", "id"),
    ("order_items", "id"),
    ("material_orders", "id"),
    ("material_order_items", "id"),
    ("pricing", "model"),
    ("company_profile", "id"),
    ("app_users", "id"),
    ("cash_flow_settings", "key"),
    ("invoices", "id"),
    ("invoice_meta", "invoice_id"),
    ("invoice_allocations", "id"),
    ("ksef_documents", "invoice_id"),
    ("drivers", "id"),
    ("driver_accounts", "driver_id"),
    ("vehicles", "id"),
    ("wz_documents", "id"),
    ("wz_items", "id"),
    ("transports", "id"),
    ("transport_items", "id"),
    ("delivery_photos", "id"),
    ("loading_bays", "id"),
    ("dispatch_appointments", "id"),
    ("appointment_status_history", "id"),
    ("audit_log", "id"),
    ("audit_events", "id"),
    ("email_events", "id"),
    ("departments", "id"),
    ("material_usage", "id"),
    ("fuel_entries", "id"),
    ("expense_categories", "id"),
    ("vehicle_expenses", "id"),
]

# KolejnoĹ›Ä‡ PULL jest waĹĽna: najpierw rodzice, potem dzieci.
SUPABASE_PULL_TABLES = [
    ("company_profile", "id"),
    ("pricing", "model"),
    ("app_users", "id"),
    ("cash_flow_settings", "key"),
    ("customers", "id"),
    ("products", "id"),
    ("stock", "product_id"),
    ("drivers", "id"),
    ("driver_accounts", "driver_id"),
    ("orders", "id"),
    ("material_orders", "id"),
    ("order_items", "id"),
    ("material_order_items", "id"),
    ("invoices", "id"),
    ("invoice_meta", "invoice_id"),
    ("invoice_allocations", "id"),
    ("ksef_documents", "invoice_id"),
    ("vehicles", "id"),
    ("wz_documents", "id"),
    ("wz_items", "id"),
    ("transports", "id"),
    ("transport_items", "id"),
    ("delivery_photos", "id"),
    ("loading_bays", "id"),
    ("dispatch_appointments", "id"),
    ("appointment_status_history", "id"),
    ("audit_log", "id"),
    ("audit_events", "id"),
    ("email_events", "id"),
    ("departments", "id"),
    ("expense_categories", "id"),
    ("material_usage", "id"),
    ("fuel_entries", "id"),
    ("vehicle_expenses", "id"),
]

_supabase_sync_lock = threading.Lock()
_supabase_sync_state = {
    "running": False,
    "last_started_ts": 0.0,
    "last_result": None,
}

def supabase_enabled() -> bool:
    return bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)


def cloud_row_id() -> int:
    """Return a bigint ID for Supabase tables without an identity generator."""
    return int(time.time() * 1000) * 1000 + secrets.randbelow(1000)

def _chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]

def supabase_upsert_rows(table: str, rows: list, on_conflict: str):
    if not rows:
        return
    if not supabase_enabled():
        raise RuntimeError("Brak konfiguracji SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")

    qs = urllib.parse.urlencode({"on_conflict": on_conflict})
    url = f"{SUPABASE_URL}/rest/v1/{table}?{qs}"
    payload = json.dumps(rows, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=payload, method="POST")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "resolution=merge-duplicates,return=minimal")
    with urllib.request.urlopen(req, timeout=60) as resp:
        if resp.status >= 300:
            raise RuntimeError(f"Supabase HTTP {resp.status}")

def sqlite_table_rows(table: str):
    c = conn()
    cur = c.cursor()
    cur.execute(f"SELECT * FROM {table}")
    rows = [dict(r) for r in cur.fetchall()]
    c.close()
    return rows

def sync_all_to_supabase():
    if not supabase_enabled():
        return {"ok": False, "error": "Brak konfiguracji SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY"}

    out = {"ok": True, "tables": {}, "synced_at": now_iso()}
    for table, conflict_col in SUPABASE_SYNC_TABLES:
        try:
            rows = sqlite_table_rows(table)
            for pack in _chunks(rows, 500):
                supabase_upsert_rows(table, pack, conflict_col)
            out["tables"][table] = {"rows": len(rows), "status": "ok"}
        except Exception as e:
            out["ok"] = False
            out["tables"][table] = {"status": "error", "error": str(e)}
    return out

def trigger_background_supabase_sync(reason: str = "write"):
    if not SUPABASE_AUTO_SYNC_ON_WRITE:
        return False, "disabled"
    if not supabase_enabled():
        return False, "not_configured"

    now_ts = time.time()
    with _supabase_sync_lock:
        if _supabase_sync_state["running"]:
            return False, "already_running"
        if (now_ts - float(_supabase_sync_state["last_started_ts"])) < SUPABASE_MIN_SYNC_INTERVAL_SEC:
            return False, "throttled"
        _supabase_sync_state["running"] = True
        _supabase_sync_state["last_started_ts"] = now_ts

    def _job():
        try:
            result = sync_all_to_supabase()
            result["reason"] = reason
        except Exception as e:
            result = {"ok": False, "error": str(e), "reason": reason, "synced_at": now_iso()}
        finally:
            with _supabase_sync_lock:
                _supabase_sync_state["running"] = False
                _supabase_sync_state["last_result"] = result

    th = threading.Thread(target=_job, daemon=True)
    th.start()
    return True, "started"



def supabase_json_value(value):
    """Prepare values for PostgREST without turning whole quantities into 14.0.

    Several existing Supabase quantity columns are bigint. Python converts an
    HTML number such as ``14`` to float ``14.0`` while processing m³ values;
    PostgreSQL then rejects the JSON value for a bigint column. Whole values
    are therefore serialized as integers, recursively for all API payloads.
    """
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, dict):
        return {key: supabase_json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [supabase_json_value(item) for item in value]
    return value


def supabase_request(path: str, method: str = "GET", params: dict | None = None, payload=None, prefer: str | None = None, timeout: int = 60, use_anon_key: bool = False):
    if not supabase_enabled():
        raise RuntimeError("Brak konfiguracji SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")

    url = f"{SUPABASE_URL}{path}"
    if params:
        qs = urllib.parse.urlencode(params, doseq=True)
        url = f"{url}?{qs}"

    data = None
    if payload is not None:
        data = json.dumps(supabase_json_value(payload), ensure_ascii=False).encode("utf-8")

    # Zwykłe logowanie hasłem jest endpointem publicznego Auth API. Musi użyć
    # anon key, a nie service-role key (ten drugi służy wyłącznie serwerowi).
    api_key = SUPABASE_ANON_KEY if use_anon_key and SUPABASE_ANON_KEY else SUPABASE_SERVICE_ROLE_KEY
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("apikey", api_key)
    req.add_header("Authorization", f"Bearer {api_key}")
    req.add_header("Content-Type", "application/json")
    if prefer:
        req.add_header("Prefer", prefer)

    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            if not raw:
                return None
            ctype = (resp.headers.get("Content-Type") or "").lower()
            if "application/json" in ctype or raw[:1] in (b"[", b"{"):
                return json.loads(raw.decode("utf-8"))
            return raw.decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        # Nie chowaj użytecznej odpowiedzi Supabase za samym kodem HTTP 422.
        raw = exc.read().decode("utf-8", errors="replace")[:500]
        try:
            body = json.loads(raw)
            raw = str(body.get("msg") or body.get("message") or body.get("error") or raw)
        except Exception:
            pass
        raise RuntimeError(f"Supabase HTTP {exc.code}: {raw or 'brak szczegółów'}") from exc


def supabase_storage_ref(object_path: str, bucket: str | None = None) -> str:
    bucket = bucket or SUPABASE_STORAGE_BUCKET
    return f"supabase://{bucket}/{object_path.lstrip('/')}"


def parse_supabase_storage_ref(value: str) -> tuple[str, str] | None:
    raw = norm(value)
    if not raw.startswith("supabase://"):
        return None
    rest = raw[len("supabase://"):]
    if "/" not in rest:
        return None
    bucket, object_path = rest.split("/", 1)
    return bucket, object_path


def supabase_storage_object_url(bucket: str, object_path: str) -> str:
    quoted_path = urllib.parse.quote(object_path.lstrip("/"), safe="/")
    return f"{SUPABASE_URL}/storage/v1/object/{urllib.parse.quote(bucket, safe='')}/{quoted_path}"


def ensure_supabase_storage_bucket(bucket: str | None = None):
    bucket = bucket or SUPABASE_STORAGE_BUCKET
    if not supabase_enabled() or not bucket:
        return
    payload = json.dumps({"id": bucket, "name": bucket, "public": False}).encode("utf-8")
    req = urllib.request.Request(f"{SUPABASE_URL}/storage/v1/bucket", data=payload, method="POST")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", "application/json")
    try:
        urllib.request.urlopen(req, timeout=30).read()
    except urllib.error.HTTPError as e:
        if e.code not in (400, 409):
            raise


def supabase_storage_upload_file(local_path: str, object_path: str, bucket: str | None = None, content_type: str = "application/pdf") -> str:
    if not supabase_enabled():
        raise RuntimeError("Brak konfiguracji Supabase")
    bucket = bucket or SUPABASE_STORAGE_BUCKET
    ensure_supabase_storage_bucket(bucket)
    with open(local_path, "rb") as f:
        data = f.read()
    req = urllib.request.Request(supabase_storage_object_url(bucket, object_path), data=data, method="POST")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", content_type)
    req.add_header("x-upsert", "true")
    with urllib.request.urlopen(req, timeout=90) as resp:
        resp.read()
    return supabase_storage_ref(object_path, bucket)

def supabase_storage_upload_bytes(data: bytes, object_path: str, bucket: str | None = None, content_type: str = "application/octet-stream") -> str:
    """Upload used by the driver portal; the browser never receives Supabase keys."""
    if not supabase_enabled():
        raise RuntimeError("Brak konfiguracji Supabase")
    bucket = bucket or SUPABASE_STORAGE_BUCKET
    ensure_supabase_storage_bucket(bucket)
    req = urllib.request.Request(supabase_storage_object_url(bucket, object_path), data=data, method="POST")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    req.add_header("Content-Type", content_type or "application/octet-stream")
    req.add_header("x-upsert", "false")
    with urllib.request.urlopen(req, timeout=90) as resp:
        if resp.status >= 300:
            raise RuntimeError(f"Supabase Storage HTTP {resp.status}")
    return supabase_storage_ref(object_path, bucket)


def supabase_storage_download_bytes(storage_ref: str) -> tuple[bytes, str]:
    parsed = parse_supabase_storage_ref(storage_ref)
    if not parsed:
        raise RuntimeError("Nieprawidłowa ścieżka Supabase Storage")
    bucket, object_path = parsed
    req = urllib.request.Request(supabase_storage_object_url(bucket, object_path), method="GET")
    req.add_header("apikey", SUPABASE_SERVICE_ROLE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
    with urllib.request.urlopen(req, timeout=90) as resp:
        return resp.read(), os.path.basename(object_path)


def supabase_insert_row(table: str, row: dict):
    res = supabase_request(
        f"/rest/v1/{table}",
        method="POST",
        payload=[row],
        prefer="return=representation",
    )
    if isinstance(res, list):
        return res[0] if res else None
    return res


def supabase_update_rows(table: str, values: dict, filters: dict):
    params = {k: f"eq.{v}" for k, v in filters.items()}
    return supabase_request(
        f"/rest/v1/{table}",
        method="PATCH",
        params=params,
        payload=values,
        prefer="return=minimal",
    )


def supabase_delete_rows(table: str, filters: dict):
    params = {k: f"eq.{v}" for k, v in filters.items()}
    return supabase_request(
        f"/rest/v1/{table}",
        method="DELETE",
        params=params,
        prefer="return=minimal",
    )


def supabase_select_rows(table: str, order_by: str = "id", page_size: int = 1000, extra_params: dict | None = None):
    rows = []
    offset = 0
    while True:
        params = {"select": "*", "limit": page_size, "offset": offset}
        if order_by:
            params["order"] = f"{order_by}.asc"
        if extra_params:
            params.update(extra_params)
        chunk = supabase_request(f"/rest/v1/{table}", method="GET", params=params) or []
        if not isinstance(chunk, list):
            raise RuntimeError(f"NieprawidĹ‚owa odpowiedĹş Supabase dla tabeli {table}")
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        offset += page_size
    return rows


def local_client_search_rows(limit: int = 5000):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT customer_email, customer_name, query, product_sku, product_model, product_name, results_count, source, created_at
      FROM client_search_logs
      ORDER BY created_at DESC, id DESC
      LIMIT ?
    """, (limit,))
    rows = [dict(r) for r in cur.fetchall()]
    c.close()
    return rows


def supabase_client_search_rows(limit: int = 5000):
    if not supabase_enabled():
        return []
    rows = supabase_request(
        "/rest/v1/client_search_logs",
        method="GET",
        params={
            "select": "customer_email,customer_name,query,product_sku,product_model,product_name,results_count,source,created_at",
            "order": "created_at.desc",
            "limit": str(limit),
        },
        timeout=30,
    ) or []
    return rows if isinstance(rows, list) else []


def load_client_search_rows(limit: int = 5000):
    local_rows = local_client_search_rows(limit=limit)
    cloud_rows = []
    cloud_ok = False
    if supabase_enabled():
        try:
            cloud_rows = supabase_client_search_rows(limit=limit)
            cloud_ok = True
        except Exception:
            cloud_rows = []

    merged = []
    seen = set()
    for row in list(cloud_rows) + list(local_rows):
        cleaned = {
            "customer_email": norm((row or {}).get("customer_email")).lower(),
            "customer_name": norm((row or {}).get("customer_name")),
            "query": norm((row or {}).get("query")),
            "product_sku": norm((row or {}).get("product_sku")),
            "product_model": norm((row or {}).get("product_model")),
            "product_name": norm((row or {}).get("product_name")),
            "results_count": to_int((row or {}).get("results_count"), 0),
            "source": norm((row or {}).get("source")) or "stock",
            "created_at": norm((row or {}).get("created_at")),
        }
        if not cleaned["query"]:
            continue
        key = (
            cleaned["customer_email"],
            cleaned["customer_name"],
            cleaned["query"].lower(),
            cleaned["product_sku"].lower(),
            cleaned["product_model"].lower(),
            cleaned["product_name"].lower(),
            cleaned["results_count"],
            cleaned["source"],
            cleaned["created_at"],
        )
        if key in seen:
            continue
        seen.add(key)
        merged.append(cleaned)

    merged.sort(key=lambda r: r.get("created_at") or "", reverse=True)
    source_label = "Supabase + kopia lokalna" if cloud_ok else "Kopia lokalna"
    return merged[:limit], source_label


def save_client_search_log_local(row: dict):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      INSERT INTO client_search_logs(customer_email, customer_name, query, product_sku, product_model, product_name, results_count, source, created_at)
      VALUES(?,?,?,?,?,?,?,?,?)
    """, (
        row.get("customer_email", ""),
        row.get("customer_name", ""),
        row.get("query", ""),
        row.get("product_sku", ""),
        row.get("product_model", ""),
        row.get("product_name", ""),
        to_int(row.get("results_count"), 0),
        row.get("source", "stock"),
        row.get("created_at") or now_iso(),
    ))
    c.commit()
    c.close()


def save_client_search_log_supabase(row: dict) -> bool:
    if not supabase_enabled():
        return False
    payload = {
        "customer_email": row.get("customer_email", ""),
        "customer_name": row.get("customer_name", ""),
        "query": row.get("query", ""),
        "product_sku": row.get("product_sku", ""),
        "product_model": row.get("product_model", ""),
        "product_name": row.get("product_name", ""),
        "results_count": to_int(row.get("results_count"), 0),
        "source": row.get("source", "stock"),
        "created_at": row.get("created_at") or now_iso(),
    }
    supabase_insert_row("client_search_logs", payload)
    return True


def sqlite_table_columns(table: str):
    c = conn()
    cur = c.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = [r[1] for r in cur.fetchall()]
    c.close()
    return cols


def sqlite_upsert_rows(table: str, rows: list, conflict_col: str):
    if not rows:
        return 0

    table_cols = sqlite_table_columns(table)
    usable_cols = [c for c in table_cols if any(c in row for row in rows)]
    if not usable_cols:
        return 0

    placeholders = ",".join(["?"] * len(usable_cols))
    update_cols = [c for c in usable_cols if c != conflict_col]
    if update_cols:
        update_sql = ", ".join([f"{c}=excluded.{c}" for c in update_cols])
        sql = f"INSERT INTO {table}({','.join(usable_cols)}) VALUES({placeholders}) ON CONFLICT({conflict_col}) DO UPDATE SET {update_sql}"
    else:
        sql = f"INSERT INTO {table}({','.join(usable_cols)}) VALUES({placeholders}) ON CONFLICT({conflict_col}) DO NOTHING"

    c = conn()
    cur = c.cursor()
    cnt = 0
    for row in rows:
        values = [row.get(col) for col in usable_cols]
        cur.execute(sql, values)
        cnt += 1
    c.commit()
    c.close()
    return cnt


def sqlite_delete_missing_rows(table: str, conflict_col: str, remote_keys: list):
    c = conn()
    cur = c.cursor()
    if not remote_keys:
        cur.execute(f"DELETE FROM {table}")
        deleted = cur.rowcount if cur.rowcount is not None else 0
        c.commit()
        c.close()
        return deleted

    cur.execute(f"SELECT {conflict_col} FROM {table}")
    local_keys = [r[0] for r in cur.fetchall()]
    remote_set = {str(x) for x in remote_keys}
    to_delete = [x for x in local_keys if str(x) not in remote_set]
    deleted = 0
    if to_delete:
        for i in range(0, len(to_delete), 800):
            pack = to_delete[i:i+800]
            ph = ",".join(["?"] * len(pack))
            cur.execute(f"DELETE FROM {table} WHERE {conflict_col} IN ({ph})", tuple(pack))
            deleted += cur.rowcount if cur.rowcount is not None else 0
    c.commit()
    c.close()
    return deleted


def pull_shared_tables_from_supabase(force: bool = False):
    if not supabase_enabled():
        return {"ok": False, "error": "not_configured"}

    now_ts = time.time()
    with _supabase_sync_lock:
        last_started = float(_supabase_sync_state.get("last_pull_started_ts") or 0.0)
        if (not force) and (now_ts - last_started) < SUPABASE_MIN_PULL_INTERVAL_SEC:
            return {"ok": True, "status": "throttled"}
        _supabase_sync_state["last_pull_started_ts"] = now_ts

    result = {"ok": True, "tables": {}, "pulled_at": now_iso()}
    fetched = {}

    # 1) pobierz wszystko z Supabase
    for table, conflict_col in SUPABASE_PULL_TABLES:
        try:
            fetched[(table, conflict_col)] = supabase_select_rows(table, order_by=conflict_col)
        except Exception as e:
            result["ok"] = False
            result["tables"][table] = {"status": "error", "stage": "fetch", "error": str(e)}

    # 2) upsert do lokalnego SQLite
    for table, conflict_col in SUPABASE_PULL_TABLES:
        if (table, conflict_col) not in fetched:
            continue
        try:
            remote_rows = fetched[(table, conflict_col)]
            sqlite_upsert_rows(table, remote_rows, conflict_col)
            result["tables"].setdefault(table, {})["rows"] = len(remote_rows)
            result["tables"][table]["upsert"] = "ok"
        except Exception as e:
            result["ok"] = False
            result["tables"].setdefault(table, {})
            result["tables"][table].update({"status": "error", "stage": "upsert", "error": str(e)})

    # 3) usuĹ„ lokalne rekordy, ktĂłrych juĹĽ nie ma w Supabase
    for table, conflict_col in reversed(SUPABASE_PULL_TABLES):
        if (table, conflict_col) not in fetched:
            continue
        if table == "ksef_documents":
            result["tables"].setdefault(table, {})
            result["tables"][table]["deleted_local"] = 0
            if result["tables"][table].get("upsert") == "ok":
                result["tables"][table]["status"] = "ok"
            continue
        try:
            remote_rows = fetched[(table, conflict_col)]
            remote_keys = [row.get(conflict_col) for row in remote_rows if row.get(conflict_col) is not None]
            deleted = sqlite_delete_missing_rows(table, conflict_col, remote_keys)
            result["tables"].setdefault(table, {})
            result["tables"][table]["deleted_local"] = deleted
            if result["tables"][table].get("upsert") == "ok":
                result["tables"][table]["status"] = "ok"
        except Exception as e:
            result["ok"] = False
            result["tables"].setdefault(table, {})
            result["tables"][table].update({"status": "error", "stage": "cleanup", "error": str(e)})

    try:
        normalize_temp_order_numbers()
        link_orders_to_customers_by_email(sync_remote=True)
    except Exception:
        pass
    return result


def maybe_pull_shared_from_supabase(force: bool = False):
    try:
        if request.method == "GET":
            pull_shared_tables_from_supabase(force=force)
    except Exception:
        pass


def sync_local_rows_to_supabase(table: str, conflict_col: str, ids: list):
    ids = [x for x in ids if x is not None]
    if not ids or not supabase_enabled():
        return 0

    c = conn()
    cur = c.cursor()
    ph = ",".join(["?"] * len(ids))
    cur.execute(f"SELECT * FROM {table} WHERE {conflict_col} IN ({ph})", tuple(ids))
    rows = [dict(r) for r in cur.fetchall()]
    c.close()
    if rows:
        supabase_upsert_rows(table, rows, conflict_col)
    return len(rows)


def sync_order_to_supabase(order_id: int):
    sync_local_rows_to_supabase("orders", "id", [order_id])
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id FROM order_items WHERE order_id=?", (order_id,))
    item_ids = [int(r["id"]) for r in cur.fetchall()]
    c.close()
    if item_ids:
        sync_local_rows_to_supabase("order_items", "id", item_ids)


def remote_first_create_customer(name: str, address: str, phone: str, email: str, nip: str):
    created = supabase_insert_row("customers", {
        "name": name,
        "address": address,
        "phone": phone,
        "email": email,
        "nip": nip,
        "created_at": now_iso(),
    })
    if not created or "id" not in created:
        raise RuntimeError("Supabase nie zwrĂłciĹ‚ ID dla klienta")

    customer_id = int(created["id"])
    c = conn()
    cur = c.cursor()
    cur.execute(
        "INSERT INTO customers(id, name, address, phone, email, nip, created_at) VALUES (?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET name=excluded.name, address=excluded.address, phone=excluded.phone, email=excluded.email, nip=excluded.nip, created_at=excluded.created_at",
        (customer_id, name, address, phone, email, nip, created.get("created_at") or now_iso())
    )
    c.commit()
    c.close()
    return customer_id


def remote_first_create_order(customer_id, customer_name, customer_address, customer_phone, customer_email, note, items, idempotency_key=None, delivery_date="", delivery_time=""):
    created_at = now_iso()
    order_id = cloud_row_id()
    order_payload = {
        "id": order_id,
        "order_no": "TEMP",
        "customer_id": customer_id if customer_id else None,
        "customer_name": customer_name,
        "customer_address": customer_address,
        "customer_phone": customer_phone,
        "customer_email": customer_email,
        "delivery_date": delivery_date or None,
        "delivery_time": delivery_time or None,
        "status": "new",
        "note": note,
        "created_at": created_at,
    }
    if idempotency_key:
        order_payload["idempotency_key"] = idempotency_key
    created_order = supabase_insert_row("orders", order_payload)
    if not created_order:
        raise RuntimeError("Supabase nie zwrĂłciĹ‚ ID dla zamĂłwienia")

    order_no = make_order_no(order_id, created_at)
    supabase_update_rows("orders", {"order_no": order_no}, {"id": order_id})

    c = conn()
    try:
        cur = c.cursor()
        cur.execute(
            "INSERT INTO orders(id, order_no, customer_id, customer_name, customer_address, customer_phone, customer_email, delivery_date, delivery_time, status, note, created_at, idempotency_key) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET order_no=excluded.order_no, customer_id=excluded.customer_id, customer_name=excluded.customer_name, customer_address=excluded.customer_address, customer_phone=excluded.customer_phone, customer_email=excluded.customer_email, delivery_date=excluded.delivery_date, delivery_time=excluded.delivery_time, status=excluded.status, note=excluded.note, created_at=excluded.created_at, idempotency_key=excluded.idempotency_key",
            (order_id, order_no, customer_id if customer_id else None, customer_name, customer_address, customer_phone, customer_email, delivery_date or None, delivery_time or None, "new", note, created_at, idempotency_key)
        )

        for pid, qty in items:
            cur.execute("SELECT sku FROM products WHERE id=?", (pid,))
            p = cur.fetchone()
            if not p:
                # Formularz może już widzieć świeżo dodany produkt z Supabase,
                # zanim kopia lokalna Rendera zostanie odświeżona. W takim
                # przypadku pobieramy wyłącznie tę pozycję, zamiast odrzucać
                # prawidłowe zamówienie.
                remote_rows = supabase_request(
                    "/rest/v1/products",
                    method="GET",
                    params={"select": "*", "id": f"eq.{pid}", "limit": 1},
                ) or []
                remote_product = remote_rows[0] if isinstance(remote_rows, list) and remote_rows else None
                if not remote_product:
                    raise ValueError(f"Nie istnieje produkt ID {pid} w Supabase")
                cur.execute(
                    """INSERT INTO products(id,sku,model,ean,name,unit,unit_net_price,unit_gross_price,
                       unit_material_cost,unit_production_cost,unit_transport_cost,unit_other_cost,created_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                       ON CONFLICT(id) DO UPDATE SET sku=excluded.sku,model=excluded.model,ean=excluded.ean,
                       name=excluded.name,unit=excluded.unit,unit_net_price=excluded.unit_net_price,
                       unit_gross_price=excluded.unit_gross_price,unit_material_cost=excluded.unit_material_cost,
                       unit_production_cost=excluded.unit_production_cost,unit_transport_cost=excluded.unit_transport_cost,
                       unit_other_cost=excluded.unit_other_cost,created_at=excluded.created_at""",
                    (
                        int(remote_product["id"]), remote_product["sku"], remote_product.get("model"),
                        remote_product.get("ean"), remote_product.get("name"), remote_product.get("unit") or "m3",
                        to_float(remote_product.get("unit_net_price"), 0.0),
                        to_float(remote_product.get("unit_gross_price"), 0.0),
                        to_float(remote_product.get("unit_material_cost"), 0.0),
                        to_float(remote_product.get("unit_production_cost"), 0.0),
                        to_float(remote_product.get("unit_transport_cost"), 0.0),
                        to_float(remote_product.get("unit_other_cost"), 0.0),
                        remote_product.get("created_at") or now_iso(),
                    ),
                )
                cur.execute("SELECT sku FROM products WHERE id=?", (pid,))
                p = cur.fetchone()
            item_id = cloud_row_id()
            created_item = supabase_insert_row("order_items", {
                "id": item_id,
                "order_id": order_id,
                "product_id": pid,
                "sku": p["sku"],
                "qty": qty,
                "created_at": now_iso(),
            })
            if not created_item:
                raise RuntimeError("Supabase nie zwrócił ID dla pozycji zamówienia")
            cur.execute(
                "INSERT INTO order_items(id, order_id, product_id, sku, qty, created_at) VALUES (?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET order_id=excluded.order_id, product_id=excluded.product_id, sku=excluded.sku, qty=excluded.qty, created_at=excluded.created_at",
                (int(created_item.get("id") or item_id), order_id, pid, p["sku"], qty, created_item.get("created_at") or now_iso())
            )
        c.commit()
    except Exception:
        c.rollback()
        try:
            supabase_delete_rows("order_items", {"order_id": order_id})
            supabase_delete_rows("orders", {"id": order_id})
        except Exception as rollback_exc:
            app.logger.error("Niepełny rollback zamówienia order_id=%s: %s", order_id, rollback_exc)
        raise
    finally:
        c.close()
    try:
        normalize_temp_order_numbers()
    except Exception:
        pass
    return order_id


def get_stock(product_id):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT qty FROM stock WHERE product_id=?", (product_id,))
    r = cur.fetchone()
    c.close()
    return int(r["qty"]) if r else 0

def change_stock(product_id, delta):
    c = conn()
    cur = c.cursor()
    cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (product_id,))
    cur.execute("UPDATE stock SET qty = qty + ? WHERE product_id=?", (delta, product_id))
    c.commit()
    c.close()

def safe_filename(s):
    s = re.sub(r"[^a-zA-Z0-9_\-\.]+", "_", s)
    return s[:80] if s else "file"


def invoice_dir_for_customer(customer_name: str) -> str:
    root = os.path.join(DATA_DIR, "faktury")
    os.makedirs(root, exist_ok=True)
    customer_dir = os.path.join(root, safe_filename(customer_name or "klient"))
    os.makedirs(customer_dir, exist_ok=True)
    return customer_dir


def get_pdf_font_names():
    regular = "Helvetica"
    bold = "Helvetica-Bold"

    # Szukaj czcionek Unicode takĹĽe po wildcardach i lokalnym katalogu app/fonts.
    regular_candidates = [
        # Lokalne fonty aplikacji (najwyĹĽszy priorytet)
        ("AppFont-Regular", os.path.join(APP_DIR, "fonts", "regular.ttf")),

        # Linux
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        ("DejaVuSansCondensed", "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf"),
        ("LiberationSans", "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
        ("NotoSans", "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),

        # Windows
        ("Arial", r"C:\Windows\Fonts\arial.ttf"),
        ("Calibri", r"C:\Windows\Fonts\calibri.ttf"),
        ("Tahoma", r"C:\Windows\Fonts\tahoma.ttf"),

        # macOS
        ("ArialMT", "/System/Library/Fonts/Supplemental/Arial.ttf"),
        ("HelveticaNeue", "/System/Library/Fonts/Helvetica.ttc"),
    ]
    bold_candidates = [
        ("AppFont-Bold", os.path.join(APP_DIR, "fonts", "bold.ttf")),

        # Linux
        ("DejaVuSans-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("DejaVuSansCondensed-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed-Bold.ttf"),
        ("LiberationSans-Bold", "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
        ("NotoSans-Bold", "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),

        # Windows
        ("Arial-Bold", r"C:\Windows\Fonts\arialbd.ttf"),
        ("Calibri-Bold", r"C:\Windows\Fonts\calibrib.ttf"),
        ("Tahoma-Bold", r"C:\Windows\Fonts\tahomabd.ttf"),

        # macOS
        ("Arial-BoldMT", "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
    ]

    # Dodatkowe wildcardy gdy Ĺ›cieĹĽki systemowe rĂłĹĽniÄ… siÄ™ miÄ™dzy maszynami.
    for path in glob.glob('/usr/share/fonts/**/*DejaVuSans*.ttf', recursive=True)[:6]:
        regular_candidates.append((f"AutoReg-{safe_filename(os.path.basename(path))}", path))
    for path in glob.glob('/usr/share/fonts/**/*NotoSans*.ttf', recursive=True)[:6]:
        regular_candidates.append((f"AutoReg-{safe_filename(os.path.basename(path))}", path))
    for path in glob.glob('/usr/share/fonts/**/*LiberationSans*.ttf', recursive=True)[:6]:
        regular_candidates.append((f"AutoReg-{safe_filename(os.path.basename(path))}", path))

    for path in glob.glob('/usr/share/fonts/**/*Bold*.ttf', recursive=True)[:10]:
        bold_candidates.append((f"AutoBold-{safe_filename(os.path.basename(path))}", path))

    def register_first(candidates):
        for name, path in candidates:
            if not path or not os.path.exists(path):
                continue
            try:
                if name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue
        return None

    reg = register_first(regular_candidates)
    bld = register_first(bold_candidates)

    if reg:
        regular = reg
    if bld:
        bold = bld
    elif reg:
        bold = reg

    return regular, bold


def generate_sales_invoice(order_row, items):
    customer_dir = invoice_dir_for_customer(order_row["customer_name"])
    fname = f"FV_{safe_filename(canonical_order_no(order_row['id'] if 'id' in order_row.keys() else None, order_row['created_at'] if 'created_at' in order_row.keys() else '', order_row['order_no']))}.pdf"
    fpath = os.path.join(customer_dir, fname)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM company_profile WHERE id=1")
    company = cur.fetchone()
    cur.execute("SELECT model, net_price, gross_price FROM pricing")
    pricing_rows = cur.fetchall()
    cur.execute("SELECT sku, model, name FROM products")
    product_rows = cur.fetchall()
    c.close()

    pricing_map = {norm(r["model"]): r for r in pricing_rows}
    product_map = {norm(r["sku"]): r for r in product_rows}

    def pdf_txt(value) -> str:
        return fix_polish_mojibake(norm(value))

    def fit_pdf_text(value, font_name, font_size, max_width, suffix="...") -> str:
        text = pdf_txt(value)
        if pdfmetrics.stringWidth(text, font_name, font_size) <= max_width:
            return text
        while text and pdfmetrics.stringWidth(text + suffix, font_name, font_size) > max_width:
            text = text[:-1]
        return (text + suffix) if text else ""

    w = 210 * mm
    h = 297 * mm
    cpdf = canvas.Canvas(fpath, pagesize=(w, h))

    pdf_font, pdf_font_bold = get_pdf_font_names()

    y = h - 18 * mm
    cpdf.setFont(pdf_font_bold, 14)
    cpdf.drawString(15 * mm, y, f"Faktura sprzedaĹĽowa: {canonical_order_no(order_row['id'] if 'id' in order_row.keys() else None, order_row['created_at'] if 'created_at' in order_row.keys() else '', order_row['order_no'])}")

    y -= 8 * mm
    cpdf.setFont(pdf_font, 10)
    cpdf.drawString(15 * mm, y, f"Data: {order_row['created_at']}")

    y -= 9 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, y, "Sprzedawca:")
    y -= 6 * mm
    cpdf.setFont(pdf_font, 8.5)
    if company:
        cpdf.drawString(15 * mm, y, f"{company['company_name'] or '-'}")
        y -= 5 * mm
        for ln in (company["address"] or "-").splitlines():
            cpdf.drawString(15 * mm, y, ln)
            y -= 5 * mm
        cpdf.drawString(15 * mm, y, f"NIP: {company['nip'] or '-'}")
        y -= 5 * mm
        cpdf.drawString(15 * mm, y, f"Tel: {company['phone'] or '-'}  Email: {company['email'] or '-'}")
        y -= 5 * mm
        cpdf.drawString(15 * mm, y, f"Konto: {company['bank_account'] or '-'}")
    else:
        cpdf.drawString(15 * mm, y, "Brak danych firmy (uzupeĹ‚nij w zakĹ‚adce: Dane mojej firmy)")

    y -= 8 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, y, "Nabywca:")
    y -= 6 * mm
    cpdf.setFont(pdf_font, 9)
    cpdf.drawString(15 * mm, y, f"{order_row['customer_name'] or '-'}")
    y -= 5 * mm
    for ln in (order_row["customer_address"] or "-").splitlines():
        cpdf.drawString(15 * mm, y, ln)
        y -= 5 * mm
    cpdf.drawString(15 * mm, y, f"Tel: {order_row['customer_phone'] or '-'}  Email: {order_row['customer_email'] or '-'}")

    y -= 10 * mm
    cpdf.setFont(pdf_font_bold, 9)
    cpdf.drawString(15 * mm, y, "SKU")
    cpdf.drawString(45 * mm, y, "Model")
    cpdf.drawString(95 * mm, y, "Ilość [m³]")
    cpdf.drawString(112 * mm, y, "Netto/m³")
    cpdf.drawString(140 * mm, y, "Brutto/m³")
    cpdf.drawString(170 * mm, y, "WartoĹ›Ä‡ brutto")
    y -= 5 * mm

    total_net = 0.0
    total_gross = 0.0
    cpdf.setFont(pdf_font, 9)

    for it in items:
        model = norm(it["model"])
        pr = pricing_map.get(model)
        net = float(pr["net_price"]) if pr else 0.0
        gross = float(pr["gross_price"]) if pr else 0.0
        qty = to_float(it["qty"], 0.0)
        line_net = net * qty
        line_gross = gross * qty
        total_net += line_net
        total_gross += line_gross

        cpdf.drawString(15 * mm, y, it["sku"])
        cpdf.drawString(45 * mm, y, (model or "-")[:24])
        cpdf.drawRightString(108 * mm, y, f"{qty:g} m³")
        cpdf.drawRightString(136 * mm, y, f"{net:.2f}")
        cpdf.drawRightString(164 * mm, y, f"{gross:.2f}")
        cpdf.drawRightString(195 * mm, y, f"{line_gross:.2f}")
        y -= 5 * mm

        if y < 28 * mm:
            cpdf.showPage()
            y = h - 20 * mm
            cpdf.setFont(pdf_font, 9)

    y -= 6 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawRightString(195 * mm, y, f"Suma netto: {total_net:.2f} PLN")
    y -= 5 * mm
    cpdf.drawRightString(195 * mm, y, f"Suma brutto: {total_gross:.2f} PLN")

    y -= 8 * mm
    cpdf.setFont(pdf_font, 9)
    cpdf.drawString(15 * mm, y, "Ceny pobrane z zakĹ‚adki Cennik (model, netto, brutto).")

    cpdf.save()
    return fpath


def generate_order_invoice_pdf(order_row, items, meta):
    customer_dir = invoice_dir_for_customer(meta.get("buyer_name") or (order_row["customer_name"] if order_row and "customer_name" in order_row.keys() else "") or "Klient")
    fname = f"{safe_filename(meta['invoice_no'])}.pdf"
    fpath = os.path.join(customer_dir, fname)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM company_profile WHERE id=1")
    company = cur.fetchone()
    cur.execute("SELECT model, net_price, gross_price FROM pricing")
    pricing_rows = cur.fetchall()
    cur.execute("SELECT sku, model, name FROM products")
    product_rows = cur.fetchall()
    c.close()

    pricing_map = {norm(r["model"]): r for r in pricing_rows}
    product_map = {norm(r["sku"]): r for r in product_rows}

    def pdf_txt(value) -> str:
        return fix_polish_mojibake(norm(value))

    def fit_pdf_text(value, font_name, font_size, max_width, suffix="...") -> str:
        text = pdf_txt(value)
        if pdfmetrics.stringWidth(text, font_name, font_size) <= max_width:
            return text
        while text and pdfmetrics.stringWidth(text + suffix, font_name, font_size) > max_width:
            text = text[:-1]
        return (text + suffix) if text else ""

    def wrap_pdf_text(value, font_name, font_size, max_width, max_lines=None):
        text = pdf_txt(value)
        if not text:
            return []
        out = []
        for raw_line in str(text).replace("\r", "\n").split("\n"):
            words = raw_line.split()
            if not words:
                out.append("")
                continue
            line = ""
            for word in words:
                candidate = word if not line else f"{line} {word}"
                if pdfmetrics.stringWidth(candidate, font_name, font_size) <= max_width:
                    line = candidate
                    continue
                if line:
                    out.append(line)
                    line = word
                else:
                    out.append(fit_pdf_text(word, font_name, font_size, max_width))
                    line = ""
                if max_lines and len(out) >= max_lines:
                    out[-1] = fit_pdf_text(out[-1], font_name, font_size, max_width)
                    return out
            if line:
                out.append(line)
            if max_lines and len(out) >= max_lines:
                out = out[:max_lines]
                out[-1] = fit_pdf_text(out[-1], font_name, font_size, max_width)
                return out
        return out

    w = 210 * mm
    h = 297 * mm
    cpdf = canvas.Canvas(fpath, pagesize=(w, h))
    pdf_font, pdf_font_bold = get_pdf_font_names()

    header_y = h - 20 * mm
    cpdf.setFont(pdf_font_bold, 14)
    cpdf.drawString(15 * mm, header_y, f"Faktura VAT: {meta['invoice_no']}")

    y = h - 34 * mm
    logo = find_logo_path()
    if logo:
        try:
            logo_img = ImageReader(logo)
            img_w, img_h = logo_img.getSize()
            max_w = 60 * mm
            max_h = 24 * mm
            scale = min(max_w / float(img_w), max_h / float(img_h)) if img_w and img_h else 1.0
            draw_w = float(img_w) * scale
            draw_h = float(img_h) * scale
            draw_x = 195 * mm - draw_w
            draw_y = h - 10 * mm - draw_h
            cpdf.drawImage(logo_img, draw_x, draw_y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    y -= 7 * mm
    cpdf.setFont(pdf_font, 10)
    cpdf.drawString(15 * mm, y, f"Miejsce: {pdf_txt(meta.get('place') or '-')}")
    cpdf.drawString(85 * mm, y, f"Data wystawienia: {pdf_txt(meta['issue_date'])}")
    cpdf.drawString(150 * mm, y, f"Data sprzedaży: {pdf_txt(meta['sell_date'])}")

    y -= 7 * mm
    cpdf.drawString(15 * mm, y, f"Forma płatności: {pdf_txt(payment_type_pl(meta.get('payment_type')))}")
    cpdf.drawString(85 * mm, y, f"Termin płatności: {pdf_txt(meta.get('payment_to') or '-')}")

    y -= 10 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, y, "Sprzedawca")
    cpdf.drawString(110 * mm, y, "Nabywca")

    y -= 6 * mm
    cpdf.setFont(pdf_font, 9)
    seller_name = pdf_txt((company["company_name"] if company else "") or "-")
    seller_nip = pdf_txt((company["nip"] if company else "") or "-")
    seller_addr = pdf_txt((company["address"] if company else "") or "-")
    seller_phone = pdf_txt((company["phone"] if company else "") or "")
    seller_email = pdf_txt((company["email"] if company else "") or "")
    seller_bank = pdf_txt((company["bank_account"] if company else "") or "")

    buyer_name = pdf_txt(meta.get("buyer_name") or (order_row["customer_name"] if order_row and "customer_name" in order_row.keys() else "") or "-")
    buyer_tax_no = pdf_txt(meta.get("buyer_tax_no") or "-")
    buyer_street = pdf_txt(meta.get("buyer_street") or "-")
    buyer_post = pdf_txt(meta.get("buyer_post_code") or "")
    buyer_city = pdf_txt(meta.get("buyer_city") or "")
    buyer_country = pdf_txt(meta.get("buyer_country") or "PL")
    buyer_email = pdf_txt(meta.get("buyer_email") or "")
    buyer_phone = pdf_txt(meta.get("buyer_phone") or "")

    seller_lines = [seller_name, f"NIP: {seller_nip}", seller_addr]
    if seller_phone:
        seller_lines.append(f"tel: {seller_phone}")
    if seller_email:
        seller_lines.append(f"email: {seller_email}")
    if seller_bank:
        seller_lines.append(f"konto: {seller_bank}")

    buyer_lines = [buyer_name, f"NIP: {buyer_tax_no}", buyer_street, f"{buyer_post} {buyer_city}".strip(), buyer_country]
    if buyer_phone:
        buyer_lines.append(f"tel: {buyer_phone}")
    if buyer_email:
        buyer_lines.append(f"email: {buyer_email}")

    seller_x = 15 * mm
    buyer_x = 108 * mm
    seller_width = 84 * mm
    buyer_width = 87 * mm
    line_gap = 4.8 * mm
    seller_wrapped = []
    buyer_wrapped = []
    for line in seller_lines:
        seller_wrapped.extend(wrap_pdf_text(line, pdf_font, 8.7, seller_width, max_lines=2))
    for line in buyer_lines:
        buyer_wrapped.extend(wrap_pdf_text(line, pdf_font, 8.7, buyer_width, max_lines=2))

    max_len = max(len(seller_wrapped), len(buyer_wrapped))
    cpdf.setFont(pdf_font, 8.7)
    for i in range(max_len):
        if i < len(seller_wrapped):
            cpdf.drawString(seller_x, y, seller_wrapped[i])
        if i < len(buyer_wrapped):
            cpdf.drawString(buyer_x, y, buyer_wrapped[i])
        y -= line_gap

    y -= 4 * mm
    table_left = 12 * mm
    table_right = 198 * mm
    row_h = 12 * mm
    # L.p. | Nazwa/SKU | Ilo?? | Netto/szt | Brutto/szt | Wart. netto | VAT
    col_x = [12 * mm, 20 * mm, 100 * mm, 113 * mm, 136 * mm, 159 * mm, 182 * mm, 198 * mm]

    def cell_center(x1, x2):
        return (x1 + x2) / 2.0

    def cell_baseline(y_top, h_cell, font_name, font_size):
        asc = pdfmetrics.getAscent(font_name, font_size)
        desc = pdfmetrics.getDescent(font_name, font_size)
        text_h = asc - desc
        y_bottom = y_top - h_cell + 1
        return y_bottom + (h_cell - text_h) / 2.0 - desc

    cpdf.setFillColorRGB(0.96, 0.96, 0.96)
    cpdf.rect(table_left, y - row_h + 1, table_right - table_left, row_h, stroke=0, fill=1)
    cpdf.setFillColorRGB(0, 0, 0)
    header_font = 7.6
    cpdf.setFont(pdf_font_bold, header_font)
    header_y = cell_baseline(y, row_h, pdf_font_bold, header_font)
    cpdf.drawCentredString(cell_center(col_x[0], col_x[1]), header_y, "L.p.")
    cpdf.drawCentredString(cell_center(col_x[1], col_x[2]), header_y, "Nazwa/SKU")
    cpdf.drawCentredString(cell_center(col_x[2], col_x[3]), header_y, "Ilość [m³]")
    cpdf.drawCentredString(cell_center(col_x[3], col_x[4]), header_y, "Netto/m³")
    cpdf.drawCentredString(cell_center(col_x[4], col_x[5]), header_y, "Brutto/m³")
    cpdf.drawCentredString(cell_center(col_x[5], col_x[6]), header_y, "Wartość netto")
    cpdf.drawCentredString(cell_center(col_x[6], col_x[7]), header_y, "VAT")
    cpdf.line(table_left, y + 1, table_right, y + 1)
    cpdf.line(table_left, y - row_h + 1, table_right, y - row_h + 1)
    for cx in col_x:
        cpdf.line(cx, y + 1, cx, y - row_h + 1)
    y -= row_h

    total_net = 0.0
    total_net_dec = Decimal("0.00")
    discount_pct = max(0.0, to_float(meta.get("discount_percent"), 0.0))
    body_font = 8.2
    cpdf.setFont(pdf_font, body_font)

    lp = 1
    for it in items:
        sku = pdf_txt(it.get("sku"))
        product_row = product_map.get(norm(sku))
        model = pdf_txt(it.get("model") or (product_row["model"] if product_row else ""))
        name = pdf_txt(it.get("name") or (product_row["name"] if product_row else ""))
        common_name = name or model
        pr = pricing_map.get(model) or pricing_map.get(sku)
        net_dec = money_dec(pr["net_price"] if pr else it.get("net_price"))
        qty_dec = Decimal(str(to_float(it["qty"], 0.0)))
        qty = float(qty_dec)
        line_net_dec = (net_dec * qty_dec).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        if discount_pct > 0:
            line_net_dec = (line_net_dec * (Decimal("100.0") - Decimal(str(discount_pct))) / Decimal("100.0")).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        unit_gross_dec = gross_from_net_23(net_dec)

        net = money_float(net_dec)
        gross = money_float(unit_gross_dec)
        line_net = money_float(line_net_dec)

        total_net += line_net
        total_net_dec += line_net_dec

        text_y = cell_baseline(y, row_h, pdf_font, body_font)
        cpdf.drawCentredString(cell_center(col_x[0], col_x[1]), text_y, str(lp))
        name_left = col_x[1] + 1.5 * mm
        name_width = (col_x[2] - col_x[1]) - 3 * mm
        cpdf.setFont(pdf_font_bold, body_font)
        cpdf.drawString(name_left, y - 4.4 * mm, fit_pdf_text(sku or "-", pdf_font_bold, body_font, name_width))
        cpdf.setFont(pdf_font, body_font)
        if common_name:
            label = common_name if common_name.lower() == model.lower() else f"{common_name} / {model}".strip(" /")
            cpdf.drawString(name_left, y - 8.7 * mm, fit_pdf_text(label, pdf_font, body_font, name_width))
        cpdf.drawCentredString(cell_center(col_x[2], col_x[3]), text_y, f"{qty:g} m³")
        cpdf.drawRightString(col_x[4] - 1.5 * mm, text_y, f"{net:.2f}")
        cpdf.drawRightString(col_x[5] - 1.5 * mm, text_y, f"{gross:.2f}")
        cpdf.drawRightString(col_x[6] - 1.5 * mm, text_y, f"{line_net:.2f}")
        cpdf.drawCentredString(cell_center(col_x[6], col_x[7]), text_y, "23%")
        cpdf.line(table_left, y - row_h + 1, table_right, y - row_h + 1)
        for cx in col_x:
            cpdf.line(cx, y + 1, cx, y - row_h + 1)
        y -= row_h
        lp += 1
        if y < 26 * mm:
            cpdf.showPage()
            y = h - 20 * mm
            cpdf.setFont(pdf_font, body_font)

    total_net_dec = total_net_dec.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    total_tax_dec = vat23_from_net(total_net_dec)
    total_gross_dec = (total_net_dec + total_tax_dec).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
    total_net = money_float(total_net_dec)
    total_tax = money_float(total_tax_dec)
    total_gross = money_float(total_gross_dec)
    y -= 6 * mm
    cpdf.setFont(pdf_font_bold, 10)
    if discount_pct > 0:
        cpdf.drawRightString(198 * mm, y, f"Rabat: {discount_pct:.2f}%")
        y -= 5 * mm
    cpdf.drawRightString(198 * mm, y, f"Suma netto: {total_net:.2f} PLN")
    y -= 5 * mm
    cpdf.drawRightString(198 * mm, y, f"VAT 23%: {total_tax:.2f} PLN")
    y -= 5 * mm
    cpdf.drawRightString(198 * mm, y, f"Suma brutto: {total_gross:.2f} PLN")

    ksef_number = norm(meta.get("ksef_number") or "")
    if ksef_number:
        y -= 10 * mm
        if y < 22 * mm:
            cpdf.showPage()
            y = h - 20 * mm
        cpdf.setFont(pdf_font_bold, 9)
        cpdf.drawString(15 * mm, y, "KSeF")
        y -= 5 * mm
        cpdf.setFont(pdf_font, 8.5)
        cpdf.drawString(15 * mm, y, "Faktura została wystawiona i jest dostępna w Krajowym Systemie e-Faktur.")
        y -= 5 * mm
        cpdf.setFont(pdf_font_bold, 8.5)
        cpdf.drawString(15 * mm, y, f"Numer KSeF: {pdf_txt(ksef_number)}")

    cpdf.save()
    return fpath, round(total_net,2), round(total_gross,2)


def packing_list_pdf_path_for_invoice(invoice_pdf_path: str, invoice_no: str) -> str:
    if parse_supabase_storage_ref(invoice_pdf_path):
        invoice_pdf_path = ""
    base_dir = os.path.dirname(invoice_pdf_path) if invoice_pdf_path else os.path.join(DATA_DIR, "faktury")
    return os.path.join(base_dir, f"{safe_filename(invoice_no)}_lista_pakowania.pdf")


def invoice_storage_object_path(invoice_id: int, invoice_no: str, suffix: str = ".pdf") -> str:
    return f"invoices/{int(invoice_id)}/{safe_filename(invoice_no)}{suffix}"


def invoice_packing_storage_object_path(invoice_id: int, invoice_no: str) -> str:
    return invoice_storage_object_path(invoice_id, invoice_no, "_lista_pakowania.pdf")


def upload_invoice_pdfs_to_supabase(invoice_id: int, invoice_no: str, invoice_pdf_path: str, packing_pdf_path: str = "") -> str:
    if not supabase_enabled():
        return invoice_pdf_relpath(invoice_pdf_path)
    invoice_ref = supabase_storage_upload_file(
        invoice_pdf_path,
        invoice_storage_object_path(invoice_id, invoice_no),
        content_type="application/pdf",
    )
    if packing_pdf_path and os.path.exists(packing_pdf_path):
        try:
            supabase_storage_upload_file(
                packing_pdf_path,
                invoice_packing_storage_object_path(invoice_id, invoice_no),
                content_type="application/pdf",
            )
        except Exception:
            pass
    return invoice_ref


def generate_invoice_packing_list_pdf(order_row, items, meta, invoice_pdf_path: str = "") -> str:
    customer_dir = invoice_dir_for_customer(meta.get("buyer_name") or (order_row["customer_name"] if order_row and "customer_name" in order_row.keys() else "") or "Klient")
    fpath = packing_list_pdf_path_for_invoice(invoice_pdf_path or os.path.join(customer_dir, f"{safe_filename(meta['invoice_no'])}.pdf"), meta["invoice_no"])

    w = 210 * mm
    h = 297 * mm
    cpdf = canvas.Canvas(fpath, pagesize=(w, h))
    pdf_font, pdf_font_bold = get_pdf_font_names()

    y = h - 18 * mm
    cpdf.setFont(pdf_font_bold, 15)
    cpdf.drawString(15 * mm, y, "Lista pakowania")
    y -= 7 * mm
    cpdf.setFont(pdf_font, 10)
    cpdf.drawString(15 * mm, y, f"Do faktury: {meta.get('invoice_no') or '-'}")
    y -= 5 * mm
    if order_row:
        cpdf.drawString(15 * mm, y, f"Klient: {order_row['customer_name']}")
        y -= 5 * mm
    cpdf.drawString(15 * mm, y, f"Data: {meta.get('issue_date') or app_now().strftime('%Y-%m-%d')}")
    y -= 8 * mm

    def fit_text(text, max_width, font_name, font_size):
        text = norm(text or "")
        if not text:
            return "-"
        if cpdf.stringWidth(text, font_name, font_size) <= max_width:
            return text
        ell = "..."
        while text and cpdf.stringWidth(text + ell, font_name, font_size) > max_width:
            text = text[:-1]
        return (text + ell) if text else ell

    def strip_note_from_order_no(order_no, note):
        order_no = norm(order_no or "")
        note = norm(note or "")
        if note and order_no.lower().endswith((" " + note).lower()):
            return order_no[:-(len(note) + 1)].strip()
        return order_no

    cpdf.setFont(pdf_font_bold, 9)
    headers = [("Lp.", 15), ("Zamówienie", 27), ("Notatka", 62), ("SKU", 92), ("Model / nazwa", 122), ("Ilość", 184)]
    for label, x_mm in headers:
        cpdf.drawString(x_mm * mm, y, label)
    y -= 3 * mm
    cpdf.line(15 * mm, y, 198 * mm, y)
    y -= 5 * mm

    cpdf.setFont(pdf_font, 9)
    total_qty = 0
    for lp, it in enumerate(items, 1):
        qty = int(it.get("qty") or 0)
        if qty <= 0:
            continue
        total_qty += qty
        order_no = norm(it.get("source_order_no") or "")
        note = norm(it.get("source_order_note") or "")
        sku = norm(it.get("sku") or "")
        model_name = norm(it.get("model") or it.get("name") or "")
        order_no = strip_note_from_order_no(order_no, note)
        product_text = (sku + "  " + model_name).strip()

        if y < 22 * mm:
            cpdf.showPage()
            y = h - 18 * mm
            cpdf.setFont(pdf_font, 8.5)

        cpdf.drawString(15 * mm, y, str(lp))
        cpdf.drawString(27 * mm, y, fit_text(order_no, 32 * mm, pdf_font, 8.5))
        cpdf.drawString(62 * mm, y, fit_text(note if note else "-", 27 * mm, pdf_font, 8.5))
        cpdf.drawString(92 * mm, y, fit_text(product_text, 86 * mm, pdf_font, 8.5))
        cpdf.drawRightString(198 * mm, y, str(qty))
        y -= 5 * mm

    y -= 4 * mm
    cpdf.line(15 * mm, y, 198 * mm, y)
    y -= 6 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawRightString(198 * mm, y, f"Razem szt.: {total_qty}")

    cpdf.save()
    return fpath


def invoice_pdf_relpath(abs_path: str) -> str:
    try:
        return os.path.relpath(abs_path, DATA_DIR)
    except Exception:
        return abs_path

def invoice_pdf_abspath(rel_path: str) -> str:
    return os.path.join(DATA_DIR, rel_path)

def find_invoice_pdf_fallback(invoice_no: str) -> str:
    root = os.path.join(DATA_DIR, "faktury")
    target = f"{safe_filename(invoice_no or '')}.pdf"
    if not target or target == ".pdf" or not os.path.isdir(root):
        return ""
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn == target:
                return os.path.join(dirpath, fn)
    return ""

def invoice_pdf_exists(pdf_path: str, invoice_no: str = "") -> tuple[bool, str]:
    abs_path = ""
    raw_pdf = norm(pdf_path)
    if parse_supabase_storage_ref(raw_pdf):
        try:
            supabase_storage_download_bytes(raw_pdf)
            return True, raw_pdf
        except Exception:
            return False, raw_pdf
    if raw_pdf:
        abs_path = raw_pdf if os.path.isabs(raw_pdf) else invoice_pdf_abspath(raw_pdf)
    if abs_path and os.path.exists(abs_path):
        return True, abs_path
    fallback = find_invoice_pdf_fallback(invoice_no)
    if fallback and os.path.exists(fallback):
        return True, fallback
    return False, ""


def load_invoice_meta(invoice_id: int):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM invoice_meta WHERE invoice_id=?", (invoice_id,))
    row = cur.fetchone()
    c.close()
    return dict(row) if row else None

def upsert_invoice_meta(
    invoice_id: int,
    pdf_path: str = "",
    invoice_items_json: str = "",
    sent_to_client: int | None = None,
    seen_by_client: int | None = None,
    seen_at: str | None = None,
    payment_reminder: int | None = None,
    paid: int | None = None,
    paid_at: str | None = None
):
    current = load_invoice_meta(invoice_id) or {}
    if sent_to_client is None:
        sent_to_client = int(current.get("sent_to_client") or 0)
    if seen_by_client is None:
        seen_by_client = int(current.get("seen_by_client") or 0)
    if seen_at is None:
        seen_at = current.get("seen_at")
    if payment_reminder is None:
        payment_reminder = int(current.get("payment_reminder") or 0)
    if paid is None:
        paid = int(current.get("paid") or 0)
    if paid_at is None:
        paid_at = current.get("paid_at")

    c = conn()
    cur = c.cursor()
    cur.execute("""
      INSERT INTO invoice_meta(invoice_id, pdf_path, invoice_items_json, sent_to_client, seen_by_client, payment_reminder, paid, paid_at, seen_at, updated_at)
      VALUES(?,?,?,?,?,?,?,?,?,?)
      ON CONFLICT(invoice_id) DO UPDATE SET
        pdf_path=excluded.pdf_path,
        invoice_items_json=excluded.invoice_items_json,
        sent_to_client=excluded.sent_to_client,
        seen_by_client=excluded.seen_by_client,
        payment_reminder=excluded.payment_reminder,
        paid=excluded.paid,
        paid_at=excluded.paid_at,
        seen_at=excluded.seen_at,
        updated_at=excluded.updated_at
    """, (invoice_id, pdf_path, invoice_items_json, int(sent_to_client), int(seen_by_client), int(payment_reminder), int(paid), paid_at, seen_at, now_iso()))
    c.commit()
    c.close()


def sync_invoice_meta_to_supabase(invoice_id: int):
    if not supabase_enabled():
        return
    meta = load_invoice_meta(invoice_id)
    if not meta:
        return
    try:
        sync_local_rows_to_supabase("invoice_meta", "invoice_id", [invoice_id])
        return
    except Exception:
        pass

    # Fallback dla Supabase bez najnowszych kolumn payment_reminder/paid/paid_at.
    legacy = {
        "invoice_id": meta.get("invoice_id"),
        "pdf_path": meta.get("pdf_path") or "",
        "invoice_items_json": meta.get("invoice_items_json") or "",
        "sent_to_client": int(meta.get("sent_to_client") or 0),
        "seen_by_client": int(meta.get("seen_by_client") or 0),
        "seen_at": meta.get("seen_at"),
        "updated_at": meta.get("updated_at") or now_iso(),
    }
    supabase_upsert_rows("invoice_meta", [legacy], "invoice_id")

def prepare_invoice_items(order_items: list[dict], form):
    prepared = []
    for it in order_items:
        remaining_qty = to_float(it.get("remaining_qty") if it.get("remaining_qty") is not None else it.get("qty"), 0.0)
        qty = to_float(form.get(f"invoice_qty_{it['id']}"), 0.0)
        if qty <= 0:
            continue
        qty = min(qty, remaining_qty)
        if qty <= 0:
            continue
        row = dict(it)
        row["order_item_id"] = int(it.get("id") or 0)
        row["source_order_id"] = int(it.get("order_id") or it.get("source_order_id") or 0)
        row["source_order_no"] = it.get("source_order_no") or ""
        row["source_order_note"] = it.get("source_order_note") or ""
        row["ordered_qty"] = to_float(it.get("qty"), 0.0)
        row["invoiced_qty_before"] = to_float(it.get("invoiced_qty"), 0.0)
        row["qty"] = qty
        line_net = money_dec(row.get("net_price")) * Decimal(str(qty))
        line_net = line_net.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        line_vat = vat23_from_net(line_net)
        line_gross = (line_net + line_vat).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        row["gross_price"] = money_float(gross_from_net_23(row.get("net_price")))
        row["vat_rate"] = 23
        row["line_value_net"] = money_float(line_net)
        row["line_value_vat"] = money_float(line_vat)
        row["line_value_gross"] = money_float(line_gross)
        prepared.append(row)
    return prepared


def invoiced_qty_by_order_item_ids(order_item_ids: list[int]):
    ids = [int(x) for x in order_item_ids if x is not None]
    out = {x: 0 for x in ids}
    if not ids:
        return out

    c = conn()
    cur = c.cursor()
    ph = ",".join(["?"] * len(ids))
    cur.execute(f"""
      SELECT order_item_id, COALESCE(SUM(qty),0) AS qty
      FROM invoice_allocations
      WHERE order_item_id IN ({ph})
      GROUP BY order_item_id
    """, tuple(ids))
    for r in cur.fetchall():
        out[int(r["order_item_id"])] = int(r["qty"] or 0)
    c.close()
    return out


def replace_invoice_allocations(invoice_id: int, invoice_items: list[dict]):
    c = conn()
    cur = c.cursor()
    cur.execute("DELETE FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    allocation_ids = []
    for it in invoice_items:
        order_item_id = int(it.get("order_item_id") or it.get("id") or 0)
        order_id = int(it.get("source_order_id") or it.get("order_id") or 0)
        qty = int(it.get("qty") or 0)
        if order_item_id <= 0 or order_id <= 0 or qty <= 0:
            continue
        cur.execute("""
          INSERT INTO invoice_allocations(invoice_id, order_id, order_item_id, product_id, sku, qty, created_at)
          VALUES(?,?,?,?,?,?,?)
        """, (
            invoice_id,
            order_id,
            order_item_id,
            int(it.get("product_id") or 0) or None,
            it.get("sku") or "",
            qty,
            now_iso()
        ))
        allocation_ids.append(int(cur.lastrowid))
    c.commit()
    c.close()
    return allocation_ids


def order_fully_invoiced(cur, order_id: int) -> bool:
    cur.execute("SELECT id, qty FROM order_items WHERE order_id=?", (order_id,))
    rows = [dict(r) for r in cur.fetchall()]
    if not rows:
        return False
    item_ids = [int(r["id"]) for r in rows]
    ph = ",".join(["?"] * len(item_ids))
    cur.execute(f"""
      SELECT order_item_id, COALESCE(SUM(qty),0) AS qty
      FROM invoice_allocations
      WHERE order_item_id IN ({ph})
      GROUP BY order_item_id
    """, tuple(item_ids))
    done = {int(r["order_item_id"]): int(r["qty"] or 0) for r in cur.fetchall()}
    return all(int(row["qty"] or 0) > 0 and int(done.get(int(row["id"]), 0)) >= int(row["qty"] or 0) for row in rows)


def finalize_fully_invoiced_orders(order_ids: list[int]):
    touched = sorted({int(x) for x in order_ids if x})
    if not touched:
        return [], []

    c = conn()
    cur = c.cursor()
    completed_order_ids = []
    changed_product_ids = []

    for order_id in touched:
        if not order_fully_invoiced(cur, order_id):
            continue
        cur.execute("SELECT id, status, warehouse_issued FROM orders WHERE id=?", (order_id,))
        order_row = cur.fetchone()
        if not order_row:
            continue

        warehouse_issued = int(order_row["warehouse_issued"] or 0)
        if warehouse_issued == 0 and INVENTORY_AUTOMATION_ENABLED:
            cur.execute("SELECT product_id, qty FROM order_items WHERE order_id=?", (order_id,))
            for it in cur.fetchall():
                pid = int(it["product_id"])
                qty = int(it["qty"] or 0)
                cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
                cur.execute("UPDATE stock SET qty = qty - ? WHERE product_id=?", (qty, pid))
                changed_product_ids.append(pid)
        # FV kończy obieg handlowy; nie rozchoduje automatycznie materiałów produkcyjnych.
        warehouse_issued = 1

        if norm(order_row["status"]).lower() != "issued" or int(order_row["warehouse_issued"] or 0) != warehouse_issued:
            cur.execute("UPDATE orders SET status='issued', warehouse_issued=? WHERE id=?", (warehouse_issued, order_id))
            completed_order_ids.append(order_id)

    c.commit()
    c.close()

    if supabase_enabled():
        for order_id in completed_order_ids:
            try:
                supabase_update_rows("orders", {"status": "issued", "warehouse_issued": 1}, {"id": order_id})
            except Exception:
                pass
        if changed_product_ids:
            try:
                sync_local_rows_to_supabase("stock", "product_id", list(set(changed_product_ids)))
            except Exception:
                pass

    return completed_order_ids, list(set(changed_product_ids))


def reconcile_orders_after_invoice_change(order_ids: list[int]):
    touched = sorted({int(x) for x in order_ids if x})
    if not touched:
        return [], []

    c = conn()
    cur = c.cursor()
    changed_order_ids = []
    changed_product_ids = []

    for order_id in touched:
        cur.execute("SELECT id, status, warehouse_issued FROM orders WHERE id=?", (order_id,))
        order_row = cur.fetchone()
        if not order_row:
            continue

        fully = order_fully_invoiced(cur, order_id)
        warehouse_issued = int(order_row["warehouse_issued"] or 0)
        current_status = norm(order_row["status"]).lower()

        if fully and warehouse_issued == 0:
            cur.execute("UPDATE orders SET status='issued', warehouse_issued=1 WHERE id=?", (order_id,))
            changed_order_ids.append(order_id)

        elif not fully and warehouse_issued == 1:
            next_status = "confirmed" if current_status == "issued" else (current_status or "confirmed")
            cur.execute("UPDATE orders SET status=?, warehouse_issued=0 WHERE id=?", (next_status, order_id))
            changed_order_ids.append(order_id)

    c.commit()
    c.close()

    if supabase_enabled():
        if changed_order_ids:
            try:
                sync_local_rows_to_supabase("orders", "id", changed_order_ids)
            except Exception:
                pass
        if changed_product_ids:
            try:
                sync_local_rows_to_supabase("stock", "product_id", list(set(changed_product_ids)))
            except Exception:
                pass

    return changed_order_ids, list(set(changed_product_ids))


def invoice_edit_items(invoice_id: int, invoice_row: dict):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT order_id, order_item_id, qty FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    current_alloc_rows = [dict(r) for r in cur.fetchall()]
    current_qty_by_item = {int(r["order_item_id"]): int(r["qty"] or 0) for r in current_alloc_rows}
    allocated_order_ids = {int(r["order_id"]) for r in current_alloc_rows if int(r.get("order_id") or 0)}

    email = _email_key(invoice_row.get("buyer_email"))
    if not email and invoice_row.get("order_id"):
        cur.execute("SELECT customer_email FROM orders WHERE id=?", (invoice_row.get("order_id"),))
        rr = cur.fetchone()
        email = _email_key(rr["customer_email"]) if rr else ""

    order_ids = set(allocated_order_ids)
    if invoice_row.get("order_id"):
        order_ids.add(int(invoice_row.get("order_id")))
    if email:
        status_ph = ",".join(["?"] * len(CURRENT_ORDER_STATUSES))
        cur.execute(f"""
          SELECT id
          FROM orders
          WHERE LOWER(COALESCE(customer_email,'')) = ?
            AND (
              LOWER(COALESCE(status,'')) IN ({status_ph})
              OR id IN (SELECT order_id FROM invoice_allocations WHERE invoice_id=?)
            )
          ORDER BY created_at DESC, id DESC
        """, (email, *sorted(CURRENT_ORDER_STATUSES), invoice_id))
        order_ids.update(int(r["id"]) for r in cur.fetchall())

    if not order_ids:
        c.close()
        return []

    ids = sorted(order_ids)
    ph = ",".join(["?"] * len(ids))
    cur.execute(f"""
      SELECT oi.*, p.model, p.name,
             oo.order_no AS source_order_no,
             oo.created_at AS source_order_created_at,
             oo.note AS source_order_note,
             COALESCE(pr.net_price, 0) AS net_price,
             COALESCE(pr.gross_price, 0) AS gross_price
      FROM order_items oi
      JOIN orders oo ON oo.id=oi.order_id
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      WHERE oi.order_id IN ({ph})
      ORDER BY oo.created_at DESC, oo.id DESC, oi.id
    """, ids)
    items = [dict(r) for r in cur.fetchall()]
    c.close()

    invoiced_by_item = invoiced_qty_by_order_item_ids([int(it["id"]) for it in items])
    out = []
    for it in items:
        item_id = int(it["id"])
        current_qty = int(current_qty_by_item.get(item_id, 0))
        ordered_qty = int(it.get("qty") or 0)
        invoiced_total = int(invoiced_by_item.get(item_id, 0))
        invoiced_other = max(0, invoiced_total - current_qty)
        max_qty = max(0, ordered_qty - invoiced_other)
        row = dict(it)
        row["order_item_id"] = item_id
        row["source_order_id"] = int(it.get("order_id") or 0)
        row["source_order_no"] = order_display_no(
            row["source_order_id"],
            it.get("source_order_created_at"),
            it.get("source_order_no"),
            it.get("source_order_note") or ""
        )
        row["source_order_note"] = it.get("source_order_note") or ""
        row["ordered_qty"] = ordered_qty
        row["invoiced_other_qty"] = invoiced_other
        row["current_invoice_qty"] = current_qty
        row["remaining_qty"] = max_qty
        if max_qty > 0 or current_qty > 0:
            out.append(row)
    return out


def prepare_invoice_edit_items(edit_items: list[dict], form):
    prepared = []
    for it in edit_items:
        max_qty = to_float(it.get("remaining_qty"), 0.0)
        qty = to_float(form.get(f"invoice_qty_{it['id']}"), 0.0)
        qty = max(0, min(qty, max_qty))
        if qty <= 0:
            continue
        row = dict(it)
        row["order_item_id"] = int(it.get("id") or it.get("order_item_id") or 0)
        row["source_order_id"] = int(it.get("order_id") or it.get("source_order_id") or 0)
        row["ordered_qty"] = to_float(it.get("ordered_qty") or it.get("qty"), 0.0)
        row["invoiced_qty_before"] = to_float(it.get("invoiced_other_qty"), 0.0)
        row["qty"] = qty
        line_net = money_dec(row.get("net_price")) * Decimal(str(qty))
        line_net = line_net.quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        line_vat = vat23_from_net(line_net)
        line_gross = (line_net + line_vat).quantize(MONEY_Q, rounding=ROUND_HALF_UP)
        row["gross_price"] = money_float(gross_from_net_23(row.get("net_price")))
        row["vat_rate"] = 23
        row["line_value_net"] = money_float(line_net)
        row["line_value_vat"] = money_float(line_vat)
        row["line_value_gross"] = money_float(line_gross)
        prepared.append(row)
    return prepared


# =========================
# TEMPLATES (BASE as "file")
# =========================

CLIENT_API_PATHS = {
    "/api/client_search_log", "/api/client/orders",
    "/api/order_lookup", "/api/client_invoices", "/api/client_order_email",
}
_rate_lock = threading.Lock()
_rate_hits = {}


def _rate_limit(bucket: str, limit: int, window_seconds: int):
    now = time.time()
    key = (bucket, request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip())
    with _rate_lock:
        hits = [ts for ts in _rate_hits.get(key, []) if now - ts < window_seconds]
        if len(hits) >= limit:
            return False
        hits.append(now)
        _rate_hits[key] = hits
    return True


def _admin_password_ok(candidate: str) -> bool:
    if ADMIN_PASSWORD_HASH:
        try:
            return check_password_hash(ADMIN_PASSWORD_HASH, candidate)
        except Exception:
            return False
    return bool(ADMIN_PASSWORD) and hmac.compare_digest(ADMIN_PASSWORD, candidate)


def role_may_write(role: str, path: str) -> bool:
    """Role separation: every write is limited to the person's operational area."""
    role = (role or "").lower()
    if role == "admin":
        return True
    if path.startswith("/admin/"):
        return False
    if role == "manager":
        return not (path.startswith("/ksef") or path.startswith("/invoices"))
    if role == "accounting":
        return (path.startswith("/invoices") or path.startswith("/ksef") or path.startswith("/cash-flow")
                or path.startswith("/customers") or path.startswith("/company") or path.startswith("/pricing")
                or (path.startswith("/orders/") and "/invoice" in path))
    if role == "warehouse":
        return path.startswith("/beton/wz") or path.startswith("/beton/transports") or path.startswith("/dispatch") or path.startswith("/operations")
    if role == "office":
        return ((path == "/orders/new" or path.startswith("/orders/"))
                and "/invoice" not in path and "/status" not in path) or path.startswith("/customers")
    return False


@app.before_request
def security_gate():
    path = request.path
    if path == "/healthz":
        return None
    if path == "/login":
        if request.method == "POST" and not _rate_limit("admin_login", 8, 15 * 60):
            return "Zbyt wiele prób logowania. Spróbuj później.", 429
        return None

    is_driver_api = path.startswith("/api/driver/")
    is_legacy_client_api = path in CLIENT_API_PATHS or path.startswith("/api/invoices/")
    if is_legacy_client_api and not LEGACY_CLIENT_API_ENABLED:
        return jsonify(ok=False, error="Panel klienta jest wyłączony"), 404
    is_client_api = is_driver_api or is_legacy_client_api
    if is_client_api:
        if request.method == "OPTIONS":
            return None
        if path == "/api/driver/login":
            if not _rate_limit("driver_login", 8, 15 * 60):
                return jsonify(ok=False, error="Zbyt wiele prób logowania. Spróbuj później."), 429
            return None
        if path == "/api/client/orders" and not _rate_limit("client_orders", 12, 10 * 60):
            return jsonify(ok=False, error="Zbyt wiele prób złożenia zamówienia"), 429
        if not _rate_limit("client_api", 180, 60):
            return jsonify(ok=False, error="Zbyt wiele żądań"), 429
        user = _authenticated_client_user()
        if not user:
            return jsonify(ok=False, error="Brak autoryzacji"), 401
        g.client_user = user
        return None

    if path == "/logout":
        return None
    if not session.get("admin_authenticated"):
        if path.startswith("/api/"):
            return jsonify(ok=False, error="Brak autoryzacji administratora"), 401
        return redirect(url_for("login", next=request.full_path if request.query_string else path))

    # Role and account activity are read from the database on each request, so an
    # administrator's change takes effect immediately without waiting for logout.
    if session.get("user_id"):
        c = conn()
        current_user = c.execute("SELECT role,active FROM app_users WHERE id=? AND deleted_at IS NULL", (session["user_id"],)).fetchone()
        c.close()
        if not current_user or not int(current_user["active"]):
            session.clear()
            return redirect(url_for("login"))
        session["role"] = current_user["role"]

    if request.method in {"POST", "PUT", "PATCH", "DELETE"} and not role_may_write(session.get("role"), path):
        return "Brak uprawnienia dla tej roli.", 403

    if request.method in {"POST", "PUT", "PATCH", "DELETE"}:
        if request.is_json:
            origin = norm(request.headers.get("Origin")).rstrip("/")
            expected = request.host_url.rstrip("/")
            if origin and origin != expected:
                return jsonify(ok=False, error="Nieprawidłowe źródło żądania"), 403
        else:
            supplied = norm(request.form.get("csrf_token") or request.headers.get("X-CSRF-Token"))
            if not supplied or not hmac.compare_digest(supplied, session.get("csrf_token", "")):
                return "Nieprawidłowy token bezpieczeństwa formularza. Odśwież stronę.", 403


@app.get("/healthz")
def healthz():
    return jsonify(ok=True, service="beton-lagow")


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        username = norm(request.form.get("username"))
        c = conn()
        user = c.execute("""SELECT * FROM app_users WHERE username=? COLLATE NOCASE
                            AND active=1 AND deleted_at IS NULL""", (username,)).fetchone()
        if not app.secret_key:
            error = "Brak konfiguracji sesji na serwerze."
        elif user and check_password_hash(user["password_hash"], request.form.get("password") or ""):
            session.clear()
            session["admin_authenticated"] = True
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["display_name"] = user["display_name"]
            session["role"] = user["role"]
            session["csrf_token"] = secrets.token_urlsafe(32)
            session.permanent = True
            c.execute("UPDATE app_users SET last_login_at=?,updated_at=? WHERE id=?", (now_iso(), now_iso(), user["id"]))
            c.commit()
            c.close()
            # Konto i ostatnie logowanie są wspólne dla wszystkich lokalizacji.
            sync_local_rows_to_supabase("app_users", "id", [user["id"]])
            target = norm(request.args.get("next"))
            return redirect(target if target.startswith("/") and not target.startswith("//") else url_for("home"))
        else:
            error = "Nieprawidłowy login lub hasło."
        c.close()
    return render_template_string(r'''<!doctype html><html lang="pl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Logowanie — Beton Łagów</title><style>body{margin:0;font-family:Inter,Segoe UI,sans-serif;background:#f5f6fa;color:#17233c;display:grid;place-items:center;min-height:100vh}.box{width:min(420px,calc(100% - 28px));background:#fff;padding:30px;border-radius:24px;box-shadow:0 18px 55px rgba(20,35,65,.13)}.logo{display:block;max-width:210px;max-height:74px;object-fit:contain;margin:0 0 15px}.muted{color:#718096;font-size:13px;margin-bottom:22px}label{display:block;font-size:12px;font-weight:700;margin:12px 0 6px}input{width:100%;box-sizing:border-box;padding:12px;border:1px solid #dfe3ec;border-radius:13px;font:inherit}button{width:100%;margin-top:18px;padding:12px;border:0;border-radius:13px;background:#5577ee;color:#fff;font-weight:700}.error{background:#fff1f2;color:#b9384c;padding:10px;border-radius:12px;font-size:12px}</style></head><body><form class="box" method="post"><img class="logo" src="{{ url_for('brand_logo') }}" alt="Beton Łagów"><div class="muted">Zaloguj się jako administrator.</div>{% if error %}<div class="error">{{ error }}</div>{% endif %}<label>Login</label><input name="username" autocomplete="username" required><label>Hasło</label><input name="password" type="password" autocomplete="current-password" required><button type="submit">Zaloguj</button></form></body></html>''', error=error)


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.after_request
def security_headers_and_csrf(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(self), microphone=(), geolocation=()")
    response.headers.setdefault("Content-Security-Policy", "default-src 'self'; img-src 'self' data: blob:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline' https://unpkg.com https://cdn.jsdelivr.net; connect-src 'self' https://*.supabase.co https://api.resend.com")
    if session.get("admin_authenticated") and response.content_type and response.content_type.startswith("text/html"):
        body = response.get_data(as_text=True)
        token = session.get("csrf_token", "")
        if token:
            hidden = f'<input type="hidden" name="csrf_token" value="{token}">'
            body = re.sub(r'(<form\b[^>]*\bmethod=["\']post["\'][^>]*>)', r'\1' + hidden, body, flags=re.I)
            response.set_data(body)
    return response

BASE = r"""
<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ title or "Beton Łagów Orders" }}</title>
  <style>
    :root{--navy:#12213d;--navy2:#0b1730;--blue:#5577ee;--blue2:#3f63dc;--mint:#31b98b;--amber:#f5a524;--red:#e05263;--ink:#17233c;--muted:#718096;--bg:#f5f6fa;--line:#e7eaf2;--card:#fff;--radius:22px;--shadow:0 12px 35px rgba(31,45,78,.07)}
    *{box-sizing:border-box}html{background:var(--bg)}body{font-family:Inter,ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;margin:0;background:radial-gradient(circle at 85% -10%,#eaf0ff 0,transparent 28%),var(--bg);color:var(--ink);line-height:1.45}
    .top{position:fixed;inset:10px auto 10px 10px;width:238px;background:linear-gradient(165deg,var(--navy),var(--navy2));color:#fff;padding:22px 14px;border-radius:26px;display:flex;flex-direction:column;z-index:1100;box-shadow:0 24px 50px rgba(10,24,54,.22);overflow-y:auto}
    .brand{padding:4px 10px 20px}.brand img{display:block;width:100%;max-width:185px;max-height:72px;object-fit:contain;object-position:left center}
    .nav{display:flex!important;flex-direction:column;align-items:stretch!important;gap:5px!important;flex-wrap:nowrap!important;width:100%}.nav a,.nav-drop-btn{display:flex;align-items:center;color:#dce5f7;text-decoration:none;padding:11px 12px;border:0;border-radius:13px;background:transparent;font:inherit;font-size:14px;font-weight:600;cursor:pointer;transition:.18s ease}.nav a:hover,.nav-drop-btn:hover,.nav a.active{background:rgba(93,128,246,.24);color:#fff;transform:translateX(2px)}
    .nav a:before{width:25px;font-size:16px;opacity:.9}.nav a:nth-child(1):before{content:"⌂"}.nav a:nth-child(2):before{content:"▣"}.nav a:nth-child(3):before{content:"＋"}.nav a:nth-child(4):before{content:"▤"}.nav a:nth-child(5):before{content:"K"}.nav a:nth-child(6):before{content:"⌕"}.nav a:nth-child(7):before{content:"▦"}.nav a:nth-child(8):before{content:"◇"}.nav a:nth-child(9):before{content:"▧"}
    .nav-dropdown{position:relative;display:block}.nav-drop-btn{width:100%;text-align:left}.nav-drop-btn:before{content:"⚙";width:25px}.nav-dropdown-menu{display:none;margin:4px 0 2px 12px;border-left:1px solid rgba(255,255,255,.16);padding:2px 0 2px 8px}.nav-dropdown:hover .nav-dropdown-menu,.nav-dropdown:focus-within .nav-dropdown-menu{display:grid;gap:2px}.nav-dropdown-menu a{font-size:13px;padding:8px 10px}.nav-dropdown-menu a:before{display:none}
    .top>.right{margin:auto 8px 0!important;padding-top:16px;border-top:1px solid rgba(255,255,255,.13);color:#91a1bd!important;font-size:10px;overflow-wrap:anywhere}
    .mobile-toggle{display:none}.wrap{max-width:1500px;margin:0 0 0 258px;padding:28px 28px 18px;min-height:100vh}
    .card{background:rgba(255,255,255,.94);border:1px solid rgba(226,230,239,.9);border-radius:var(--radius);padding:20px;box-shadow:var(--shadow);margin-bottom:16px;overflow-x:auto}.card:hover{border-color:#dce2f1}
    .row{display:grid;grid-template-columns:1fr 1fr;gap:16px}h1{font-size:26px;letter-spacing:-.7px;margin:0 0 16px}h2{font-size:17px;letter-spacing:-.2px;margin:0 0 13px}.muted{color:var(--muted);font-size:12px}
    .btn{display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:10px 14px;border:1px solid #dce1eb;border-radius:13px;background:#fff;color:var(--ink);font-weight:650;text-decoration:none;cursor:pointer;box-shadow:0 3px 10px rgba(30,44,75,.04);transition:.18s ease}.btn:hover{transform:translateY(-1px);border-color:#bfc9df;box-shadow:0 7px 16px rgba(30,44,75,.09)}.btn.primary{background:linear-gradient(135deg,var(--blue),var(--blue2));color:#fff;border-color:transparent}.btn.danger{background:#fff0f2;color:#b92d43;border-color:#ffd6dc}.btn.ok{background:#e9faf4;color:#14835f;border-color:#c6f0e2}
    input,select,textarea{width:100%;padding:11px 13px;border:1px solid #dfe3ec;border-radius:13px;background:#fbfcfe;color:var(--ink);font:inherit;font-size:14px;outline:none;transition:.18s}input:focus,select:focus,textarea:focus{border-color:#7892f3;background:#fff;box-shadow:0 0 0 4px rgba(85,119,238,.11)}textarea{min-height:90px}
    table{width:100%;border-collapse:separate;border-spacing:0;min-width:660px}th,td{border-bottom:1px solid #edf0f5;padding:12px 11px;text-align:left;vertical-align:middle}th{background:#f8f9fc;color:#64718a;font-size:11px;text-transform:uppercase;letter-spacing:.45px;font-weight:750}thead th:first-child{border-radius:12px 0 0 12px}thead th:last-child{border-radius:0 12px 12px 0}tbody tr{transition:.15s}tbody tr:hover{background:#fafbff}
    .badge{display:inline-block;padding:5px 10px;border-radius:999px;border:1px solid #dfe4ef;background:#f8faff;color:#526079;font-size:11px;font-weight:700}.st-confirmed,.badge-paid{background:#e8f9f3!important;color:#16835f!important;border-color:#c9efe2!important}.st-unconfirmed{background:#fff1f2!important;color:#be3b50!important;border-color:#ffd7dc!important}.st-delivery{background:#edf3ff!important;color:#4166d3!important;border-color:#d9e4ff!important}.st-issued{background:#f0f2f6!important;color:#667085!important}
    .flex{display:flex;gap:10px;flex-wrap:wrap;align-items:center}.right{margin-left:auto}.small{font-size:12px}.grid3{display:grid;grid-template-columns:2fr 1fr 1fr;gap:16px}.line{height:1px;background:#edf0f5;margin:16px 0}.hint{background:#fff9e9;border:1px solid #f8e6ae;padding:12px 14px;border-radius:14px;color:#7e641b;font-size:13px}.kpi{display:flex;gap:10px;flex-wrap:wrap}.kpi .pill{background:#f4f7ff;border:1px solid #e1e8fb;padding:8px 11px;border-radius:999px;color:#516582;font-size:12px}.items-row{display:grid;grid-template-columns:2fr 120px 120px 120px;gap:10px;align-items:center}
    @media(max-width:980px){.top{transform:translateX(-110%);transition:.25s}.top.open{transform:translateX(0)}.mobile-toggle{display:grid;place-items:center;position:fixed;right:14px;bottom:14px;z-index:1200;width:52px;height:52px;border:0;border-radius:17px;background:var(--navy);color:#fff;font-size:22px;box-shadow:0 12px 28px rgba(12,28,58,.28)}.wrap{margin-left:0;padding:18px 14px 80px}.row,.grid3{grid-template-columns:1fr}.items-row{grid-template-columns:1fr 1fr}}
    @media(max-width:560px){.card{padding:15px;border-radius:18px}.items-row{grid-template-columns:1fr}.flex>.right{margin-left:0}h1{font-size:22px}}
  </style>
</head>
<body>
  <button class="mobile-toggle" type="button" onclick="document.querySelector('.top').classList.toggle('open')">☰</button>
  <div class="top">
    <div class="brand"><img src="{{ url_for('brand_logo') }}" alt="Beton Łagów"></div>
    <div class="nav flex">
      <a class="{% if request.endpoint == 'home' %}active{% endif %}" href="{{ url_for('home') }}">Pulpit</a>
      <a class="{% if request.endpoint in ['orders','order_view'] %}active{% endif %}" href="{{ url_for('orders') }}">Zamówienia</a>
      <a class="{% if request.endpoint == 'order_new' %}active{% endif %}" href="{{ url_for('order_new') }}">Nowe zamówienie</a>
      <a href="{{ url_for('invoices') }}">Faktury</a>
      <a href="{{ url_for('beton.wz_list') }}">Dokumenty WZ</a>
      <a href="{{ url_for('beton.transports') }}">Transporty</a>
      <a href="{{ url_for('dispatch.appointments') }}">Wydaj transport</a>
      <a href="{{ url_for('beton.drivers') }}">Kierowcy i pojazdy</a>
      <a href="{{ url_for('ops.operations') }}">Koszty i zużycie</a>
      <a href="{{ url_for('ops.analytics') }}">Analizy i raporty</a>
      <a href="{{ url_for('ksef_dashboard') }}">KSeF</a>
      <a href="{{ url_for('material_orders') }}">Zamówienia materiałów</a>
      <div class="nav-dropdown">
        <button class="nav-drop-btn" type="button">Ustawienia ▾</button>
        <div class="nav-dropdown-menu">
          <a href="{{ url_for('products') }}">Produkty</a>
          <a href="{{ url_for('customers') }}">Klienci</a>
          <a href="{{ url_for('pricing') }}">Cennik</a>
          <a href="{{ url_for('company') }}">Dane mojej firmy</a>
          <a href="{{ url_for('cash_flow') }}">Cash flow</a>
          {% if session.get('role') == 'admin' %}<a href="{{ url_for('admin_users') }}">Użytkownicy</a>{% endif %}
          <a href="{{ url_for('admin_audit') }}">Dziennik zmian</a>
          {% if session.get('role') == 'admin' %}<a href="{{ url_for('admin_test_data') }}">Dane testowe</a>{% endif %}
          <a href="{{ url_for('email_test') }}">Test maili</a>
        </div>
      </div>
      <a href="{{ url_for('logout') }}">Wyloguj</a>
    </div>
    <div class="right muted">Magazyn główny<br>{{ base_url }}</div>
  </div>

  <div class="wrap">
    {% block content %}{% endblock %}
    <div class="muted small" style="margin:14px 2px;">Dane na dysku: <b>{{ db_path }}</b></div>
  </div>

<script>
async function refreshStock(productId, targetId){
  if(!productId){ document.getElementById(targetId).innerText = "-"; return; }
  const r = await fetch("/api/product/"+productId);
  const j = await r.json();
  document.getElementById(targetId).innerText = (j.stock ?? "-");
}

function addItemRow(){
  const tpl = document.getElementById("itemRowTpl");
  const container = document.getElementById("itemsContainer");
  const node = tpl.content.cloneNode(true);
  container.appendChild(node);
}

function removeRow(btn){
  const row = btn.closest(".items-row");
  if(row) row.remove();
}
</script>

</body>
</html>
"""

# loader: BASE dostÄ™pny jako "base.html"
app.jinja_loader = DictLoader({"base.html": BASE})
app.jinja_env.globals["canonical_order_no"] = canonical_order_no
app.jinja_env.globals["order_display_no"] = order_display_no
app.jinja_env.globals["order_status_label"] = order_status_label if "order_status_label" in globals() else None
app.jinja_env.globals["order_status_css"] = order_status_css if "order_status_css" in globals() else None


def _admin_only():
    if session.get("role") != "admin":
        abort(403)


@app.route("/admin/users", methods=["GET", "POST"])
def admin_users():
    _admin_only()
    error = ""
    roles = [("admin", "Administrator"), ("manager", "Kierownik"), ("accounting", "Księgowość"), ("office", "Biuro"), ("warehouse", "Magazyn")]
    if request.method == "POST":
        username = norm(request.form.get("username"))
        display_name = norm(request.form.get("display_name"))
        password = request.form.get("password") or ""
        role = norm(request.form.get("role")) or "office"
        if role not in {r[0] for r in roles}:
            error = "Nieprawidłowa rola."
        elif not username or not display_name:
            error = "Podaj login oraz imię i nazwisko."
        elif len(password) < 12:
            error = "Hasło musi mieć co najmniej 12 znaków."
        else:
            c = conn()
            try:
                stamp = now_iso()
                cur = c.execute("INSERT INTO app_users(username,display_name,password_hash,role,created_at,updated_at) VALUES(?,?,?,?,?,?)", (username, display_name, generate_password_hash(password), role, stamp, stamp))
                user_id = cur.lastrowid
                c.commit()
                sync_local_rows_to_supabase("app_users", "id", [user_id])
                return redirect(url_for("admin_users"))
            except sqlite3.IntegrityError:
                error = "Taki login już istnieje."
            finally:
                c.close()
    c = conn()
    users = c.execute("SELECT * FROM app_users WHERE deleted_at IS NULL ORDER BY active DESC, role, display_name").fetchall()
    c.close()
    return render_template_string('''{% extends "base.html" %}{% block content %}<h1>Użytkownicy i uprawnienia</h1><div class="card"><p class="muted">Tylko administrator główny tworzy konta, zmienia role i wyłącza dostęp.</p>{% if error %}<div class="notice">{{error}}</div>{% endif %}<h2>Dodaj konto</h2><form method="post" class="grid3"><div><label>Imię i nazwisko</label><input name="display_name" required></div><div><label>Login</label><input name="username" required></div><div><label>Hasło (min. 12)</label><input name="password" type="password" minlength="12" required></div><div><label>Rola</label><select name="role">{% for value,label in roles %}<option value="{{value}}">{{label}}</option>{% endfor %}</select></div><div style="align-self:end"><button class="btn primary">Utwórz konto</button></div></form></div><div class="card"><h2>Role i dostęp</h2><table><thead><tr><th>Osoba</th><th>Login</th><th>Rola</th><th>Ostatnie logowanie</th><th>Dostęp</th></tr></thead><tbody>{% for u in users %}<tr><td><b>{{u.display_name}}</b></td><td>{{u.username}}</td><td><form method="post" action="{{url_for('admin_user_role',user_id=u.id)}}"><select name="role">{% for value,label in roles %}<option value="{{value}}" {% if value==u.role %}selected{% endif %}>{{label}}</option>{% endfor %}</select><button class="btn">Zapisz rolę</button></form></td><td>{{u.last_login_at or '—'}}</td><td>{% if u.id == session.get('user_id') %}<span class="badge">Twoje konto</span>{% else %}<form method="post" action="{{url_for('admin_user_toggle',user_id=u.id)}}" style="display:inline"><button class="btn">{{'Wyłącz' if u.active else 'Włącz'}}</button></form><form method="post" action="{{url_for('admin_user_delete',user_id=u.id)}}" style="display:inline;margin-left:6px"><button class="btn" style="color:#b42318">Usuń konto</button></form>{% endif %}</td></tr>{% else %}<tr><td colspan="5">Brak kont.</td></tr>{% endfor %}</tbody></table></div>{% endblock %}''', users=users, roles=roles, error=error, base_url=BASE_URL, db_path=DB_PATH)
    error = ""
    if request.method == "POST":
        username = norm(request.form.get("username"))
        display_name = norm(request.form.get("display_name"))
        password = request.form.get("password") or ""
        role = norm(request.form.get("role")) or "admin"
        if role not in {"admin", "manager", "accounting", "office", "warehouse"}:
            error = "Nieprawidłowa rola."
        elif not username or not display_name:
            error = "Podaj login oraz imię i nazwisko."
        elif len(password) < 12:
            error = "Hasło musi mieć co najmniej 12 znaków."
        else:
            c = conn()
            try:
                stamp = now_iso()
                c.execute("""INSERT INTO app_users(username,display_name,password_hash,role,created_at,updated_at)
                             VALUES(?,?,?,?,?,?)""", (username, display_name, generate_password_hash(password), role, stamp, stamp))
                c.commit()
                return redirect(url_for("admin_users"))
            except sqlite3.IntegrityError:
                error = "Taki login już istnieje."
            finally:
                c.close()
    c = conn(); users = c.execute("SELECT * FROM app_users WHERE deleted_at IS NULL ORDER BY active DESC,display_name").fetchall(); c.close()
    return render_template_string('''{% extends "base.html" %}{% block content %}
    <h1>Użytkownicy i administratorzy</h1><div class="row"><div class="card"><h2>Dodaj konto</h2>
    {% if error %}<div class="hint">{{error}}</div>{% endif %}<form method="post"><label>Imię i nazwisko</label><input name="display_name" required>
    <label>Login</label><input name="username" required><label>Hasło (min. 12 znaków)</label><input name="password" type="password" minlength="12" required>
    <label>Rola</label><select name="role"><option value="admin">Administrator</option><option value="manager">Kierownik</option><option value="accounting">Księgowość</option><option value="office">Biuro</option><option value="warehouse">Magazyn</option></select>
    <button class="btn primary" style="margin-top:12px">Utwórz konto</button></form></div><div class="card"><h2>Aktywne konta</h2><table><thead><tr><th>Osoba</th><th>Login</th><th>Rola</th><th>Ostatnie logowanie</th><th></th></tr></thead><tbody>
    {% for u in users %}<tr><td><b>{{u.display_name}}</b></td><td>{{u.username}}</td><td>{{u.role}}</td><td>{{u.last_login_at or '—'}}</td><td>{% if u.id != session.get('user_id') %}<form method="post" action="{{url_for('admin_user_toggle',user_id=u.id)}}"><button class="btn">{{'Wyłącz' if u.active else 'Włącz'}}</button></form>{% else %}<span class="badge">Twoje konto</span>{% endif %}</td></tr>{% endfor %}</tbody></table></div></div>{% endblock %}''', users=users,error=error,base_url=BASE_URL,db_path=DB_PATH)


@app.post("/admin/users/<int:user_id>/toggle")
def admin_user_toggle(user_id):
    _admin_only()
    if user_id == session.get("user_id"):
        abort(400)
    c = conn(); user = c.execute("SELECT * FROM app_users WHERE id=? AND deleted_at IS NULL", (user_id,)).fetchone()
    if not user:
        c.close(); abort(404)
    if user["active"] and user["role"] == "admin" and c.execute("SELECT COUNT(*) FROM app_users WHERE role='admin' AND active=1 AND deleted_at IS NULL").fetchone()[0] <= 1:
        c.close(); return "Nie można wyłączyć ostatniego administratora.", 400
    c.execute("UPDATE app_users SET active=?,updated_at=? WHERE id=?", (0 if user["active"] else 1, now_iso(), user_id)); c.commit(); c.close()
    sync_local_rows_to_supabase("app_users", "id", [user_id])
    return redirect(url_for("admin_users"))


@app.post("/admin/users/<int:user_id>/delete")
def admin_user_delete(user_id):
    """Remove a user's access while preserving audit and document history."""
    _admin_only()
    if user_id == session.get("user_id"):
        return "Nie możesz usunąć własnego konta.", 400

    c = conn()
    user = c.execute("SELECT * FROM app_users WHERE id=? AND deleted_at IS NULL", (user_id,)).fetchone()
    if not user:
        c.close()
        abort(404)
    active_admins = c.execute(
        "SELECT COUNT(*) FROM app_users WHERE role='admin' AND active=1 AND deleted_at IS NULL"
    ).fetchone()[0]
    if user["active"] and user["role"] == "admin" and active_admins <= 1:
        c.close()
        return "Nie można usunąć ostatniego aktywnego administratora.", 400

    timestamp = now_iso()
    c.execute(
        "UPDATE app_users SET active=0, deleted_at=?, updated_at=? WHERE id=?",
        (timestamp, timestamp, user_id),
    )
    c.commit()
    c.close()
    sync_local_rows_to_supabase("app_users", "id", [user_id])
    return redirect(url_for("admin_users"))


@app.post("/admin/users/<int:user_id>/role")
def admin_user_role(user_id):
    _admin_only()
    role = norm(request.form.get("role"))
    allowed = {"admin", "manager", "accounting", "office", "warehouse"}
    if role not in allowed:
        return "Nieprawidłowa rola.", 400
    c = conn()
    user = c.execute("SELECT * FROM app_users WHERE id=? AND deleted_at IS NULL", (user_id,)).fetchone()
    if not user:
        c.close()
        abort(404)
    if user["role"] == "admin" and role != "admin" and c.execute("SELECT COUNT(*) FROM app_users WHERE role='admin' AND active=1 AND deleted_at IS NULL").fetchone()[0] <= 1:
        c.close()
        return "Nie można odebrać roli ostatniemu administratorowi.", 400
    c.execute("UPDATE app_users SET role=?,updated_at=? WHERE id=?", (role, now_iso(), user_id))
    c.commit()
    c.close()
    sync_local_rows_to_supabase("app_users", "id", [user_id])
    return redirect(url_for("admin_users"))


@app.get("/admin/audit")
def admin_audit():
    # Render's local SQLite is temporary. Always refresh the history from the
    # central database before displaying it, otherwise a redeploy shows an
    # empty journal despite the records existing in Supabase.
    maybe_pull_shared_from_supabase(force=True)
    actor_filter = norm(request.args.get("actor")); path_filter = norm(request.args.get("path"))
    sql = "SELECT * FROM audit_events WHERE 1=1"; params = []
    if actor_filter: sql += " AND (actor_username LIKE ? OR actor_display_name LIKE ?)"; params += [f"%{actor_filter}%", f"%{actor_filter}%"]
    if path_filter: sql += " AND path LIKE ?"; params.append(f"%{path_filter}%")
    sql += " ORDER BY id DESC LIMIT 500"
    c=conn(); rows=c.execute(sql,params).fetchall(); c.close()
    return render_template_string('''{% extends "base.html" %}{% block content %}<h1>Dziennik zmian</h1><div class="card"><form method="get" class="flex"><input name="actor" value="{{request.args.get('actor','')}}" placeholder="Osoba" style="max-width:260px"><input name="path" value="{{request.args.get('path','')}}" placeholder="Moduł / ścieżka" style="max-width:260px"><button class="btn">Filtruj</button></form></div><div class="card"><table><thead><tr><th>Czas</th><th>Kto</th><th>Akcja</th><th>Obiekt</th><th>Wynik</th><th>IP</th></tr></thead><tbody>{% for x in rows %}<tr><td>{{x.created_at}}</td><td><b>{{x.actor_display_name or x.actor_username}}</b><br><span class="muted">{{x.actor_role or x.actor_type}}</span></td><td>{{x.action}}<br><span class="muted">{{x.method}} {{x.path}}</span></td><td>{{x.entity_type or '—'}} {{x.entity_id or ''}}</td><td><span class="badge">HTTP {{x.response_status}}</span></td><td>{{x.ip_address or '—'}}</td></tr>{% endfor %}</tbody></table></div>{% endblock %}''',rows=rows,base_url=BASE_URL,db_path=DB_PATH)


@app.route("/admin/test-data", methods=["GET", "POST"])
def admin_test_data():
    _admin_only()
    c=conn(); owner_id=c.execute("SELECT MIN(id) FROM app_users WHERE deleted_at IS NULL").fetchone()[0]; c.close()
    if int(session.get("user_id") or 0) != int(owner_id or 0):
        abort(403)
    error=""; done=False
    if request.method == "POST":
        phrase=norm(request.form.get("confirmation")); password=request.form.get("password") or ""
        c=conn(); me=c.execute("SELECT password_hash FROM app_users WHERE id=?",(session.get("user_id"),)).fetchone(); c.close()
        if phrase != "WYCZYSC DANE TESTOWE":
            error="Przepisz dokładnie: WYCZYSC DANE TESTOWE"
        elif not me or not check_password_hash(me["password_hash"],password):
            error="Nieprawidłowe hasło głównego administratora."
        else:
            c=conn()
            try:
                c.execute("PRAGMA foreign_keys=OFF")
                for table in ("delivery_photos","transport_items","transports","wz_items","wz_documents","invoice_allocations","invoice_meta","ksef_documents","invoices","order_items","orders","material_usage","fuel_entries","vehicle_expenses","audit_log","audit_events","email_events","client_search_logs"):
                    c.execute(f"DELETE FROM {table}")
                c.execute("UPDATE stock SET qty=0")
                c.commit(); done=True
            finally:
                c.execute("PRAGMA foreign_keys=ON"); c.close()
    return render_template_string('''{% extends "base.html" %}{% block content %}<h1>Czyszczenie danych testowych</h1><div class="card"><div class="hint"><b>Tylko główny administrator.</b> Operacja usuwa WZ, transporty, zamówienia, faktury, koszty, zdjęcia i historię testową oraz zeruje magazyn. Zachowuje konta użytkowników, klientów, produkty, cennik, kierowców, pojazdy i dane firmy.</div>{% if done %}<p><span class="badge">Dane testowe zostały wyczyszczone</span></p>{% else %}{% if error %}<p class="hint">{{error}}</p>{% endif %}<form method="post"><label>Wpisz: WYCZYSC DANE TESTOWE</label><input name="confirmation" autocomplete="off" required><label>Twoje hasło</label><input name="password" type="password" required><button class="btn danger" style="margin-top:14px" onclick="return confirm('Czy na pewno usunąć dane testowe?')">Trwale wyczyść dane testowe</button></form>{% endif %}</div>{% endblock %}''',error=error,done=done,base_url=BASE_URL,db_path=DB_PATH)


# =========================
# PAGES
# =========================

@app.get("/")
def home():
    maybe_pull_shared_from_supabase()
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT COUNT(*) AS n FROM orders WHERE status IN ('new','packed','confirmed','in_delivery')")
    n_orders_current = cur.fetchone()["n"]
    cur.execute("SELECT COUNT(*) AS n FROM orders WHERE date(created_at)=date('now','localtime')")
    n_orders_today = int(cur.fetchone()["n"] or 0)
    cur.execute("""
      SELECT o.id,o.order_no,o.customer_name,o.created_at,o.status,
             COALESCE(SUM(oi.qty * COALESCE(pr.net_price,0)),0) AS total_net
      FROM orders o
      LEFT JOIN order_items oi ON oi.order_id=o.id
      LEFT JOIN products p ON p.id=oi.product_id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model))=TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model))=TRIM(LOWER(p.sku)))
      GROUP BY o.id ORDER BY o.id DESC LIMIT 8
    """)
    recent_orders = [dict(r) for r in cur.fetchall()]

    # Pulpit nie opiera się na dawnym ręcznym statusie zamówienia. Etap jest
    # wyliczany z faktycznie utworzonego transportu, podpisanego WZ i faktury.
    def dashboard_delivery_status(order_id):
        if cur.execute("SELECT 1 FROM invoices WHERE order_id=? LIMIT 1", (order_id,)).fetchone():
            return "FV wystawiona"
        rows = cur.execute("""SELECT t.status FROM transports t
            JOIN wz_documents w ON w.id=t.wz_id
            WHERE w.order_id=? AND w.deleted_at IS NULL AND t.deleted_at IS NULL""", (order_id,)).fetchall()
        transport_statuses = {norm(row["status"]).lower() for row in rows}
        if "in_transit" in transport_statuses:
            return "W dostawie"
        if "issued" in transport_statuses:
            return "Wydane"
        if "assigned" in transport_statuses:
            return "Przydzielone"
        if "delivered" in transport_statuses:
            return "WZ podpisane"
        if "closed" in transport_statuses:
            return "Na miejscu"
        if "returned" in transport_statuses:
            return "Zakończone"
        wz = cur.execute("""SELECT status FROM wz_documents
            WHERE order_id=? AND deleted_at IS NULL ORDER BY id DESC LIMIT 1""", (order_id,)).fetchone()
        wz_status = norm(wz["status"]).lower() if wz else ""
        if wz_status == "ready_invoice":
            return "WZ podpisane"
        if wz_status in ("issued", "in_transport"):
            return "Wydane"
        return "Nieprzydzielone"

    all_orders = [dict(row) for row in cur.execute("SELECT id,delivery_date FROM orders WHERE lower(COALESCE(status,'')) <> 'cancelled'").fetchall()]
    for order in recent_orders:
        order["delivery_status"] = dashboard_delivery_status(order["id"])
    dashboard_counts = {}
    for order in all_orders:
        label = dashboard_delivery_status(order["id"])
        dashboard_counts[label] = dashboard_counts.get(label, 0) + 1
    status_new = dashboard_counts.get("Nieprzydzielone", 0)
    status_assigned = dashboard_counts.get("Przydzielone", 0) + dashboard_counts.get("Wydane", 0)
    status_delivery = dashboard_counts.get("W dostawie", 0) + dashboard_counts.get("Na miejscu", 0)
    status_signed = dashboard_counts.get("WZ podpisane", 0)
    status_done = dashboard_counts.get("Zakończone", 0)
    status_invoice = dashboard_counts.get("FV wystawiona", 0)
    today_iso = datetime.now().date().isoformat()
    today_remaining = sum(
        1 for order in all_orders
        if norm(order.get("delivery_date")) == today_iso
        and dashboard_delivery_status(order["id"]) not in {"Zakończone", "FV wystawiona"}
    )
    status_total = sum(dashboard_counts.values())
    status_divisor = max(1, status_total)
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <style>
        .dashboard-head{display:flex;align-items:center;gap:14px;margin-bottom:18px}.dashboard-head h1{margin:0}.search-shell{margin-left:28px;flex:1;max-width:580px;position:relative}.search-shell input{padding-left:42px;background:#fff}.search-shell:before{content:"⌕";position:absolute;left:15px;top:9px;color:#8793aa;font-size:19px;z-index:2}.metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:16px;margin-bottom:16px}.metric{display:grid;grid-template-columns:55px 1fr;gap:14px;align-items:center;background:#fff;border:1px solid #e7eaf2;border-radius:22px;padding:18px;box-shadow:var(--shadow)}.metric .icon{display:grid;place-items:center;width:55px;height:55px;border-radius:17px;background:var(--soft,#edf3ff);color:var(--tone,#5577ee);font-size:23px}.metric span{color:#718096;font-size:12px;font-weight:650}.metric b{display:block;margin-top:2px;font-size:25px;letter-spacing:-.6px}.metric small{display:block;margin-top:4px;color:#2da176;font-size:10px}.dash-grid{display:grid;grid-template-columns:minmax(0,2.15fr) minmax(300px,.9fr);gap:16px;align-items:start}.panel-title{display:flex;align-items:center;gap:9px;margin-bottom:13px}.panel-title h2{margin:0}.panel-title .btn{margin-left:auto;padding:7px 11px;font-size:11px}.orders-card{padding-bottom:8px}.orders-card table{min-width:780px}.orders-card td{font-size:12px}.customer-name{font-weight:700}.order-no{color:#4166d3;font-weight:750;text-decoration:none}.side-stack{display:grid;gap:16px}.stock-list{display:grid;gap:2px}.stock-item{display:grid;grid-template-columns:42px 1fr auto;align-items:center;gap:10px;padding:10px 2px;border-bottom:1px solid #edf0f5}.stock-icon{display:grid;place-items:center;width:39px;height:39px;border-radius:12px;background:#f1f4f9;color:#68758d}.stock-name{font-size:12px;font-weight:700}.stock-sku{font-size:9px;color:#8b96a9}.stock-qty{font-size:12px;font-weight:800}.stock-qty:after{content:"";display:inline-block;width:7px;height:7px;margin-left:8px;border-radius:50%;background:#ee5262}.donut-wrap{display:grid;grid-template-columns:145px 1fr;align-items:center;gap:14px}.donut{width:140px;height:140px;border-radius:50%;display:grid;place-items:center;background:conic-gradient(#5577ee 0 calc(var(--p1)*1%),#65a7ec calc(var(--p1)*1%) calc((var(--p1) + var(--p2))*1%),#31b98b calc((var(--p1) + var(--p2))*1%) calc((var(--p1) + var(--p2) + var(--p3))*1%),#e05263 calc((var(--p1) + var(--p2) + var(--p3))*1%) 100%)}.donut:before{content:"";width:86px;height:86px;background:#fff;border-radius:50%;position:absolute}.donut-label{position:relative;text-align:center;font-size:11px;color:#77849b}.donut-label b{display:block;color:#17233c;font-size:25px}.legend{display:grid;gap:9px}.legend-row{display:grid;grid-template-columns:9px 1fr auto;gap:7px;align-items:center;font-size:10px}.legend-dot{width:8px;height:8px;border-radius:50%}.quick-card{grid-column:1}.quick-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:10px}.quick-grid .btn{min-height:80px;flex-direction:column;background:#f8faff;border-color:#e8ecf6;font-size:11px}.quick-grid .btn b{font-size:20px}.quick-grid .btn:nth-child(1){background:#edf3ff;color:#4166d3}.quick-grid .btn:nth-child(2){background:#eaf9f4;color:#16835f}.quick-grid .btn:nth-child(3){background:#eef3ff;color:#4b6bd3}.quick-grid .btn:nth-child(4){background:#fff5e5;color:#c57a10}.quick-grid .btn:nth-child(5){background:#f3edff;color:#7650ce}@media(max-width:1200px){.metrics{grid-template-columns:1fr 1fr}.dash-grid{grid-template-columns:1fr}.quick-card{grid-column:auto}}@media(max-width:760px){.dashboard-head{flex-wrap:wrap}.search-shell{order:3;margin-left:0;flex-basis:100%}.metrics{grid-template-columns:1fr}.quick-grid{grid-template-columns:1fr 1fr}.donut-wrap{grid-template-columns:1fr}.donut{margin:auto}}
      </style>

      <div class="dashboard-head">
        <div><h1>Pulpit</h1><div class="muted">Dzisiejszy obraz realizacji zamówień i dostaw.</div></div>
        <form class="search-shell" action="{{ url_for('orders') }}"><input name="q" placeholder="Szukaj zamówień, produktów, klientów..."></form>
        <a class="btn primary" href="{{ url_for('order_new') }}">＋ Nowe zamówienie</a>
      </div>

      <div class="metrics">
        <div class="metric"><div class="icon">▣</div><div><span>Nowe zamówienia</span><b>{{ n_orders_today }}</b><small>{{ n_orders_current }} aktualnie w toku</small></div></div>
        <div class="metric" style="--soft:#eaf9f4;--tone:#1aa176"><div class="icon">◇</div><div><span>W realizacji</span><b>{{ n_orders_current }}</b><small><a href="{{ url_for('orders', tab='today') }}">Pozostało do realizacji dzisiaj ({{ today_remaining }})</a></small></div></div>
      </div>

      <div class="dash-grid">
        <div class="card orders-card">
          <div class="panel-title"><span>▣</span><h2>Ostatnie zamówienia</h2><a class="btn" href="{{ url_for('orders') }}">Zobacz wszystkie</a></div>
          <table><thead><tr><th>Nr zamówienia</th><th>Klient</th><th>Data</th><th>Wartość</th><th>Status</th><th></th></tr></thead><tbody>
          {% for o in recent_orders %}<tr><td><a class="order-no" href="{{ url_for('order_view',order_id=o.id) }}">{{ canonical_order_no(o.id,o.created_at,o.order_no) }}</a></td><td class="customer-name">{{ o.customer_name or '-' }}</td><td>{{ o.created_at[:16] }}</td><td>{{ "%.2f"|format(o.total_net) }} zł</td><td><span class="badge">{{ o.delivery_status }}</span></td><td><a class="btn" href="{{ url_for('order_view',order_id=o.id) }}">•••</a></td></tr>{% endfor %}
          {% if not recent_orders %}<tr><td colspan="6" class="muted">Brak zamówień do wyświetlenia.</td></tr>{% endif %}
          </tbody></table>
        </div>
        <div class="side-stack">
          <div class="card"><div class="panel-title"><h2>Status realizacji</h2></div><div class="donut-wrap"><div class="donut" style="--p1:{{ status_new*100/status_divisor }};--p2:{{ status_assigned*100/status_divisor }};--p3:{{ (status_delivery + status_signed)*100/status_divisor }}"><div class="donut-label"><b>{{ status_total }}</b>łącznie</div></div><div class="legend">
            <div class="legend-row"><i class="legend-dot" style="background:#5577ee"></i><span>Nieprzydzielone</span><b>{{ status_new }}</b></div><div class="legend-row"><i class="legend-dot" style="background:#65a7ec"></i><span>Wydane / przydzielone</span><b>{{ status_assigned }}</b></div><div class="legend-row"><i class="legend-dot" style="background:#31b98b"></i><span>W dostawie / na miejscu</span><b>{{ status_delivery }}</b></div><div class="legend-row"><i class="legend-dot" style="background:#c784de"></i><span>WZ podpisane</span><b>{{ status_signed }}</b></div><div class="legend-row"><i class="legend-dot" style="background:#45a879"></i><span>Zakończone</span><b>{{ status_done }}</b></div><div class="legend-row"><i class="legend-dot" style="background:#e05263"></i><span>FV wystawiona</span><b>{{ status_invoice }}</b></div>
          </div></div></div>
        </div>
        <div class="card quick-card"><div class="panel-title"><span>ϟ</span><h2>Szybkie akcje</h2></div><div class="quick-grid">
          <a class="btn" href="{{ url_for('order_new') }}"><b>＋</b><span>Nowe zamówienie</span></a><a class="btn" href="{{ url_for('products') }}"><b>◇</b><span>Dodaj produkt</span></a><a class="btn" href="{{ url_for('material_orders') }}"><b>⇢</b><span>Zamów materiały</span></a><a class="btn" href="{{ url_for('invoices') }}"><b>▤</b><span>Faktury</span></a>
        </div>
      </div>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Start", base_url=BASE_URL, db_path=DB_PATH,
                                  n_orders_current=n_orders_current, n_orders_today=n_orders_today, today_remaining=today_remaining,
                                  recent_orders=recent_orders, status_new=status_new, status_assigned=status_assigned,
                                  status_delivery=status_delivery, status_signed=status_signed, status_done=status_done,
                                  status_invoice=status_invoice, status_total=status_total,
                                  status_divisor=status_divisor)


@app.get("/searches")
def client_searches():
    q = norm(request.args.get("q"))
    rows, source_label = load_client_search_rows(limit=5000)
    if q:
        needle = q.lower()
        rows = [
            r for r in rows
            if needle in (r.get("query") or "").lower()
            or needle in (r.get("customer_email") or "").lower()
            or needle in (r.get("customer_name") or "").lower()
        ]

    global_stats = {}
    model_stats = {}
    client_stats = {}
    phrase_events_seen = set()
    model_events_seen = set()
    for r in rows:
        query = norm(r.get("query"))
        if not query:
            continue
        email = norm(r.get("customer_email")).lower()
        name = norm(r.get("customer_name"))
        client_key = email or name or "anon"
        product_sku = norm(r.get("product_sku"))
        product_model = norm(r.get("product_model"))
        product_name = norm(r.get("product_name"))
        results_count = to_int(r.get("results_count"), 0)
        created_at = norm(r.get("created_at"))

        product_key = product_sku or product_model
        model_event_key = (email, name, query.lower(), product_key.lower(), created_at)
        if product_key and 0 < results_count <= 20 and model_event_key not in model_events_seen:
            model_events_seen.add(model_event_key)
            m = model_stats.setdefault(product_key, {
                "product_model": product_key,
                "product_sku": product_sku,
                "product_name": product_name or product_model,
                "searches_count": 0,
                "clients": set(),
                "phrases": set(),
                "last_at": "",
            })
            m["searches_count"] += 1
            m["clients"].add(client_key)
            if query:
                m["phrases"].add(query)
            if product_sku and not m.get("product_sku"):
                m["product_sku"] = product_sku
            if (product_name or product_model) and not m.get("product_name"):
                m["product_name"] = product_name or product_model
            if created_at > m["last_at"]:
                m["last_at"] = created_at

        phrase_event_key = (email, name, query.lower(), created_at)
        if phrase_event_key in phrase_events_seen:
            continue
        phrase_events_seen.add(phrase_event_key)

        g = global_stats.setdefault(query, {
            "query": query,
            "searches_count": 0,
            "clients": set(),
            "no_result_count": 0,
            "max_results": 0,
            "last_at": "",
        })
        g["searches_count"] += 1
        g["clients"].add(client_key)
        if results_count == 0:
            g["no_result_count"] += 1
        g["max_results"] = max(g["max_results"], results_count)
        if created_at > g["last_at"]:
            g["last_at"] = created_at

        client_label = name or email or "Nieznany klient"
        skey = (client_label, email, query)
        s = client_stats.setdefault(skey, {
            "client_label": client_label,
            "customer_email": email,
            "query": query,
            "searches_count": 0,
            "no_result_count": 0,
            "max_results": 0,
            "last_at": "",
        })
        s["searches_count"] += 1
        if results_count == 0:
            s["no_result_count"] += 1
        s["max_results"] = max(s["max_results"], results_count)
        if created_at > s["last_at"]:
            s["last_at"] = created_at

    model_rows = []
    for r in model_stats.values():
        item = dict(r)
        item["clients_count"] = len(item.pop("clients"))
        phrases = sorted(item.pop("phrases"))
        item["phrases_preview"] = ", ".join(phrases[:5])
        model_rows.append(item)
    model_rows.sort(key=lambda r: (r["searches_count"], r["last_at"]), reverse=True)
    model_rows = model_rows[:10]

    global_rows = []
    for r in global_stats.values():
        item = dict(r)
        item["clients_count"] = len(item.pop("clients"))
        global_rows.append(item)
    global_rows.sort(key=lambda r: (r["searches_count"], r["last_at"]), reverse=True)
    global_rows = global_rows[:10]

    summary_rows = list(client_stats.values())
    summary_rows.sort(key=lambda r: r["last_at"], reverse=True)
    summary_rows = summary_rows[:50]

    latest_rows = rows[:50]
    total_count = len(rows)

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Top wyszukiwania</h1>
          <span class="badge">Łącznie: {{ total_count }}</span>
          <span class="badge">{{ source_label }}</span>
        </div>
        <form method="get" class="grid3" style="margin-top:10px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj: klient / email / fraza">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('client_searches') }}">Wyczyść</a>
        </form>
      </div>

      <div class="card">
        <h2>TOP 10 modeli / SKU</h2>
        <div class="muted" style="margin-bottom:8px;">
          Najważniejsze produkty, które klienci realnie zobaczyli po wyszukaniu w panelu — także po nazwie zwyczajowej, rozstawie albo części SKU.
        </div>
        <table>
          <thead>
            <tr><th>Model / SKU</th><th>Nazwa</th><th>Ile razy</th><th>Klientów</th><th>Ostatnio</th></tr>
          </thead>
          <tbody>
            {% for r in model_rows %}
              <tr>
                <td><b>{{ r.product_model }}</b>{% if r.product_sku and r.product_sku != r.product_model %}<div class="muted">{{ r.product_sku }}</div>{% endif %}</td>
                <td>{{ r.product_name or '-' }}</td>
                <td><span class="badge">{{ r.searches_count }}</span></td>
                <td>{{ r.clients_count }}</td>
                <td class="muted">{{ r.last_at }}</td>
              </tr>
            {% endfor %}
            {% if not model_rows %}
              <tr><td colspan="5" class="muted">Brak zapisanych wyszukiwań.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <details class="card">
        <summary style="cursor:pointer;font-weight:700;font-size:16px;">Pokaż szczegóły: frazy, klienci i ostatnie wpisy</summary>

      <div style="margin-top:14px;">
        <h2>Frazy klientów</h2>
        <div class="muted" style="margin-bottom:8px;">
          Tu zostają wpisane teksty klienta. Pomaga sprawdzić, jak klienci szukają produktów i gdzie pojawiają się literówki albo brakujące nazwy.
        </div>
        <table>
          <thead>
            <tr><th>Fraza</th><th>Wyszukań</th><th>Klientów</th><th>Bez wyników</th><th>Najwięcej wyników</th><th>Ostatnio</th></tr>
          </thead>
          <tbody>
            {% for r in global_rows %}
              <tr>
                <td><b>{{ r.query }}</b></td>
                <td><span class="badge">{{ r.searches_count }}</span></td>
                <td>{{ r.clients_count }}</td>
                <td>{% if r.no_result_count %}<span class="badge">{{ r.no_result_count }}</span>{% else %}-{% endif %}</td>
                <td>{{ r.max_results }}</td>
                <td class="muted">{{ r.last_at }}</td>
              </tr>
            {% endfor %}
            {% if not global_rows %}
              <tr><td colspan="6" class="muted">Brak zapisanych fraz.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <div style="margin-top:18px;">
        <h2>Wyszukiwania według klienta</h2>
        <div class="muted" style="margin-bottom:8px;">Tu zobaczysz, kto konkretnie szukał danej frazy.</div>
        <table>
          <thead>
            <tr><th>Klient</th><th>Email</th><th>Fraza</th><th>Ile razy</th><th>Bez wyników</th><th>Ostatnio</th></tr>
          </thead>
          <tbody>
            {% for r in summary_rows %}
              <tr>
                <td><b>{{ r.client_label }}</b></td>
                <td>{{ r.customer_email or '-' }}</td>
                <td>{{ r.query }}</td>
                <td><span class="badge">{{ r.searches_count }}</span></td>
                <td>{% if r.no_result_count %}<span class="badge">{{ r.no_result_count }}</span>{% else %}-{% endif %}</td>
                <td class="muted">{{ r.last_at }}</td>
              </tr>
            {% endfor %}
            {% if not summary_rows %}
              <tr><td colspan="6" class="muted">Brak zapisanych wyszukiwań.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <div style="margin-top:18px;">
        <h2>Ostatnie wpisy</h2>
        <table>
          <thead>
            <tr><th>Czas</th><th>Klient</th><th>Email</th><th>Fraza</th><th>Model / SKU</th><th>Wyniki</th></tr>
          </thead>
          <tbody>
            {% for r in latest_rows %}
              <tr>
                <td class="muted">{{ r.created_at }}</td>
                <td>{{ r.customer_name or '-' }}</td>
                <td>{{ r.customer_email or '-' }}</td>
                <td><b>{{ r.query }}</b></td>
                <td>{{ r.product_model or r.product_sku or '-' }}</td>
                <td>{{ r.results_count }}</td>
              </tr>
            {% endfor %}
            {% if not latest_rows %}
              <tr><td colspan="6" class="muted">Brak wpisów.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
      </details>
    {% endblock %}
    """
    return render_template_string(tpl, title="Top wyszukiwania", base_url=BASE_URL, db_path=DB_PATH,
                                  model_rows=model_rows, global_rows=global_rows, summary_rows=summary_rows, latest_rows=latest_rows,
                                  total_count=total_count, q=q, source_label=source_label)


def client_searches_v2():
    q = norm(request.args.get("q"))
    rows, source_label = load_client_search_rows(limit=5000)

    customer_name_by_email = {}
    order_name_by_email = {}
    product_name_by_sku = {}
    known_product_names = {}
    try:
        c = conn()
        cur = c.cursor()
        cur.execute("SELECT email, name FROM customers WHERE TRIM(COALESCE(email,''))<>''")
        for rr in cur.fetchall():
            email_key = _email_key(rr["email"])
            company_name = norm(rr["name"])
            if email_key and company_name and not _order_name_is_fallback(company_name, email_key):
                customer_name_by_email[email_key] = company_name

        cur.execute("""
          SELECT customer_email, customer_name
          FROM orders
          WHERE TRIM(COALESCE(customer_email,''))<>''
          ORDER BY id DESC
        """)
        for rr in cur.fetchall():
            email_key = _email_key(rr["customer_email"])
            company_name = norm(rr["customer_name"])
            if email_key and company_name and email_key not in order_name_by_email and not _order_name_is_fallback(company_name, email_key):
                order_name_by_email[email_key] = company_name

        cur.execute("SELECT sku, name FROM products WHERE TRIM(COALESCE(name,''))<>''")
        for rr in cur.fetchall():
            sku = norm(rr["sku"])
            product_name = norm(rr["name"])
            if sku and product_name:
                product_name_by_sku[sku.lower()] = product_name
                known_product_names.setdefault(product_name.lower(), product_name)
        c.close()
    except Exception:
        try:
            c.close()
        except Exception:
            pass

    def display_customer_name(row):
        email = _email_key(row.get("customer_email"))
        raw_name = norm(row.get("customer_name"))
        for candidate in (customer_name_by_email.get(email), order_name_by_email.get(email), raw_name):
            candidate = norm(candidate)
            if candidate and not _order_name_is_fallback(candidate, email) and "@" not in candidate:
                return candidate
        return "-"

    def canonical_product_name(row):
        product_name = norm(row.get("product_name"))
        if product_name and product_name != "-":
            return product_name
        product_sku = norm(row.get("product_sku")).lower()
        product_model = norm(row.get("product_model")).lower()
        query_key = norm(row.get("query")).lower()
        if product_sku and product_sku in product_name_by_sku:
            return product_name_by_sku[product_sku]
        if product_model and product_model in product_name_by_sku:
            return product_name_by_sku[product_model]
        if query_key and query_key in known_product_names:
            return known_product_names[query_key]
        return ""

    for row in rows:
        row["_client_label"] = display_customer_name(row)
        row["_product_label"] = canonical_product_name(row)

    if q:
        needle = q.lower()
        rows = [
            r for r in rows
            if needle in (r.get("query") or "").lower()
            or needle in (r.get("customer_name") or "").lower()
            or needle in (r.get("_client_label") or "").lower()
            or needle in (r.get("_product_label") or "").lower()
            or needle in (r.get("product_sku") or "").lower()
            or needle in (r.get("product_model") or "").lower()
        ]

    phrase_stats = {}
    name_stats = {}
    client_stats = {}
    phrase_events_seen = set()
    name_events_seen = set()

    for r in rows:
        query = norm(r.get("query"))
        if not query:
            continue
        email = norm(r.get("customer_email")).lower()
        client_label = norm(r.get("_client_label"))
        client_key = email or client_label or "anon"
        product_name = norm(r.get("_product_label"))
        results_count = to_int(r.get("results_count"), 0)
        created_at = norm(r.get("created_at"))

        name_key = product_name.lower()
        name_event_key = (client_key, query.lower(), name_key, created_at)
        if name_key and results_count > 0 and name_event_key not in name_events_seen:
            name_events_seen.add(name_event_key)
            item = name_stats.setdefault(name_key, {
                "product_name": product_name,
                "searches_count": 0,
                "clients": set(),
                "last_at": "",
            })
            item["searches_count"] += 1
            item["clients"].add(client_key)
            if created_at > item["last_at"]:
                item["last_at"] = created_at

        phrase_event_key = (client_key, query.lower(), created_at)
        if phrase_event_key in phrase_events_seen:
            continue
        phrase_events_seen.add(phrase_event_key)

        phrase = phrase_stats.setdefault(query, {
            "query": query,
            "searches_count": 0,
            "clients": set(),
            "no_result_count": 0,
            "max_results": 0,
            "last_at": "",
        })
        phrase["searches_count"] += 1
        phrase["clients"].add(client_key)
        if results_count == 0:
            phrase["no_result_count"] += 1
        phrase["max_results"] = max(phrase["max_results"], results_count)
        if created_at > phrase["last_at"]:
            phrase["last_at"] = created_at

        summary_name = client_label if client_label and client_label != "-" else "Nieznany klient"
        skey = (summary_name, query)
        summary = client_stats.setdefault(skey, {
            "client_label": summary_name,
            "query": query,
            "searches_count": 0,
            "no_result_count": 0,
            "max_results": 0,
            "last_at": "",
        })
        summary["searches_count"] += 1
        if results_count == 0:
            summary["no_result_count"] += 1
        summary["max_results"] = max(summary["max_results"], results_count)
        if created_at > summary["last_at"]:
            summary["last_at"] = created_at

    name_rows = []
    for r in name_stats.values():
        item = dict(r)
        item["clients_count"] = len(item.pop("clients"))
        name_rows.append(item)
    name_rows.sort(key=lambda r: (r["searches_count"], r["last_at"]), reverse=True)
    name_rows = name_rows[:10]

    phrase_rows = []
    for r in phrase_stats.values():
        item = dict(r)
        item["clients_count"] = len(item.pop("clients"))
        phrase_rows.append(item)
    phrase_rows.sort(key=lambda r: (r["searches_count"], r["last_at"]), reverse=True)
    phrase_rows = phrase_rows[:10]

    summary_rows = list(client_stats.values())
    summary_rows.sort(key=lambda r: r["last_at"], reverse=True)
    summary_rows = summary_rows[:50]

    latest_rows = rows[:50]
    total_count = len(rows)

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Top wyszukiwania</h1>
          <span class="badge">Łącznie: {{ total_count }}</span>
          <span class="badge">{{ source_label }}</span>
        </div>
        <form method="get" class="grid3" style="margin-top:10px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj: klient / fraza / nazwa">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('client_searches') }}">Wyczyść</a>
        </form>
      </div>

      <div class="card">
        <h2>TOP 10 nazw zwyczajowych</h2>
        <div class="muted" style="margin-bottom:8px;">
          Ranking jest zsumowany po nazwie zwyczajowej, np. Winsor, Carl, Cerne — bez rozbijania na każdy rozmiar SKU.
        </div>
        <table>
          <thead>
            <tr><th>Nazwa zwyczajowa</th><th>Ile razy</th><th>Klientów</th><th>Ostatnio</th></tr>
          </thead>
          <tbody>
            {% for r in name_rows %}
              <tr>
                <td><b>{{ r.product_name or '-' }}</b></td>
                <td><span class="badge">{{ r.searches_count }}</span></td>
                <td>{{ r.clients_count }}</td>
                <td class="muted">{{ r.last_at }}</td>
              </tr>
            {% endfor %}
            {% if not name_rows %}
              <tr><td colspan="4" class="muted">Brak zapisanych wyszukiwań.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

      <details class="card">
        <summary style="cursor:pointer;font-weight:700;font-size:16px;">Pokaż szczegóły: frazy, klienci i ostatnie wpisy</summary>

        <div style="margin-top:14px;">
          <h2>Frazy klientów</h2>
          <div class="muted" style="margin-bottom:8px;">
            Tu zostają wpisane teksty klienta. Pomaga sprawdzić, jak klienci szukają produktów i gdzie pojawiają się literówki albo brakujące nazwy.
          </div>
          <table>
            <thead>
              <tr><th>Fraza</th><th>Wyszukań</th><th>Klientów</th><th>Bez wyników</th><th>Najwięcej wyników</th><th>Ostatnio</th></tr>
            </thead>
            <tbody>
              {% for r in phrase_rows %}
                <tr>
                  <td><b>{{ r.query }}</b></td>
                  <td><span class="badge">{{ r.searches_count }}</span></td>
                  <td>{{ r.clients_count }}</td>
                  <td>{% if r.no_result_count %}<span class="badge">{{ r.no_result_count }}</span>{% else %}-{% endif %}</td>
                  <td>{{ r.max_results }}</td>
                  <td class="muted">{{ r.last_at }}</td>
                </tr>
              {% endfor %}
              {% if not phrase_rows %}
                <tr><td colspan="6" class="muted">Brak zapisanych fraz.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>

        <div style="margin-top:18px;">
          <h2>Wyszukiwania według klienta</h2>
          <div class="muted" style="margin-bottom:8px;">Tu zobaczysz, która firma szukała danej frazy.</div>
          <table>
            <thead>
              <tr><th>Klient</th><th>Fraza</th><th>Ile razy</th><th>Bez wyników</th><th>Ostatnio</th></tr>
            </thead>
            <tbody>
              {% for r in summary_rows %}
                <tr>
                  <td><b>{{ r.client_label }}</b></td>
                  <td>{{ r.query }}</td>
                  <td><span class="badge">{{ r.searches_count }}</span></td>
                  <td>{% if r.no_result_count %}<span class="badge">{{ r.no_result_count }}</span>{% else %}-{% endif %}</td>
                  <td class="muted">{{ r.last_at }}</td>
                </tr>
              {% endfor %}
              {% if not summary_rows %}
                <tr><td colspan="5" class="muted">Brak zapisanych wyszukiwań.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>

        <div style="margin-top:18px;">
          <h2>Ostatnie wpisy</h2>
          <table>
            <thead>
              <tr><th>Czas</th><th>Klient</th><th>Fraza</th><th>Nazwa</th><th>Model / SKU</th><th>Wyniki</th></tr>
            </thead>
            <tbody>
              {% for r in latest_rows %}
                <tr>
                  <td class="muted">{{ r.created_at }}</td>
                  <td>{{ r._client_label or '-' }}</td>
                  <td><b>{{ r.query }}</b></td>
                  <td>{{ r._product_label or '-' }}</td>
                  <td>{{ r.product_model or r.product_sku or '-' }}</td>
                  <td>{{ r.results_count }}</td>
                </tr>
              {% endfor %}
              {% if not latest_rows %}
                <tr><td colspan="6" class="muted">Brak wpisów.</td></tr>
              {% endif %}
            </tbody>
          </table>
        </div>
      </details>
    {% endblock %}
    """
    return render_template_string(tpl, title="Top wyszukiwania", base_url=BASE_URL, db_path=DB_PATH,
                                  name_rows=name_rows, phrase_rows=phrase_rows, summary_rows=summary_rows, latest_rows=latest_rows,
                                  total_count=total_count, q=q, source_label=source_label)


app.view_functions["client_searches"] = client_searches_v2


register_cash_flow(app, {
    "conn": conn,
    "now_iso": now_iso,
    "app_now": app_now,
    "to_float": to_float,
    "maybe_pull_shared_from_supabase": maybe_pull_shared_from_supabase,
    "sync_local_rows_to_supabase": sync_local_rows_to_supabase,
    "BASE_URL": BASE_URL,
    "DB_PATH": DB_PATH,
})

register_beton_logistics(app, {
    "conn": conn,
    "now_iso": now_iso,
    "supabase_enabled": supabase_enabled,
    "supabase_request": supabase_request,
    "supabase_storage_upload_bytes": supabase_storage_upload_bytes,
    "supabase_storage_download_bytes": supabase_storage_download_bytes,
    "pull_shared_tables_from_supabase": pull_shared_tables_from_supabase,
    "sync_local_rows_to_supabase": sync_local_rows_to_supabase,
    "BASE_URL": BASE_URL,
    "DB_PATH": DB_PATH,
})

register_dispatch(app, {
    "conn": conn,
    "now_iso": now_iso,
    "sync_local_rows_to_supabase": sync_local_rows_to_supabase,
    "BASE_URL": BASE_URL,
    "DB_PATH": DB_PATH,
})

register_operations(app, conn, now_iso, maybe_pull_shared_from_supabase)


@app.after_request
def auto_sync_after_write(response):
    try:
        is_client_api = request.path in CLIENT_API_PATHS or request.path.startswith("/api/invoices/") or request.path.startswith("/api/driver/")
        if is_client_api:
            origin = norm(request.headers.get("Origin")).rstrip("/")
            if origin and origin in CLIENT_ALLOWED_ORIGINS:
                response.headers["Access-Control-Allow-Origin"] = origin
                response.headers["Vary"] = "Origin"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, Idempotency-Key"
        elif request.path.startswith("/api/"):
            response.headers["Access-Control-Allow-Origin"] = "*"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"

        no_auto_sync_paths = {"/api/client_search_log", "/api/client_order_email"}
        if response.status_code < 400 and request.method in ("POST", "PUT", "PATCH", "DELETE") and request.path not in no_auto_sync_paths:
            if SUPABASE_SYNC_BEFORE_RESPONSE and supabase_enabled():
                result = sync_all_to_supabase()
                response.headers["X-Supabase-Sync"] = "ok" if result.get("ok") else "failed"
                if not result.get("ok"):
                    app.logger.error("Niepełna synchronizacja Supabase po %s %s: %s", request.method, request.path, result.get("tables"))
            else:
                trigger_background_supabase_sync(reason=f"{request.method} {request.path}")
    except Exception:
        pass
    return response


@app.after_request
def write_audit_event(response):
    """Niezależny ślad każdej operacji zmieniającej dane, także nieudanej."""
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return response
    try:
        sensitive = ("password", "token", "secret", "authorization", "csrf")
        if request.is_json:
            raw = request.get_json(silent=True) or {}
        else:
            raw = request.form.to_dict(flat=False)
        payload = {}
        for key, value in raw.items():
            payload[key] = "[UKRYTO]" if any(word in key.lower() for word in sensitive) else value
        actor_type, actor_id = "anonymous", None
        username, display_name, role = "niezalogowany", None, None
        if session.get("admin_authenticated"):
            actor_type, actor_id = "staff", str(session.get("user_id") or "")
            username = session.get("username") or "pracownik"
            display_name, role = session.get("display_name"), session.get("role")
        elif getattr(g, "client_user", None):
            driver = g.client_user
            actor_type = "driver"
            actor_id = str(driver.get("id") or driver.get("user_id") or "")
            username = driver.get("email") or driver.get("name") or "kierowca"
            display_name, role = driver.get("name"), "driver"
        parts = [p for p in request.path.split("/") if p]
        entity_type = parts[0] if parts else "application"
        entity_id = next((p for p in reversed(parts) if p.isdigit()), None)
        event_id = cloud_row_id()
        c = conn()
        c.execute("""INSERT INTO audit_events(id,request_id,actor_type,actor_id,actor_username,actor_display_name,
                    actor_role,action,method,path,entity_type,entity_id,payload_json,response_status,ip_address,user_agent,created_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                    event_id, request.headers.get("X-Request-ID") or str(uuid.uuid4()), actor_type, actor_id, username,
                    display_name, role, request.endpoint or f"{request.method} {request.path}", request.method,
                    request.path, entity_type, entity_id, json.dumps(payload, ensure_ascii=False, default=str)[:12000],
                    response.status_code, (request.headers.get("X-Forwarded-For") or request.remote_addr or "").split(",")[0].strip(),
                    (request.user_agent.string or "")[:500], now_iso()))
        c.commit(); c.close()
        # This record must not wait for a later request: it is the evidence of
        # the operation that has just happened.
        if supabase_enabled():
            sync_local_rows_to_supabase("audit_events", "id", [event_id])
    except Exception:
        app.logger.exception("Nie udało się zapisać zdarzenia audytowego")
    return response


# -------------------------
# COMPANY
# -------------------------

@app.get("/company")
def company():
    maybe_pull_shared_from_supabase()
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM company_profile WHERE id=1")
    row = cur.fetchone()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <h1>Dane mojej firmy</h1>
        <div class="muted">Te dane trafiÄ… na fakturÄ™ sprzedaĹĽowÄ….</div>
      </div>

      <div class="card">
        <form method="post" action="{{ url_for('company_save') }}" class="row">
          <div><label class="muted small">Nazwa firmy</label><input name="company_name" value="{{ row['company_name'] if row else '' }}"></div>
          <div><label class="muted small">NIP</label><input name="nip" value="{{ row['nip'] if row else '' }}"></div>
          <div><label class="muted small">Telefon</label><input name="phone" value="{{ row['phone'] if row else '' }}"></div>
          <div><label class="muted small">Email</label><input name="email" value="{{ row['email'] if row else '' }}"></div>
          <div><label class="muted small">Konto bankowe</label><input name="bank_account" value="{{ row['bank_account'] if row else '' }}"></div>
          <div><label class="muted small">Adres</label><textarea name="address">{{ row['address'] if row else '' }}</textarea></div>
          <div class="flex" style="align-items:flex-end;"><button class="btn primary" type="submit">Zapisz dane firmy</button></div>
        </form>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Dane mojej firmy", base_url=BASE_URL, db_path=DB_PATH, row=row)

@app.post("/company/save")
def company_save():
    company_name = norm(request.form.get("company_name"))
    address = norm(request.form.get("address"))
    nip = norm(request.form.get("nip"))
    phone = norm(request.form.get("phone"))
    email = norm(request.form.get("email"))
    bank_account = norm(request.form.get("bank_account"))

    c = conn()
    cur = c.cursor()
    cur.execute("""
      INSERT INTO company_profile(id, company_name, address, nip, phone, email, bank_account, updated_at)
      VALUES(1,?,?,?,?,?,?,?)
      ON CONFLICT(id) DO UPDATE SET
        company_name=excluded.company_name,
        address=excluded.address,
        nip=excluded.nip,
        phone=excluded.phone,
        email=excluded.email,
        bank_account=excluded.bank_account,
        updated_at=excluded.updated_at
    """, (company_name, address, nip, phone, email, bank_account, now_iso()))
    c.commit()
    c.close()
    return redirect(url_for("company"))


# -------------------------
# PRICING
# -------------------------

@app.get("/pricing")
def pricing():
    maybe_pull_shared_from_supabase()
    q = norm(request.args.get("q"))
    c = conn()
    cur = c.cursor()
    if q:
        like = f"%{q}%"
        cur.execute("SELECT * FROM pricing WHERE model LIKE ? ORDER BY model LIMIT 2000", (like,))
    else:
        cur.execute("SELECT * FROM pricing ORDER BY model LIMIT 2000")
    rows = cur.fetchall()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <h1>Cennik</h1>
        <div class="muted">Import pliku cen (kolumny: model, netto, brutto). ObsĹ‚uga CSV i XLSX (jeĹ›li dostÄ™pny openpyxl).</div>
      </div>

      <div class="card">
        <h2>Import cennika</h2>
        <form method="post" action="{{ url_for('pricing_import') }}" enctype="multipart/form-data" class="row">
          <div>
            <input type="file" name="file" accept=".csv,.xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv" required>
          </div>
          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Importuj cennik</button>
          </div>
        </form>
      </div>

      <div class="card">
        <form method="get" class="grid3" style="margin-bottom:10px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj modelu">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('pricing') }}">WyczyĹ›Ä‡</a>
        </form>
        <h2>Pozycje cennika</h2>
        <table>
          <thead><tr><th>Model</th><th>Netto</th><th>Brutto</th></tr></thead>
          <tbody>
            {% for r in rows %}
              <tr>
                <td><b>{{ r['model'] }}</b></td>
                <td>{{ "%.2f"|format(r['net_price']) }}</td>
                <td>{{ "%.2f"|format(r['gross_price']) }}</td>
              </tr>
            {% endfor %}
            {% if not rows %}
              <tr><td colspan="3" class="muted">Brak pozycji cennika.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Cennik", base_url=BASE_URL, db_path=DB_PATH, rows=rows, q=q)

@app.post("/pricing/import")
def pricing_import():
    f = request.files.get("file")
    if not f:
        return "Brak pliku", 400

    filename = norm(f.filename).lower()
    parsed_rows = []

    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception:
            return "Brak biblioteki openpyxl do odczytu XLSX. UĹĽyj CSV albo doinstaluj openpyxl.", 400

        wb = load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return "Pusty plik", 400
        headers = [norm(x) for x in rows[0]]
        data = rows[1:]
        i_model = guess_col(headers, ["model"])
        i_sku = guess_col(headers, ["sku", "symbol", "index", "indeks", "kod", "code"])
        i_name = guess_col(headers, ["nazwa", "name", "produkt", "product"])
        i_ean = guess_col(headers, ["ean", "gtin"])
        i_net = guess_col(headers, ["netto", "net", "cena netto"])
        i_gross = guess_col(headers, ["brutto", "gross", "cena brutto"])
        if i_model is None or i_net is None or i_gross is None:
            return "Plik musi mieÄ‡ kolumny: model, netto, brutto", 400
        for r in data:
            if not r:
                continue
            model = norm(r[i_model]) if len(r) > i_model else ""
            if not model:
                continue
            sku = norm(r[i_sku]) if i_sku is not None and len(r) > i_sku else model
            name = norm(r[i_name]) if i_name is not None and len(r) > i_name else ""
            ean = norm(r[i_ean]) if i_ean is not None and len(r) > i_ean else ""
            net = to_float(r[i_net] if len(r) > i_net else "", 0.0)
            gross = to_float(r[i_gross] if len(r) > i_gross else "", 0.0)
            parsed_rows.append((sku, model, name, ean, net, gross))

    else:
        raw = f.read()
        try:
            text = raw.decode("utf-8-sig")
        except Exception:
            text = raw.decode("latin2", errors="replace")
        sample = text[:5000]
        delim = ";" if sample.count(";") >= sample.count(",") else ","
        rdr = csv.reader(io.StringIO(text), delimiter=delim)
        rows = list(rdr)
        if not rows:
            return "Pusty plik", 400
        headers = rows[0]
        data = rows[1:]
        i_model = guess_col(headers, ["model"])
        i_sku = guess_col(headers, ["sku", "symbol", "index", "indeks", "kod", "code"])
        i_name = guess_col(headers, ["nazwa", "name", "produkt", "product"])
        i_ean = guess_col(headers, ["ean", "gtin"])
        i_net = guess_col(headers, ["netto", "net", "cena netto"])
        i_gross = guess_col(headers, ["brutto", "gross", "cena brutto"])
        if i_model is None or i_net is None or i_gross is None:
            return "Plik musi mieÄ‡ kolumny: model, netto, brutto", 400
        for r in data:
            if not r:
                continue
            model = norm(r[i_model]) if len(r) > i_model else ""
            if not model:
                continue
            sku = norm(r[i_sku]) if i_sku is not None and len(r) > i_sku else model
            name = norm(r[i_name]) if i_name is not None and len(r) > i_name else ""
            ean = norm(r[i_ean]) if i_ean is not None and len(r) > i_ean else ""
            net = to_float(r[i_net] if len(r) > i_net else "", 0.0)
            gross = to_float(r[i_gross] if len(r) > i_gross else "", 0.0)
            parsed_rows.append((sku, model, name, ean, net, gross))

    c = conn()
    cur = c.cursor()
    changed_product_ids = []
    for sku, model, name, ean, net, gross in parsed_rows:
        cur.execute("""
          INSERT INTO pricing(model, net_price, gross_price, created_at)
          VALUES(?,?,?,?)
          ON CONFLICT(model) DO UPDATE SET
            net_price=excluded.net_price,
            gross_price=excluded.gross_price,
            created_at=excluded.created_at
        """, (model, net, gross, now_iso()))
        if sku:
            cur.execute("SELECT id FROM products WHERE sku=? LIMIT 1", (sku,))
            existing = cur.fetchone()
            if existing:
                cur.execute("""
                  UPDATE products
                  SET model=COALESCE(NULLIF(?, ''), model),
                      ean=COALESCE(NULLIF(?, ''), ean),
                      name=COALESCE(NULLIF(?, ''), name)
                  WHERE sku=?
                """, (model, ean, name, sku))
                pid = int(existing["id"])
            else:
                cur.execute(
                    "INSERT INTO products(sku, model, ean, name, created_at) VALUES (?,?,?,?,?)",
                    (sku, model, ean, name, now_iso())
                )
                pid = int(cur.lastrowid)
            changed_product_ids.append(pid)
            cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
    c.commit()
    c.close()
    if supabase_enabled():
        try:
            sync_local_table_to_supabase("pricing", "model")
        except Exception:
            pass
        try:
            sync_local_rows_to_supabase("products", "id", changed_product_ids)
        except Exception:
            pass
        try:
            sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
        except Exception:
            pass
    return redirect(url_for("pricing"))


# -------------------------
# CUSTOMERS
# -------------------------

@app.get("/customers")
def customers():
    maybe_pull_shared_from_supabase()
    q = norm(request.args.get("q"))
    c = conn()
    cur = c.cursor()
    if q:
        like = f"%{q}%"
        cur.execute("""
          SELECT * FROM customers
          WHERE name LIKE ? OR phone LIKE ? OR email LIKE ? OR address LIKE ? OR nip LIKE ?
          ORDER BY id DESC
          LIMIT 500
        """, (like, like, like, like, like))
    else:
        cur.execute("SELECT * FROM customers ORDER BY id DESC LIMIT 500")
    rows = cur.fetchall()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <h1>Klienci stali</h1>
        <form method="get" class="grid3" style="margin-top:10px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj: nazwa / telefon / email / adres / NIP">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('customers') }}">WyczyĹ›Ä‡</a>
        </form>
      </div>

      <div class="card">
        <h2>Dodaj klienta</h2>
        <form method="post" action="{{ url_for('customers_create') }}" class="row">
          <div>
            <label class="muted small">Nazwa</label>
            <input name="name" required>
          </div>
          <div>
            <label class="muted small">Telefon</label>
            <input name="phone">
          </div>
          <div>
            <label class="muted small">Email</label>
            <input name="email">
          </div>
          <div>
            <label class="muted small">NIP</label>
            <input name="nip" placeholder="np. 1234567890">
          </div>
          <div>
            <label class="muted small">Adres</label>
            <textarea name="address" placeholder="Ulica, kod, miasto"></textarea>
          </div>
          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Zapisz klienta</button>
          </div>
        </form>
      </div>

      <div class="card">
        <h2>Lista klientĂłw</h2>
        <table>
          <thead>
            <tr><th>Nazwa</th><th>Telefon</th><th>Email</th><th>NIP</th><th>Adres</th><th>Akcje</th></tr>
          </thead>
          <tbody>
            {% for r in rows %}
              <tr>
                <td><b>{{ r['name'] }}</b></td>
                <td>{{ r['phone'] or '-' }}</td>
                <td>{{ r['email'] or '-' }}</td>
                <td>{{ r['nip'] or '-' }}</td>
                <td style="white-space:pre-line;">{{ r['address'] or '-' }}</td>
                <td>
                  <div class="flex">
                    <a class="btn" href="{{ url_for('customers_edit', customer_id=r['id']) }}">Edytuj</a>
                    <form method="post" action="{{ url_for('customers_delete', customer_id=r['id']) }}" onsubmit="return confirm('UsunÄ…Ä‡ klienta?')">
                      <button class="btn danger" type="submit">UsuĹ„</button>
                    </form>
                  </div>
                </td>
              </tr>
            {% endfor %}
            {% if not rows %}
              <tr><td colspan="6" class="muted">Brak klientĂłw.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Klienci", base_url=BASE_URL, db_path=DB_PATH, rows=rows, q=q)

@app.post("/customers/create")
def customers_create():
    name = norm(request.form.get("name"))
    address = norm(request.form.get("address"))
    phone = norm(request.form.get("phone"))
    email = norm(request.form.get("email"))
    nip = norm(request.form.get("nip"))
    if not name:
        return "Brak nazwy klienta", 400

    if supabase_enabled():
        remote_first_create_customer(name, address, phone, email, nip)
    else:
        c = conn()
        cur = c.cursor()
        cur.execute(
            "INSERT INTO customers(name, address, phone, email, nip, created_at) VALUES (?,?,?,?,?,?)",
            (name, address, phone, email, nip, now_iso())
        )
        c.commit()
        c.close()

    try:
        link_orders_to_customers_by_email(sync_remote=True)
    except Exception:
        pass
    return redirect(url_for("customers"))

@app.get("/customers/<int:customer_id>/edit")
def customers_edit(customer_id):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM customers WHERE id=?", (customer_id,))
    row = cur.fetchone()
    c.close()
    if not row:
        return "Nie znaleziono klienta", 404

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <h1>Edycja klienta</h1>
        <div class="muted">ZmieĹ„ dane zapisane dla staĹ‚ego klienta.</div>
      </div>

      <div class="card">
        <form method="post" action="{{ url_for('customers_update', customer_id=row['id']) }}" class="row">
          <div>
            <label class="muted small">Nazwa</label>
            <input name="name" value="{{ row['name'] }}" required>
          </div>
          <div>
            <label class="muted small">Telefon</label>
            <input name="phone" value="{{ row['phone'] or '' }}">
          </div>
          <div>
            <label class="muted small">Email</label>
            <input name="email" value="{{ row['email'] or '' }}">
          </div>
          <div>
            <label class="muted small">NIP</label>
            <input name="nip" value="{{ row['nip'] or '' }}" placeholder="np. 1234567890">
          </div>
          <div>
            <label class="muted small">Adres</label>
            <textarea name="address" placeholder="Ulica, kod, miasto">{{ row['address'] or '' }}</textarea>
          </div>
          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Zapisz zmiany</button>
            <a class="btn" href="{{ url_for('customers') }}">PowrĂłt</a>
          </div>
        </form>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Edycja klienta", base_url=BASE_URL, db_path=DB_PATH, row=row)

@app.post("/customers/<int:customer_id>/update")
def customers_update(customer_id):
    name = norm(request.form.get("name"))
    address = norm(request.form.get("address"))
    phone = norm(request.form.get("phone"))
    email = norm(request.form.get("email"))
    nip = norm(request.form.get("nip"))
    if not name:
        return "Brak nazwy klienta", 400

    c = conn()
    cur = c.cursor()
    cur.execute("""
      UPDATE customers
      SET name=?, address=?, phone=?, email=?, nip=?
      WHERE id=?
    """, (name, address, phone, email, nip, customer_id))
    c.commit()
    c.close()

    if supabase_enabled():
        supabase_update_rows("customers", {
            "name": name,
            "address": address,
            "phone": phone,
            "email": email,
            "nip": nip,
        }, {"id": customer_id})

    try:
        link_orders_to_customers_by_email(sync_remote=True)
    except Exception:
        pass
    return redirect(url_for("customers"))

@app.post("/customers/<int:customer_id>/delete")
def customers_delete(customer_id):
    if supabase_enabled():
        supabase_delete_rows("customers", {"id": customer_id})

    c = conn()
    cur = c.cursor()
    cur.execute("DELETE FROM customers WHERE id=?", (customer_id,))
    c.commit()
    c.close()
    return redirect(url_for("customers"))


# -------------------------
# PRODUCTS
# -------------------------

@app.get("/products")
def products():
    maybe_pull_shared_from_supabase()
    q = norm(request.args.get("q"))
    c = conn()
    cur = c.cursor()
    if q:
        like = f"%{q}%"
        cur.execute("""
          SELECT p.*
          FROM products p
          WHERE p.sku LIKE ? OR p.model LIKE ? OR p.ean LIKE ? OR p.name LIKE ?
          ORDER BY p.sku
          LIMIT 1000
        """, (like, like, like, like))
    else:
        cur.execute("""
          SELECT p.*
          FROM products p
          ORDER BY p.sku
          LIMIT 1000
        """)
    rows = cur.fetchall()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Produkty końcowe / receptury</h1>
          <div class="right"></div>
        </div>
        <form method="get" class="grid3" style="margin-top:10px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj: SKU / model / EAN / nazwa">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('products') }}">WyczyĹ›Ä‡</a>
        </form>
      </div>

      <div class="card">
        <h2>Import CSV wyrobów gotowych</h2>
        <div class="muted">Wybierz plik CSV z Excela. Minimalnie: kolumna SKU (unikalna). PozostaĹ‚e: model, ean, name/nazwa.</div>
        <div class="notice" style="margin-top:10px;">Kolumny dla betoniarni: <b>kod</b>, <b>produkt</b>, <b>jednostka</b> (m3, t, kg, szt.), <b>cena_netto</b>, <b>cena_brutto</b>, <b>koszt_materialu</b>, <b>koszt_produkcji</b>, <b>koszt_transportu</b>, <b>koszt_inny</b>. Ceny i koszty są podawane za jedną jednostkę.</div>
        <form method="post" action="{{ url_for('products_import') }}" enctype="multipart/form-data" class="row" style="margin-top:10px;">
          <div>
            <input type="file" name="file" accept=".csv,text/csv" required>
            <div class="muted small" style="margin-top:6px;">Kodowanie: najlepiej UTF-8. Separator zwykle â€ž;â€ť lub â€ž,â€ť â€“ program sam sprĂłbuje.</div>
          </div>
          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Importuj</button>
          </div>
        </form>
      </div>

      <div class="card">
        <h2>Lista (max 1000)</h2>
        <table>
          <thead>
            <tr>
              <th>SKU</th>
              <th>Produkt</th>
              <th>Jednostka</th>
              <th>Cena netto / j.</th>
              <th>Cena brutto / j.</th>
              <th>Koszt całkowity / j.</th>
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
            <tr>
              <td><b>{{ r["sku"] }}</b></td>
              <td>{{ r["name"] or r["model"] or "" }}</td>
              <td>{{ r["unit"] or "m3" }}</td>
              <td>{{ "%.2f"|format(r["unit_net_price"] or 0) }} PLN</td>
              <td>{{ "%.2f"|format(r["unit_gross_price"] or 0) }} PLN</td>
              <td>{{ "%.2f"|format((r["unit_material_cost"] or 0) + (r["unit_production_cost"] or 0) + (r["unit_transport_cost"] or 0) + (r["unit_other_cost"] or 0)) }} PLN</td>
            </tr>
            {% endfor %}
            {% if not rows %}
              <tr><td colspan="5" class="muted">Brak produktĂłw. ZrĂłb import CSV.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Produkty", base_url=BASE_URL, db_path=DB_PATH, rows=rows, q=q)

@app.post("/products/import")
def products_import():
    f = request.files.get("file")
    if not f:
        return "Brak pliku", 400

    raw = f.read()
    # SprĂłbuj UTF-8, jak nie pĂłjdzie to latin2
    try:
        text = raw.decode("utf-8-sig")
    except:
        text = raw.decode("latin2", errors="replace")

    # SprĂłbuj wykryÄ‡ delimiter
    sample = text[:5000]
    delim = ";" if sample.count(";") >= sample.count(",") else ","

    rdr = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(rdr)
    if not rows:
        return "Pusty CSV", 400

    headers = rows[0]
    data = rows[1:]

    i_sku = guess_col(headers, ["sku", "symbol", "index", "indeks", "kod", "code"])
    i_model = guess_col(headers, ["model", "model_uchwytu", "nazwa_modelu"])
    i_ean = guess_col(headers, ["ean", "gtin"])
    i_name = guess_col(headers, ["name", "nazwa", "produkt", "product"])
    i_unit = guess_col(headers, ["jednostka", "jm", "unit", "miara"])
    i_net = guess_col(headers, ["cena_netto", "cena jednostkowa netto", "netto", "price_net", "unit_net_price"])
    i_gross = guess_col(headers, ["cena_brutto", "cena jednostkowa brutto", "brutto", "price_gross", "unit_gross_price"])
    i_material_cost = guess_col(headers, ["koszt_materialu", "koszt materiału", "material_cost"])
    i_production_cost = guess_col(headers, ["koszt_produkcji", "production_cost"])
    i_transport_cost = guess_col(headers, ["koszt_transportu", "transport_cost"])
    i_other_cost = guess_col(headers, ["koszt_inny", "koszt pozostały", "other_cost"])

    # A cement-plant export may identify a product by its name only.
    # The internal SKU then uses that name as a stable import key.
    if i_sku is None and i_name is not None:
        i_sku = i_name

    if i_sku is None:
        return "CSV musi mieÄ‡ kolumnÄ™ SKU / Symbol / Indeks", 400

    c = conn()
    cur = c.cursor()
    added = 0
    updated = 0

    for row in data:
        if not row or len(row) <= i_sku:
            continue
        sku = norm(row[i_sku]) if i_sku is not None and len(row) > i_sku else ""
        if not sku:
            sku = f"BETON-{data.index(row)+1}"
        if not sku:
            continue
        model = norm(row[i_model]) if i_model is not None and len(row) > i_model else ""
        ean = norm(row[i_ean]) if i_ean is not None and len(row) > i_ean else ""
        name = norm(row[i_name]) if i_name is not None and len(row) > i_name else ""
        unit = norm(row[i_unit]) if i_unit is not None and len(row) > i_unit else "m3"
        net = to_float(row[i_net], 0) if i_net is not None and len(row) > i_net else 0
        gross = to_float(row[i_gross], 0) if i_gross is not None and len(row) > i_gross else round(net * 1.23, 2)
        material_cost = to_float(row[i_material_cost], 0) if i_material_cost is not None and len(row) > i_material_cost else 0
        production_cost = to_float(row[i_production_cost], 0) if i_production_cost is not None and len(row) > i_production_cost else 0
        transport_cost = to_float(row[i_transport_cost], 0) if i_transport_cost is not None and len(row) > i_transport_cost else 0
        other_cost = to_float(row[i_other_cost], 0) if i_other_cost is not None and len(row) > i_other_cost else 0

        cur.execute("SELECT id FROM products WHERE sku=?", (sku,))
        exists = cur.fetchone()
        if exists:
            cur.execute("UPDATE products SET model=?, ean=?, name=?, unit=?, unit_net_price=?, unit_gross_price=?, unit_material_cost=?, unit_production_cost=?, unit_transport_cost=?, unit_other_cost=? WHERE sku=?", (model, ean, name, unit, net, gross, material_cost, production_cost, transport_cost, other_cost, sku))
            updated += 1
            pid = exists["id"]
        else:
            cur.execute(
                "INSERT INTO products(sku, model, ean, name, unit, unit_net_price, unit_gross_price, unit_material_cost, unit_production_cost, unit_transport_cost, unit_other_cost, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (sku, model, ean, name, unit, net, gross, material_cost, production_cost, transport_cost, other_cost, now_iso())
            )
            pid = cur.lastrowid
            added += 1

        cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
        cur.execute("INSERT INTO pricing(model, net_price, gross_price, created_at) VALUES (?,?,?,?) ON CONFLICT(model) DO UPDATE SET net_price=excluded.net_price, gross_price=excluded.gross_price, created_at=excluded.created_at", (model or sku, net, gross, now_iso()))

    c.commit()
    c.close()

    return redirect(url_for("products", q=""))


# -------------------------
# STOCK
# -------------------------

def stock():
    maybe_pull_shared_from_supabase()
    q = norm(request.args.get("q"))
    c = conn()
    cur = c.cursor()

    if q:
        like = f"%{q}%"
        cur.execute("""
          SELECT x.*,
                 CASE WHEN x.ordered_new - x.qty > 0 THEN x.ordered_new - x.qty ELSE 0 END AS reserved_in_delivery,
                 CASE WHEN x.in_delivery - (CASE WHEN x.ordered_new - x.qty > 0 THEN x.ordered_new - x.qty ELSE 0 END) > 0
                      THEN x.in_delivery - (CASE WHEN x.ordered_new - x.qty > 0 THEN x.ordered_new - x.qty ELSE 0 END)
                      ELSE 0
                 END AS available_in_delivery
          FROM (
            SELECT p.id, p.sku, p.model, p.ean, p.name,
                   COALESCE(s.qty,0) AS qty,
                   COALESCE((
                      SELECT SUM(ci.qty)
                      FROM material_order_items ci
                      JOIN material_orders cp ON cp.id=ci.package_id
                      WHERE ci.product_id=p.id
                        AND cp.status IN ('planned', 'ordered', 'shipped')
                   ), 0) AS in_delivery,
                   COALESCE((
                      SELECT SUM(oi.qty)
                      FROM order_items oi
                      JOIN orders o ON o.id=oi.order_id
                      WHERE oi.product_id=p.id
                        AND o.status='new'
                   ), 0) AS ordered_new
            FROM products p
            LEFT JOIN stock s ON s.product_id=p.id
            WHERE p.sku LIKE ? OR p.model LIKE ? OR p.ean LIKE ? OR p.name LIKE ?
          ) x
          ORDER BY x.sku
          LIMIT 1000
        """, (like, like, like, like))
    else:
        cur.execute("""
          SELECT x.*,
                 CASE WHEN x.ordered_new - x.qty > 0 THEN x.ordered_new - x.qty ELSE 0 END AS reserved_in_delivery,
                 CASE WHEN x.in_delivery - (CASE WHEN x.ordered_new - x.qty > 0 THEN x.ordered_new - x.qty ELSE 0 END) > 0
                      THEN x.in_delivery - (CASE WHEN x.ordered_new - x.qty > 0 THEN x.ordered_new - x.qty ELSE 0 END)
                      ELSE 0
                 END AS available_in_delivery
          FROM (
            SELECT p.id, p.sku, p.model, p.ean, p.name,
                   COALESCE(s.qty,0) AS qty,
                   COALESCE((
                      SELECT SUM(ci.qty)
                      FROM material_order_items ci
                      JOIN material_orders cp ON cp.id=ci.package_id
                      WHERE ci.product_id=p.id
                        AND cp.status IN ('planned', 'ordered', 'shipped')
                   ), 0) AS in_delivery,
                   COALESCE((
                      SELECT SUM(oi.qty)
                      FROM order_items oi
                      JOIN orders o ON o.id=oi.order_id
                      WHERE oi.product_id=p.id
                        AND o.status='new'
                   ), 0) AS ordered_new
            FROM products p
            LEFT JOIN stock s ON s.product_id=p.id
          ) x
          ORDER BY x.sku
          LIMIT 1000
        """)
    rows = cur.fetchall()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Magazyn</h1>
        </div>
        <form method="get" class="grid3" style="margin-top:10px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj produktu: SKU / model / EAN / nazwa">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('stock') }}">WyczyĹ›Ä‡</a>
        </form>
      </div>

      <div class="card">
        <h2>Korekta stanu</h2>
        <div class="row">
          <div>
            <label class="muted small">Produkt (SKU)</label>
            <input list="skuList" id="skuInput" placeholder="np. CH010-BB-N28">
            <datalist id="skuList">
              {% for r in rows %}
                <option value="{{ r['sku'] }}">{{ r['sku'] }}</option>
              {% endfor %}
            </datalist>
          </div>
          <div>
            <label class="muted small">Zmiana (np. +10 albo -3)</label>
            <input id="deltaInput" placeholder="+10">
          </div>
        </div>
        <div class="flex" style="margin-top:10px;">
          <button class="btn ok" onclick="applyDelta(); return false;">Zapisz korektÄ™</button>
          <div class="muted" id="deltaMsg"></div>
        </div>
      </div>

      <div class="card">
        <h2>Stany (max 1000)</h2>
        <div class="muted" style="margin-bottom:8px;">
          Najpierw realizowane sÄ… iloĹ›ci z magazynu. Niedobory z otwartych zamĂłwieĹ„ (status <b>new</b>) rezerwujÄ… towar â€žw drodzeâ€ť.
        </div>
        <table>
          <thead>
            <tr><th>SKU</th><th>Model</th><th>EAN</th><th>Nazwa</th><th>Stan</th><th>W drodze</th><th>Zarezerwowane w drodze</th><th>DostÄ™pne w drodze</th></tr>
          </thead>
          <tbody>
            {% for r in rows %}
              <tr>
                <td><b>{{ r['sku'] }}</b></td>
                <td>{{ r['model'] or "" }}</td>
                <td>{{ r['ean'] or "" }}</td>
                <td>{{ r['name'] or "" }}</td>
                <td><span class="badge">{{ r['qty'] }}</span></td>
                <td><span class="badge">{{ r['in_delivery'] }}</span></td>
                <td><span class="badge">{{ r['reserved_in_delivery'] }}</span></td>
                <td><span class="badge">{{ r['available_in_delivery'] }}</span></td>
              </tr>
            {% endfor %}
            {% if not rows %}
              <tr><td colspan="8" class="muted">Brak produktĂłw.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>

<script>
async function applyDelta(){
  const sku = document.getElementById("skuInput").value.trim();
  const delta = document.getElementById("deltaInput").value.trim();
  const msg = document.getElementById("deltaMsg");
  msg.innerText = "";
  if(!sku){ msg.innerText = "Podaj SKU"; return; }
  if(!delta){ msg.innerText = "Podaj zmianÄ™"; return; }

  const r = await fetch("/api/stock_delta", {
    method:"POST",
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({sku, delta})
  });
  const j = await r.json();
  if(!j.ok){ msg.innerText = "BĹ‚Ä…d: " + (j.error || ""); return; }
  msg.innerText = "OK. Nowy stan: " + j.new_qty;
  setTimeout(()=>location.reload(), 500);
}
</script>

    {% endblock %}
    """
    return render_template_string(tpl, title="Magazyn", base_url=BASE_URL, db_path=DB_PATH, rows=rows, q=q)

def api_stock_delta():
    data = request.get_json(force=True, silent=True) or {}
    sku = norm(data.get("sku"))
    delta_raw = norm(data.get("delta"))

    if not sku:
        return jsonify(ok=False, error="Brak SKU"), 400

    delta = to_int(delta_raw, None)
    if delta is None:
        # sprĂłbuj +10 / -3
        try:
            delta = int(delta_raw)
        except:
            return jsonify(ok=False, error="NieprawidĹ‚owa zmiana (np. +10 lub -3)"), 400

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id FROM products WHERE sku=?", (sku,))
    p = cur.fetchone()
    if not p:
        c.close()
        return jsonify(ok=False, error="Nie ma takiego SKU"), 404
    pid = p["id"]
    cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
    cur.execute("UPDATE stock SET qty = qty + ? WHERE product_id=?", (delta, pid))
    cur.execute("SELECT qty FROM stock WHERE product_id=?", (pid,))
    new_qty = cur.fetchone()["qty"]
    c.commit()
    c.close()
    return jsonify(ok=True, new_qty=new_qty)

@app.get("/api/product/<int:product_id>")
def api_product(product_id):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT p.*, COALESCE(s.qty,0) AS stock
      FROM products p
      LEFT JOIN stock s ON s.product_id=p.id
      WHERE p.id=?
    """, (product_id,))
    r = cur.fetchone()
    c.close()
    if not r:
        return jsonify(ok=False), 404
    return jsonify(ok=True, id=r["id"], sku=r["sku"], model=r["model"], ean=r["ean"], name=r["name"], stock=r["stock"])


# -------------------------
# ORDERS
# -------------------------

def render_orders_workboard(q):
    """Prosta lista pracy: bez licznikow czasowych i technicznych etapow."""
    selected = norm(request.args.get("tab")) or "in_progress"
    allowed = {"in_progress", "planned", "today", "tomorrow", "completed"}
    if selected not in allowed:
        selected = "in_progress"

    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    c = conn()
    cur = c.cursor()
    cur.execute("""
        SELECT o.id, o.order_no, o.customer_name, o.created_at,
               o.delivery_date, o.delivery_time,
               w.id AS wz_id, w.wz_no, w.status AS wz_status,
               t.id AS transport_id, t.transport_no, t.status AS transport_status,
               d.name AS driver_name, v.registration_no
        FROM orders o
        LEFT JOIN wz_documents w ON w.id=(
            SELECT x.id FROM wz_documents x
            WHERE x.order_id=o.id AND x.deleted_at IS NULL
            ORDER BY x.id DESC LIMIT 1
        )
        LEFT JOIN transports t ON t.id=(
            SELECT x.id FROM transports x
            WHERE x.wz_id=w.id AND x.deleted_at IS NULL
            ORDER BY x.id DESC LIMIT 1
        )
        LEFT JOIN drivers d ON d.id=t.driver_id
        LEFT JOIN vehicles v ON v.id=t.vehicle_id
        ORDER BY o.created_at DESC LIMIT 300
    """)
    rows = [dict(r) for r in cur.fetchall()]
    c.close()

    counters = {key: 0 for key in allowed}
    visible = []
    for row in rows:
        date_text = norm(row.get("delivery_date"))
        try:
            delivery_day = datetime.strptime(date_text, "%Y-%m-%d").date() if date_text else None
        except ValueError:
            delivery_day = None
        transport_status = norm(row.get("transport_status"))
        wz_status = norm(row.get("wz_status"))

        if transport_status == "returned" or wz_status in {"returned", "ready_invoice", "invoiced"}:
            bucket = "completed"
            row["status_label"] = "Zrealizowane"
        elif delivery_day == today:
            bucket = "today"
            row["status_label"] = {"in_transit": "W dostawie", "delivered": "WZ podpisane"}.get(
                transport_status, "Do wydania dziś"
            )
        elif delivery_day == tomorrow:
            bucket = "tomorrow"
            row["status_label"] = "Do wydania jutro"
        elif delivery_day and delivery_day > tomorrow:
            bucket = "planned"
            row["status_label"] = "Zaplanowane"
        else:
            bucket = "in_progress"
            row["status_label"] = {
                "in_transit": "W dostawie",
                "delivered": "WZ podpisane",
                "assigned": "Oczekuje na wyjazd",
                "issued": "Oczekuje na wyjazd",
            }.get(transport_status, "W trakcie realizacji")
        row["bucket"] = bucket
        row["delivery_label"] = delivery_day.strftime("%d.%m.%Y") if delivery_day else "Termin nieustalony"
        row["transport_label"] = row.get("transport_no") or "Nieprzydzielony"
        row["driver_label"] = " · ".join(x for x in (row.get("driver_name"), row.get("registration_no")) if x) or "—"
        counters[bucket] += 1
        if bucket != "completed":
            counters["in_progress"] += 1

        haystack = " ".join(str(row.get(k) or "") for k in ("order_no", "customer_name", "wz_no", "transport_no", "driver_name", "registration_no")).lower()
        matches = not q or q.lower() in haystack
        selected_match = bucket != "completed" if selected == "in_progress" else bucket == selected
        if matches and selected_match:
            visible.append(row)

    labels = {
        "in_progress": "W trakcie realizacji",
        "planned": "Zaplanowane",
        "today": "Do wydania dziś",
        "tomorrow": "Do wydania jutro",
        "completed": "Zrealizowane",
    }
    return render_template_string("""
    {% extends 'base.html' %}
    {% block content %}
    <div class="card">
      <div class="row between"><div><h1>Zamówienia</h1><p class="muted">Bieżąca lista realizacji i planowanych wydań.</p></div><a class="btn" href="{{ url_for('order_new') }}">+ Nowe zamówienie</a></div>
      <div class="tabs" style="margin:18px 0">
      {% for key, label in labels.items() %}
        <a class="btn {% if selected == key %}primary{% endif %}" href="{{ url_for('orders', tab=key, q=q) }}">{{ label }} ({{ counters[key] }})</a>
      {% endfor %}
      </div>
      <form method="get" class="row">
        <input type="hidden" name="tab" value="{{ selected }}"><input name="q" value="{{ q }}" placeholder="Szukaj: klient, WZ, transport, kierowca">
        <button class="btn primary">Szukaj</button><a class="btn" href="{{ url_for('orders', tab=selected) }}">Wyczyść</a>
      </form>
    </div>
    <div class="card"><table><thead><tr><th>Zamówienie / klient</th><th>Termin dostawy</th><th>Status</th><th>WZ / transport</th><th>Kierowca / auto</th></tr></thead><tbody>
    {% for row in rows %}<tr>
      <td><a href="{{ url_for('order_view', order_id=row.id) }}"><b>{{ row.order_no }}</b></a><br><span class="muted">{{ row.customer_name }}</span></td>
      <td>{{ row.delivery_label }}</td><td><span class="badge">{{ row.status_label }}</span></td>
      <td>{% if row.wz_id %}<a href="{{ url_for('beton.wz_view', wz_id=row.wz_id) }}">{{ row.wz_no }}</a>{% else %}—{% endif %}<br><span class="muted">{{ row.transport_label }}</span></td>
      <td>{{ row.driver_label }}</td>
    </tr>{% else %}<tr><td colspan="5">Brak zamówień w wybranej kolejce.</td></tr>{% endfor %}
    </tbody></table></div>
    {% endblock %}
    """, rows=visible, counters=counters, labels=labels, selected=selected, q=q)

@app.get("/orders")
def orders():
    maybe_pull_shared_from_supabase()
    try:
        link_orders_to_customers_by_email(sync_remote=True)
    except Exception:
        pass
    q = norm(request.args.get("q"))
    return render_orders_workboard(q)
    # Operational delivery board: orders are entered by staff, not by a client portal.
    # It uses the WZ and transport timeline as the source of truth.
    delivery_tab = norm(request.args.get("tab")) or "realization"
    allowed_delivery_tabs = {"realization", "issued", "delivery", "delivered", "returned", "all"}
    if delivery_tab not in allowed_delivery_tabs:
        delivery_tab = "realization"
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT o.id,o.order_no,o.customer_name,o.created_at,
             w.id AS wz_id,w.wz_no,w.status AS wz_status,w.created_at AS wz_created_at,w.issued_at,
             t.id AS transport_id,t.transport_no,t.status AS transport_status,t.driver_id,
             d.name AS driver_name,v.registration_no,t.departed_at,t.delivered_at,t.returned_at
      FROM orders o
      LEFT JOIN wz_documents w ON w.id=(SELECT id FROM wz_documents x WHERE x.order_id=o.id AND x.deleted_at IS NULL ORDER BY x.id DESC LIMIT 1)
      LEFT JOIN transports t ON t.id=(SELECT id FROM transports x WHERE x.wz_id=w.id AND x.deleted_at IS NULL ORDER BY x.id DESC LIMIT 1)
      LEFT JOIN drivers d ON d.id=t.driver_id
      LEFT JOIN vehicles v ON v.id=t.vehicle_id
      ORDER BY o.id DESC LIMIT 500
    """)
    delivery_rows=[]
    stage_counts={x:0 for x in allowed_delivery_tabs if x != "all"}
    durations={"realization":[],"issued":[],"delivery":[],"delivered":[]}

    def _delivery_time(value):
        if not value:
            return None
        try:
            return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        except Exception:
            return None
    def _duration_label(seconds):
        if seconds is None or seconds < 0:
            return "—"
        minutes=round(seconds / 60)
        if minutes < 60:
            return f"{minutes} min"
        hours, minutes=divmod(minutes,60)
        return f"{hours} h {minutes} min"
    for raw in cur.fetchall():
        r=dict(raw)
        ts={key:_delivery_time(r.get(key)) for key in ("created_at","issued_at","departed_at","delivered_at","returned_at")}
        transport_status=(r.get("transport_status") or "").lower()
        if transport_status == "returned":
            stage, stage_label="returned","Auto wróciło na bazę"
        elif transport_status == "delivered":
            stage, stage_label="delivered","Dostarczone"
        elif transport_status == "in_transit":
            stage, stage_label="delivery","W dostawie"
        elif r.get("wz_status") in ("issued","in_transport") or transport_status in ("assigned","issued","problem"):
            stage, stage_label="issued","Wydane / oczekuje na wyjazd"
        else:
            stage, stage_label="realization","W realizacji"
        r["stage"]=stage; r["stage_label"]=stage_label
        r["t_realization"]=_duration_label((ts["issued_at"]-ts["created_at"]).total_seconds()) if ts["issued_at"] and ts["created_at"] else "—"
        r["t_wait_departure"]=_duration_label((ts["departed_at"]-ts["issued_at"]).total_seconds()) if ts["departed_at"] and ts["issued_at"] else "—"
        r["t_delivery"]=_duration_label((ts["delivered_at"]-ts["departed_at"]).total_seconds()) if ts["delivered_at"] and ts["departed_at"] else "—"
        r["t_return"]=_duration_label((ts["returned_at"]-ts["delivered_at"]).total_seconds()) if ts["returned_at"] and ts["delivered_at"] else "—"
        for key, start, end in (("realization","created_at","issued_at"),("issued","issued_at","departed_at"),("delivery","departed_at","delivered_at"),("delivered","delivered_at","returned_at")):
            if ts[start] and ts[end]: durations[key].append((ts[end]-ts[start]).total_seconds())
        stage_counts[stage]=stage_counts.get(stage,0)+1
        if (not q or q.lower() in " ".join(str(r.get(k) or "") for k in ("order_no","customer_name","wz_no","transport_no","driver_name")).lower()) and (delivery_tab == "all" or stage == delivery_tab):
            delivery_rows.append(r)
    c.close()
    averages={key:_duration_label(sum(vals)/len(vals)) if vals else "—" for key,vals in durations.items()}
    board_tpl=r"""
    {% extends "base.html" %}{% block content %}
    <div class="card"><div class="flex"><h1 style="margin:0">Realizacja dostaw</h1><a class="btn primary right" href="{{url_for('order_new')}}">+ Nowe zamówienie</a></div><div class="muted" style="margin-top:8px">Tablica operacyjna oparta na dokumentach WZ i statusach kierowców. Czasy pokazują miejsca, w których proces czeka najdłużej.</div><div class="flex" style="margin-top:14px">{% for key,label in [('realization','W realizacji'),('issued','Wydane'),('delivery','W dostawie'),('delivered','Dostarczone'),('returned','Auto na bazie'),('all','Wszystkie')] %}<a class="btn {% if tab==key %}primary{% endif %}" href="{{url_for('orders',tab=key,q=q)}}">{{label}}{% if key!='all' %} ({{counts.get(key,0)}}){% endif %}</a>{% endfor %}</div><form method="get" class="grid3" style="margin-top:12px"><input type="hidden" name="tab" value="{{tab}}"><input name="q" value="{{q}}" placeholder="Szukaj: klient, WZ, transport, kierowca"><button class="btn primary">Szukaj</button><a class="btn" href="{{url_for('orders',tab=tab)}}">Wyczyść</a></form></div>
    <div class="grid4"><div class="card"><b>Przygotowanie → wydanie</b><div style="font-size:22px">{{averages.realization}}</div></div><div class="card"><b>Wydane → wyjazd</b><div style="font-size:22px">{{averages.issued}}</div></div><div class="card"><b>Wyjazd → dostawa</b><div style="font-size:22px">{{averages.delivery}}</div></div><div class="card"><b>Dostawa → powrót</b><div style="font-size:22px">{{averages.delivered}}</div></div></div>
    <div class="card"><table><thead><tr><th>Zamówienie / klient</th><th>Etap</th><th>WZ / transport</th><th>Kierowca / auto</th><th>Przygot.</th><th>Oczek. wyjazd</th><th>Dostawa</th><th>Powrót</th><th></th></tr></thead><tbody>{% for r in rows %}<tr><td><b>{{r.order_no}}</b><br>{{r.customer_name}}</td><td><span class="badge">{{r.stage_label}}</span></td><td>{{r.wz_no or 'WZ nie wystawiono'}}{% if r.transport_no %}<br><b>{{r.transport_no}}</b>{% endif %}</td><td>{{r.driver_name or '—'}}<br>{{r.registration_no or ''}}</td><td>{{r.t_realization}}</td><td>{{r.t_wait_departure}}</td><td>{{r.t_delivery}}</td><td>{{r.t_return}}</td><td>{% if r.transport_id %}<a class="btn" href="{{url_for('beton.transport_view',transport_id=r.transport_id)}}">Transport</a>{% elif r.wz_id %}<a class="btn" href="{{url_for('beton.wz_view',wz_id=r.wz_id)}}">WZ</a>{% else %}<a class="btn" href="{{url_for('order_view',order_id=r.id)}}">Zamówienie</a>{% endif %}</td></tr>{% else %}<tr><td colspan="9" class="muted">Brak pozycji w wybranym etapie.</td></tr>{% endfor %}</tbody></table></div>
    {% endblock %}"""
    return render_template_string(board_tpl,title="Realizacja dostaw",base_url=BASE_URL,db_path=DB_PATH,rows=delivery_rows,tab=delivery_tab,q=q,counts=stage_counts,averages=averages)

    tab = norm(request.args.get("tab")) or "new"
    if tab not in {"new", "issued", "realized", "all"}:
        tab = "new"

    c = conn()
    cur = c.cursor()

    where_parts = []
    params = []

    if tab == "new":
        where_parts.append("COALESCE(o.warehouse_issued,0)=0")
        where_parts.append("LOWER(COALESCE(o.status,'')) NOT IN ('in_delivery','issued')")
    elif tab == "issued":
        where_parts.append("(LOWER(COALESCE(o.status,''))='in_delivery' OR (COALESCE(o.warehouse_issued,0)=1 AND LOWER(COALESCE(o.status,''))<>'issued'))")
    elif tab == "realized":
        where_parts.append("LOWER(COALESCE(o.status,''))='issued'")

    if q:
        where_parts.append("(order_no LIKE ? OR customer_name LIKE ?)")
        like = f"%{q}%"
        params.extend([like, like])

    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    sql = f"""
      SELECT o.*,
             COALESCE((
               SELECT SUM(oi.qty * COALESCE(pr.net_price, 0))
               FROM order_items oi
               LEFT JOIN products p ON p.id=oi.product_id
               LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
               WHERE oi.order_id=o.id
             ), 0) AS order_value_net,
             CASE WHEN EXISTS (
               SELECT 1
               FROM order_items oi
               LEFT JOIN stock s ON s.product_id=oi.product_id
               WHERE oi.order_id=o.id
                 AND (
                   COALESCE(s.qty,0) + COALESCE((
                     SELECT SUM(ci.qty)
                     FROM material_order_items ci
                     JOIN material_orders cp ON cp.id=ci.package_id
                     WHERE ci.product_id=oi.product_id
                       AND cp.status IN ('planned', 'ordered', 'shipped')
                   ),0)
                 ) < oi.qty
             ) THEN 1 ELSE 0 END AS has_shortage
      FROM orders o
      {where_sql}
      ORDER BY o.id DESC
      LIMIT 300
    """
    cur.execute(sql, tuple(params))
    rows = [dict(r) for r in cur.fetchall()]

    visible_open_ids = sorted([r["id"] for r in rows if int(r.get("warehouse_issued") or 0) == 0 and r["status"] in ("new", "packed", "confirmed", "in_delivery")])
    if visible_open_ids:
        cur.execute("SELECT id FROM orders WHERE COALESCE(warehouse_issued,0)=0 AND status IN ('new','packed','confirmed','in_delivery') AND id<=? ORDER BY id", (visible_open_ids[-1],))
        open_order_ids = [int(r["id"]) for r in cur.fetchall()]

        ph = ",".join(["?"] * len(open_order_ids))
        cur.execute(f"""
          SELECT oi.order_id, oi.product_id, SUM(oi.qty) AS qty
          FROM order_items oi
          WHERE oi.order_id IN ({ph})
          GROUP BY oi.order_id, oi.product_id
        """, tuple(open_order_ids))
        demand_rows = cur.fetchall()

        by_order = {}
        product_ids = set()
        for dr in demand_rows:
            oid = int(dr["order_id"])
            pid = int(dr["product_id"])
            qty = int(dr["qty"])
            by_order.setdefault(oid, []).append((pid, qty))
            product_ids.add(pid)

        pool_stock = {}
        pool_delivery = {}
        if product_ids:
            pph = ",".join(["?"] * len(product_ids))
            cur.execute(f"""
              SELECT p.id AS product_id,
                     COALESCE(s.qty,0) AS stock_qty,
                     COALESCE((
                       SELECT SUM(ci.qty)
                       FROM material_order_items ci
                       JOIN material_orders cp ON cp.id=ci.package_id
                       WHERE ci.product_id=p.id
                         AND cp.status IN ('planned', 'ordered', 'shipped')
                     ),0) AS in_delivery_qty
              FROM products p
              LEFT JOIN stock s ON s.product_id=p.id
              WHERE p.id IN ({pph})
            """, tuple(product_ids))
            for pr in cur.fetchall():
                pid = int(pr["product_id"])
                pool_stock[pid] = int(pr["stock_qty"])
                pool_delivery[pid] = int(pr["in_delivery_qty"])

        has_shortage = {oid: 0 for oid in open_order_ids}
        for oid in open_order_ids:
            for pid, need0 in by_order.get(oid, []):
                need = int(need0)
                stock_now = pool_stock.get(pid, 0)
                from_stock = min(stock_now, need)
                pool_stock[pid] = stock_now - from_stock
                need -= from_stock

                delivery_now = pool_delivery.get(pid, 0)
                from_delivery = min(delivery_now, need)
                pool_delivery[pid] = delivery_now - from_delivery
                need -= from_delivery

                if need > 0:
                    has_shortage[oid] = 1

        for r in rows:
            if r["status"] in ("new", "packed", "confirmed", "in_delivery"):
                r["has_shortage"] = has_shortage.get(r["id"], 0)
            else:
                r["has_shortage"] = 0

    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">ZamĂłwienia</h1>
          <a class="btn primary right" href="{{ url_for('order_new') }}">+ Nowe zamĂłwienie</a>
        </div>
        <div class="flex" style="margin-top:10px;">
          <a class="btn {% if tab=='new' %}primary{% endif %}" href="{{ url_for('orders', tab='new', q=q) }}">Do wydania</a>
          <a class="btn {% if tab=='issued' %}primary{% endif %}" href="{{ url_for('orders', tab='issued', q=q) }}">Wydane</a>
          <a class="btn {% if tab=='realized' %}primary{% endif %}" href="{{ url_for('orders', tab='realized', q=q) }}">Zrealizowane</a>
          <a class="btn {% if tab=='all' %}primary{% endif %}" href="{{ url_for('orders', tab='all', q=q) }}">Wszystkie</a>
        </div>
        <form method="get" class="grid3" style="margin-top:10px;">
          <input type="hidden" name="tab" value="{{ tab }}">
          <input name="q" value="{{ q }}" placeholder="Szukaj: numer zamĂłwienia lub klient">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('orders', tab=tab) }}">WyczyĹ›Ä‡</a>
        </form>
      </div>

      <style>
        .st-unconfirmed{background:#ef4444;color:#fff;border-color:#ef4444;}
        .st-confirmed{background:#16a34a;color:#fff;border-color:#16a34a;}
        .st-delivery{background:#2563eb;color:#fff;border-color:#2563eb;}
        .st-issued{background:#6b7280;color:#fff;border-color:#6b7280;}
      </style>

      <div class="card">
        <table>
          <thead>
            <tr><th>Nr</th><th>Klient</th><th>Status</th><th>WartoĹ›Ä‡ netto</th><th>Data</th><th>Akcje</th></tr>
          </thead>
          <tbody>
            {% for r in rows %}
              <tr {% if tab == 'new' and (r['has_shortage'] or r['status'] in ['new','pending','unconfirmed']) %}style="background:#ffe7e7;"{% endif %}>
                <td><b>{{ order_display_no(r['id'], r['created_at'], r['order_no'], r['note']) }}</b></td>
                <td>{{ r['customer_name'] }}</td>
                <td><span class="badge {{ order_status_css(r['status']) }}">{{ order_status_label(r['status']) }}</span></td>
                <td><span class="badge">{{ "%.2f"|format(r['order_value_net']) }} PLN</span></td>
                <td class="muted">{{ r['created_at'] }}</td>
                <td class="flex">
                  <a class="btn" href="{{ url_for('order_view', order_id=r['id']) }}">SzczegĂłĹ‚y</a>
                  {% if r['status'] != 'issued' %}
                    <form method="post" action="{{ url_for('order_delete', order_id=r['id']) }}" onsubmit="return confirm('UsunÄ…Ä‡ zamĂłwienie?')">
                      <button class="btn danger" type="submit">UsuĹ„</button>
                    </form>
                  {% else %}
                    <span class="muted">PodglÄ…d</span>
                  {% endif %}
                </td>
              </tr>
            {% endfor %}
            {% if not rows %}
              <tr><td colspan="5" class="muted">Brak zamĂłwieĹ„.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="ZamĂłwienia", base_url=BASE_URL, db_path=DB_PATH, rows=rows, q=q, tab=tab, order_status_label=order_status_label, order_status_css=order_status_css, canonical_order_no=canonical_order_no)

@app.get("/orders/new")
def order_new():
    maybe_pull_shared_from_supabase()
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id, sku, model, name FROM products ORDER BY sku LIMIT 5000")
    products_rows = cur.fetchall()
    cur.execute("SELECT id, name, address, phone, email, nip FROM customers ORDER BY name")
    customers_rows = cur.fetchall()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <h1>Nowe zamĂłwienie</h1>
        <div class="muted">Produkty wybierasz z bazy. Przy wyborze pokazuje stan magazynowy.</div>
      </div>

      <div class="card">
        <form method="post" action="{{ url_for('order_create') }}">
          <div class="row">
            <div>
              <label class="muted small">Wybierz staĹ‚ego klienta (opcjonalnie)</label>
              <select id="customerSelect" name="customer_id" onchange="fillCustomer(this.value)">
                <option value="">-- rÄ™cznie / nowy klient --</option>
                {% for c in customers %}
                  <option value="{{ c['id'] }}">{{ c['name'] }}</option>
                {% endfor %}
              </select>
            </div>
            <div class="muted">Po wyborze pola klienta zostanÄ… automatycznie uzupeĹ‚nione.</div>
          </div>

          <div class="row">
            <div>
              <label class="muted small">ZamawiajÄ…cy (nazwa firmy / osoba)</label>
              <input name="customer_name" required>
            </div>
            <div>
              <label class="muted small">Telefon</label>
              <input name="customer_phone">
            </div>
          </div>

          <div class="row" style="margin-top:10px;">
            <div>
              <label class="muted small">Adres dostawy</label>
              <textarea name="customer_address" placeholder="Ulica, kod, miasto"></textarea>
            </div>
            <div>
              <label class="muted small">Email</label>
              <input name="customer_email">
            </div>
          </div>

          <div class="row" style="margin-top:10px;">
            <div>
              <label class="muted small">Termin realizacji dostawy</label>
              <input name="delivery_date" type="date" required>
            </div>
            <div>
              <label class="muted small">Planowana godzina dostawy (opcjonalnie)</label>
              <input name="delivery_time" type="time">
            </div>
          </div>

          <div class="line"></div>

          <div class="flex">
            <h2 style="margin:0;">Pozycje zamĂłwienia</h2>
            <button class="btn" onclick="addItemRow(); return false;">+ Dodaj pozycjÄ™</button>
          </div>

          <div id="itemsContainer" style="margin-top:10px;"></div>

          <template id="itemRowTpl">
            <div class="items-row card" style="margin:10px 0;">
              <div>
                <label class="muted small">Produkt (SKU)</label>
                <select name="product_id[]">
                  <option value="">-- wybierz --</option>
                  {% for p in products %}
                    <option value="{{ p['id'] }}">{{ p['sku'] }}{% if p['model'] %} â€˘ {{ p['model'] }}{% endif %}{% if p['name'] %} â€˘ {{ p['name'] }}{% endif %}</option>
                  {% endfor %}
                </select>
              </div>
              <div>
                <label class="muted small">Ilość [m³]</label>
                <input name="qty[]" type="number" min="0.01" step="0.01" value="1">
              </div>
              <div>
                <label class="muted small">Produkt ręczny</label>
                <input name="manual_product[]" placeholder="np. Beton C25/30">
              </div>
              <div class="flex" style="align-items:flex-end;">
                <button class="btn danger" onclick="removeRow(this); return false;">UsuĹ„</button>
              </div>
            </div>
          </template>

          <div class="line"></div>
          <button class="btn primary" type="submit">Zapisz zamĂłwienie</button>
          <a class="btn" href="{{ url_for('orders') }}">Anuluj</a>
        </form>
      </div>

<script>
// po dodaniu wiersza trzeba podpiÄ…Ä‡ ID na badge (stan)
function addItemRow(){
  const tpl = document.getElementById("itemRowTpl");
  const container = document.getElementById("itemsContainer");
  const node = tpl.content.cloneNode(true);

  // znajdĹş select i badge w nowo wstawionym wierszu
  const wrap = node.querySelector(".items-row");
  const select = wrap.querySelector("select");
  const badge = {};

  const id = "stock_" + Math.random().toString(36).slice(2);
  badge.id = id;
  select.dataset.stockTarget = id;

  container.appendChild(node);
}

addItemRow(); // startowo 1 pozycja

const customersData = {{ customers_json|safe }};
function fillCustomer(customerId){
  if(!customerId || !customersData[customerId]) return;
  const c = customersData[customerId];
  document.querySelector('input[name="customer_name"]').value = c.name || '';
  document.querySelector('textarea[name="customer_address"]').value = c.address || '';
  document.querySelector('input[name="customer_phone"]').value = c.phone || '';
  document.querySelector('input[name="customer_email"]').value = c.email || '';
}
</script>

    {% endblock %}
    """
    customers_json = {
        str(r["id"]): {
            "name": r["name"],
            "address": r["address"],
            "phone": r["phone"],
            "email": r["email"],
        }
        for r in customers_rows
    }
    return render_template_string(
        tpl,
        title="Nowe zamĂłwienie",
        base_url=BASE_URL,
        db_path=DB_PATH,
        products=products_rows,
        customers=customers_rows,
        customers_json=json.dumps(customers_json, ensure_ascii=False)
    )

@app.post("/orders/create")
def order_create():
    customer_id = to_int(request.form.get("customer_id"), 0)
    customer_name = norm(request.form.get("customer_name"))
    if not customer_name:
        return "Brak zamawiajÄ…cego", 400

    customer_address = norm(request.form.get("customer_address"))
    customer_phone = norm(request.form.get("customer_phone"))
    customer_email = norm(request.form.get("customer_email"))
    delivery_date = norm(request.form.get("delivery_date"))
    delivery_time = norm(request.form.get("delivery_time"))
    if not delivery_date:
        return "Podaj termin realizacji dostawy", 400
    # W betoniarni adres klienta jest jednocześnie adresem dostawy.
    # Nie używamy osobnego adresu wysyłki ani etykiet kurierskich.
    note = ""

    product_ids = request.form.getlist("product_id[]")
    manual_products = request.form.getlist("manual_product[]")
    qtys = request.form.getlist("qty[]")

    items = []
    for index, (pid, q) in enumerate(zip(product_ids, qtys)):
        pid = to_int(pid, 0)
        qty = to_float(q, 0.0)
        if pid > 0 and qty > 0:
            items.append((pid, qty))
            continue
        manual_name = norm(manual_products[index] if index < len(manual_products) else "")
        if manual_name and qty > 0:
            sku = f"RECZNY-{uuid.uuid4().hex[:10].upper()}"
            created_at = now_iso()
            # W istniejącej tabeli Supabase kolumna products.id nie zawsze ma
            # generator ID. Nadajemy więc wspólny, niekolidujący identyfikator
            # zamiast oczekiwać, że Supabase sam go utworzy.
            pid = cloud_row_id()
            c = conn()
            try:
                cur = c.cursor()
                if supabase_enabled():
                    supabase_insert_row("products", {
                        "id": pid, "sku": sku, "model": manual_name, "name": manual_name,
                        "unit": "m3", "created_at": created_at,
                    })
                    cur.execute("INSERT INTO products(id,sku,model,ean,name,created_at) VALUES(?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET sku=excluded.sku,model=excluded.model,name=excluded.name", (pid, sku, manual_name, "", manual_name, created_at))
                else:
                    cur.execute("INSERT INTO products(id,sku,model,ean,name,created_at) VALUES(?,?,?,?,?,?)", (pid, sku, manual_name, "", manual_name, created_at))
                c.commit()
            finally:
                c.close()
            items.append((pid, qty))

    if not items:
        return "Dodaj minimum 1 pozycjÄ™", 400

    if supabase_enabled():
        try:
            oid = remote_first_create_order(customer_id if customer_id > 0 else None, customer_name, customer_address, customer_phone, customer_email, note, items, delivery_date=delivery_date, delivery_time=delivery_time)
        except Exception as exc:
            app.logger.exception("Nie udało się zapisać zamówienia w Supabase")
            return render_template_string("""
              <h1>Zamówienie nie zostało zapisane</h1>
              <p>Supabase odrzucił zapis danych. Zamówienie nie zostało utworzone.</p>
              <p><strong>Szczegół techniczny:</strong> {{ error }}</p>
              <p><a href="{{ url_for('order_new') }}">Wróć do formularza zamówienia</a></p>
            """, error=str(exc)), 400
    else:
        c = conn()
        cur = c.cursor()
        created_at = now_iso()
        cur.execute("""
          INSERT INTO orders(order_no, customer_id, customer_name, customer_address, customer_phone, customer_email, delivery_date, delivery_time, status, note, created_at)
          VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """, ("TEMP", customer_id if customer_id > 0 else None, customer_name, customer_address, customer_phone, customer_email, delivery_date, delivery_time or None, "new", note, created_at))
        oid = cur.lastrowid

        order_no = make_order_no(oid, created_at)
        cur.execute("UPDATE orders SET order_no=? WHERE id=?", (order_no, oid))

        for pid, qty in items:
            cur.execute("SELECT sku FROM products WHERE id=?", (pid,))
            p = cur.fetchone()
            if not p:
                continue
            sku = p["sku"]
            cur.execute("""
              INSERT INTO order_items(order_id, product_id, sku, qty, created_at)
              VALUES(?,?,?,?,?)
            """, (oid, pid, sku, qty, now_iso()))

        c.commit()
        c.close()

    try:
        normalize_temp_order_numbers()
    except Exception:
        pass

    # Każde nowe zamówienie betonu od razu dostaje roboczy dokument WZ.
    # Pracownik potwierdza wydanie później już z poziomu tego dokumentu.
    c = conn()
    try:
        wz_id, wz_item_ids = create_wz_from_order(c, oid, destination=customer_address)
        c.commit()
    except Exception as exc:
        c.rollback()
        app.logger.exception("Nie udało się utworzyć automatycznego WZ dla zamówienia %s", oid)
        return render_template_string("""
          <h1>Nie udało się utworzyć dokumentu WZ</h1>
          <p>Zamówienie zostało zapisane, ale roboczy dokument WZ nie powstał.</p>
          <p><strong>Szczegół techniczny:</strong> {{ error }}</p>
          <p><a href="{{ url_for('order_view', order_id=order_id) }}">Wróć do zamówienia</a></p>
        """, error=str(exc), order_id=oid), 500
    finally:
        c.close()
    if supabase_enabled():
        try:
            sync_local_rows_to_supabase("wz_documents", "id", [wz_id])
            sync_local_rows_to_supabase("wz_items", "id", wz_item_ids)
        except Exception as exc:
            app.logger.exception("Nie udało się zapisać automatycznego WZ %s w Supabase", wz_id)
            return render_template_string("""
              <h1>WZ nie zostało zapisane w chmurze</h1>
              <p>Zamówienie i robocze WZ są zapisane lokalnie, lecz Supabase odrzucił zapis WZ.</p>
              <p><strong>Szczegół techniczny:</strong> {{ error }}</p>
              <p>Nie twórz ponownie tego samego zamówienia. Prześlij ten komunikat — wskaże brakującą kolumnę lub tabelę w Supabase.</p>
              <p><a href="{{ url_for('beton.wz_view', wz_id=wz_id) }}">Otwórz utworzone WZ</a></p>
            """, error=str(exc), wz_id=wz_id), 500
    return redirect(url_for("beton.wz_view", wz_id=wz_id))

@app.get("/orders/<int:order_id>")
def order_view(order_id):
    maybe_pull_shared_from_supabase()
    try:
        link_orders_to_customers_by_email(sync_remote=True)
    except Exception:
        pass
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)

    cur.execute("""
      SELECT oi.*, p.model, p.ean, p.name,
             COALESCE(s.qty, 0) AS stock_qty,
             COALESCE(pr.net_price, 0) AS net_price,
             COALESCE(pr.gross_price, 0) AS gross_price,
             (oi.qty * COALESCE(pr.net_price, 0)) AS line_value_net,
             (oi.qty * COALESCE(pr.gross_price, 0)) AS line_value_gross,
             COALESCE(s.qty,0) AS stock,
             COALESCE((
                SELECT SUM(ci.qty)
                FROM material_order_items ci
                JOIN material_orders cp ON cp.id=ci.package_id
                WHERE ci.product_id=oi.product_id
                  AND cp.status IN ('planned', 'ordered', 'shipped')
             ), 0) AS in_delivery
      FROM order_items oi
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN stock s ON s.product_id=p.id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      WHERE oi.order_id=?
      ORDER BY oi.id
    """, (order_id,))
    items = [dict(r) for r in cur.fetchall()]

    for it in items:
        it["in_delivery_available"] = int(it.get("in_delivery", 0))
        it["delivery_used"] = 0
        it["line_shortage"] = 0

    if o["status"] in ("new", "packed", "confirmed", "in_delivery"):
        cur.execute("SELECT id FROM orders WHERE status IN ('new','packed','confirmed','in_delivery') AND id<=? ORDER BY id", (order_id,))
        scoped_order_ids = [int(r["id"]) for r in cur.fetchall()]
        if scoped_order_ids:
            sph = ",".join(["?"] * len(scoped_order_ids))
            cur.execute(f"""
              SELECT oi.id, oi.order_id, oi.product_id, oi.qty
              FROM order_items oi
              WHERE oi.order_id IN ({sph})
              ORDER BY oi.order_id, oi.id
            """, tuple(scoped_order_ids))
            seq_items = cur.fetchall()

            product_ids = {int(r["product_id"]) for r in seq_items}
            pool_stock = {}
            pool_delivery = {}
            if product_ids:
                pph = ",".join(["?"] * len(product_ids))
                cur.execute(f"""
                  SELECT p.id AS product_id,
                         COALESCE(s.qty,0) AS stock_qty,
                         COALESCE((
                           SELECT SUM(ci.qty)
                           FROM material_order_items ci
                           JOIN material_orders cp ON cp.id=ci.package_id
                           WHERE ci.product_id=p.id
                             AND cp.status IN ('planned', 'ordered', 'shipped')
                         ),0) AS in_delivery_qty
                  FROM products p
                  LEFT JOIN stock s ON s.product_id=p.id
                  WHERE p.id IN ({pph})
                """, tuple(product_ids))
                for pr in cur.fetchall():
                    pid = int(pr["product_id"])
                    pool_stock[pid] = int(pr["stock_qty"])
                    pool_delivery[pid] = int(pr["in_delivery_qty"])

            item_alloc = {}
            for sr in seq_items:
                pid = int(sr["product_id"])
                need = int(sr["qty"])

                stock_now = pool_stock.get(pid, 0)
                from_stock = min(stock_now, need)
                pool_stock[pid] = stock_now - from_stock
                need_after_stock = need - from_stock

                delivery_now = pool_delivery.get(pid, 0)
                from_delivery = min(delivery_now, need_after_stock)
                pool_delivery[pid] = delivery_now - from_delivery
                shortage = need_after_stock - from_delivery

                if int(sr["order_id"]) == order_id:
                    item_alloc[int(sr["id"])] = {
                        "in_delivery_available": from_delivery,
                        "delivery_used": from_delivery,
                        "line_shortage": shortage,
                    }

            for it in items:
                al = item_alloc.get(int(it["id"]))
                if al:
                    it.update(al)

    cur.execute("SELECT id, sku, model, name FROM products ORDER BY sku LIMIT 5000")
    products_rows = cur.fetchall()
    cur.execute("""
      SELECT w.id,w.wz_no,w.status,w.invoice_id,i.invoice_no,
             t.id AS transport_id,t.transport_no,t.status AS transport_status,
             (SELECT p.id FROM delivery_photos p
                WHERE p.transport_id=t.id AND p.deleted_at IS NULL
                ORDER BY p.created_at DESC LIMIT 1) AS photo_id
      FROM wz_documents w
      LEFT JOIN invoices i ON i.id=w.invoice_id
      LEFT JOIN transports t ON t.wz_id=w.id AND t.deleted_at IS NULL
      WHERE w.order_id=? AND w.deleted_at IS NULL
      ORDER BY w.id DESC
    """, (order_id,))
    wz_documents = [dict(r) for r in cur.fetchall()]
    c.close()

    order_url = build_public_url(url_for("order_view", order_id=order_id))

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">{{ order_display_no(o['id'], o['created_at'], o['order_no'], o['note']) }}</h1>
          <div class="right flex">
            <a class="btn" href="{{ url_for('orders') }}">â† Lista</a>
            <a class="btn primary" href="{{ url_for('beton.wz_new', order_id=o['id']) }}">Nowe WZ</a>
            {% if false %}
            <form method="post" action="{{ url_for('order_confirmation_resend', order_id=o['id']) }}">
              <button class="btn" type="submit">Wyślij ponownie potwierdzenie</button>
            </form>
            <form method="post" action="{{ url_for('order_status_update', order_id=o['id']) }}" class="flex">
                <select name="status" style="width:190px;">
                  <option value="new" {% if o['status'] in ['new','pending','unconfirmed'] %}selected{% endif %}>Niepotwierdzone</option>
                  <option value="confirmed" {% if o['status']=='confirmed' %}selected{% endif %}>Potwierdzone</option>
                  <option value="in_delivery" {% if o['status'] in ['packed','in_delivery'] %}selected{% endif %}>W dostawie</option>
                  <option value="issued" {% if o['status']=='issued' %}selected{% endif %}>Zrealizowane</option>
                </select>
                <button class="btn" type="submit">ZmieĹ„ status</button>
              </form>
            {% endif %}
              {% if locked %}
                <span class="badge">Wydane z magazynu</span>
              {% endif %}
              <form method="post" action="{{ url_for('order_delete', order_id=o['id']) }}" onsubmit="return confirm('UsunÄ…Ä‡ zamĂłwienie?')">
                <button class="btn danger" type="submit">UsuĹ„ zamĂłwienie</button>
              </form>
          </div>
        </div>
        <div class="muted" style="margin-top:6px;">{{ o['created_at'] }}</div>
        {% if request.args.get('confirmation_sent') == '1' %}
          <div class="hint" style="margin-top:10px;">Potwierdzenie zamówienia zostało wysłane ponownie.</div>
        {% elif request.args.get('confirmation_error') %}
          <div class="hint" style="margin-top:10px; border-color:#fecaca; background:#fff1f2;">Nie udało się wysłać potwierdzenia: {{ request.args.get('confirmation_error') }}</div>
        {% endif %}
      </div>

      <div class="card">
        <h2>Realizacja dokumentów</h2>
        {% if wz_documents %}
        <table>
          <thead><tr><th>WZ</th><th>Stan</th><th>Transport</th><th>Dokumenty i akcje</th></tr></thead>
          <tbody>{% for wz in wz_documents %}<tr>
            <td><a href="{{ url_for('beton.wz_view',wz_id=wz.id) }}"><b>{{ wz.wz_no }}</b></a></td>
            <td><span class="badge">{{ wz.status }}</span></td>
            <td>{% if wz.transport_id %}<a href="{{ url_for('beton.transport_view',transport_id=wz.transport_id) }}">{{ wz.transport_no }}</a>{% else %}<span class="muted">Nie przypisano</span>{% endif %}</td>
            <td class="flex">
              <a class="btn" href="{{ url_for('beton.wz_view',wz_id=wz.id) }}">Otwórz WZ</a>
              {% if wz.photo_id %}<a class="btn" href="{{ url_for('beton.photo_download',photo_id=wz.photo_id) }}">Pobierz zdjęcie WZ</a>{% endif %}
              {% if wz.status=='ready_invoice' %}<a class="btn primary" href="{{ url_for('order_invoice',order_id=o['id'],wz_id=wz.id) }}">Wystaw fakturę VAT</a>
              {% elif wz.invoice_id %}<a class="btn" href="{{ url_for('invoice_download_admin',invoice_id=wz.invoice_id) }}">Pobierz fakturę {{ wz.invoice_no }}</a>{% endif %}
            </td>
          </tr>{% endfor %}</tbody>
        </table>
        {% else %}
          <div class="muted">Nie ma jeszcze dokumentu WZ dla tego zamówienia.</div>
        {% endif %}
      </div>

      <div class="row">
        <div class="card">
          <h2>ZamawiajÄ…cy</h2>
          <div><b>{{ o['customer_name'] }}</b></div>
          <div class="muted" style="white-space:pre-line; margin-top:6px;">{{ o['customer_address'] or "-" }}</div>
          <div class="muted" style="margin-top:6px;">Tel: {{ o['customer_phone'] or "-" }}</div>
          <div class="muted">Email: {{ o['customer_email'] or "-" }}</div>
          <div class="line"></div>
          <div class="muted small">Kod zamĂłwienia do skanowania: <b>{{ canonical_order_no(o['id'], o['created_at'], o['order_no']) }}</b></div>
          <div class="muted small" style="margin-top:10px;">QR jest uĹĽywany do etykiety 30x50 i skanowania zamĂłwienia.</div>
        </div>

        <div class="card">
          <h2>Notatka</h2>
          <div>{{ o['note'] or "-" }}</div>
          <div class="line"></div>
          <div class="hint">
            <b>Wydaj z magazynu</b> odejmie iloĹ›ci z magazynu, ale nie zmieni automatycznie statusu klienta na â€žZrealizowaneâ€ť.<br>
            JeĹ›li brakuje stanu, pozycja moĹĽe byÄ‡ realizowana z <b>towaru w drodze z materiałów</b> (kolumna â€žW dostawieâ€ť poniĹĽej).
          </div>
        </div>
      </div>

      {% if not locked %}
      <div class="card">
        <h2>Dodaj produkt do zamĂłwienia</h2>
        <form method="post" action="{{ url_for('order_item_add', order_id=o['id']) }}" class="items-row">
          <div>
            <select name="product_id" required>
              <option value="">-- wybierz produkt --</option>
              {% for p in products %}
                <option value="{{ p['id'] }}">{{ p['sku'] }}{% if p['model'] %} â€˘ {{ p['model'] }}{% endif %}{% if p['name'] %} â€˘ {{ p['name'] }}{% endif %}</option>
              {% endfor %}
            </select>
          </div>
          <div>
            <input name="qty" value="1" required>
          </div>
          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Dodaj</button>
          </div>
        </form>
      </div>
      {% endif %}

      <div class="card">
        <h2>Pozycje</h2>
        <table>
          <thead>
            <tr><th>SKU</th><th>Model / Nazwa</th><th>IloĹ›Ä‡</th><th>Cena netto</th><th>Cena brutto</th><th>WartoĹ›Ä‡ netto</th><th>WartoĹ›Ä‡ brutto</th><th>Stan teraz</th><th>W dostawie (dostÄ™pne)</th><th>Realizacja</th><th>Akcje</th></tr>
          </thead>
          <tbody>
            {% set ns = namespace(total_net=0, total_gross=0) %}
            {% for it in items %}
              {% set ns.total_net = ns.total_net + it['line_value_net'] %}
              {% set ns.total_gross = ns.total_gross + it['line_value_gross'] %}
              <tr>
                <td><b>{{ it['sku'] }}</b></td>
                <td>
                  {{ it['model'] or "" }}
                  {% if it['name'] %}<div class="muted small">{{ it['name'] }}</div>{% endif %}
                  {% if it['ean'] %}<div class="muted small">EAN: {{ it['ean'] }}</div>{% endif %}
                </td>
                <td>
                  {% if locked %}
                    <span class="badge">{{ it['qty'] }}</span>
                  {% else %}
                    <form method="post" action="{{ url_for('order_item_update', order_id=o['id'], item_id=it['id']) }}" class="flex">
                      <input name="qty" value="{{ it['qty'] }}" style="width:90px;">
                      <button class="btn" type="submit">ZmieĹ„</button>
                    </form>
                  {% endif %}
                </td>
                <td><span class="badge">{{ "%.2f"|format(it['net_price']) }} PLN</span></td>
                <td><span class="badge">{{ "%.2f"|format(it['gross_price']) }} PLN</span></td>
                <td><span class="badge">{{ "%.2f"|format(it['line_value_net']) }} PLN</span></td>
                <td><span class="badge">{{ "%.2f"|format(it['line_value_gross']) }} PLN</span></td>
                <td><span class="badge">{{ it['stock'] }}</span></td>
                <td><span class="badge">{{ it['in_delivery_available'] }}</span></td>
                <td>
                  {% if it['line_shortage'] <= 0 and it['delivery_used'] == 0 %}
                    <span class="badge">Z magazynu</span>
                  {% elif it['line_shortage'] <= 0 %}
                    <span class="badge">CzÄ™Ĺ›Ä‡ / caĹ‚oĹ›Ä‡ z materiałów</span>
                  {% else %}
                    <span class="badge">Brak towaru</span>
                  {% endif %}
                </td>
                <td>
                  {% if not locked %}
                    <form method="post" action="{{ url_for('order_item_delete', order_id=o['id'], item_id=it['id']) }}" onsubmit="return confirm('UsunÄ…Ä‡ pozycjÄ™?')">
                      <button class="btn danger" type="submit">UsuĹ„</button>
                    </form>
                  {% else %}
                    <span class="muted">PodglÄ…d</span>
                  {% endif %}
                </td>
              </tr>
            {% endfor %}
            {% if items %}
              <tr>
                <td colspan="5" style="text-align:right;"><b>Suma netto:</b></td>
                <td><span class="badge"><b>{{ "%.2f"|format(ns.total_net) }} PLN</b></span></td>
                <td colspan="5"></td>
              </tr>
              <tr>
                <td colspan="6" style="text-align:right;"><b>Suma brutto:</b></td>
                <td><span class="badge"><b>{{ "%.2f"|format(ns.total_gross) }} PLN</b></span></td>
                <td colspan="4"></td>
              </tr>
            {% else %}
              <tr><td colspan="11" class="muted">Brak pozycji w zamĂłwieniu.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title=canonical_order_no(o["id"], o["created_at"], o["order_no"]), base_url=BASE_URL, db_path=DB_PATH, o=o, items=items, order_url=order_url, products=products_rows, wz_documents=wz_documents, locked=(int(o["warehouse_issued"] or 0)==1), order_status_label=order_status_label, order_status_css=order_status_css, canonical_order_no=canonical_order_no)


@app.post("/orders/<int:order_id>/confirmation/resend")
def order_confirmation_resend(order_id):
    try:
        maybe_pull_shared_from_supabase(force=True)
    except Exception:
        pass
    result = _send_saved_order_confirmation(order_id, force=True)
    if result.get("ok"):
        return redirect(url_for("order_view", order_id=order_id, confirmation_sent="1"))
    return redirect(url_for("order_view", order_id=order_id, confirmation_error=norm(result.get("error")) or "Nieznany błąd"))

@app.post("/orders/<int:order_id>/items/add")
def order_item_add(order_id):
    product_id = to_int(request.form.get("product_id"), 0)
    qty = to_float(request.form.get("qty"), 0.0)
    if product_id <= 0 or qty <= 0:
        return "NieprawidĹ‚owy produkt lub iloĹ›Ä‡", 400

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT status, warehouse_issued FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)
    if INVENTORY_AUTOMATION_ENABLED and int(o["warehouse_issued"] or 0) == 1:
        c.close()
        return "ZamĂłwienie wydane z magazynu jest tylko do podglÄ…du", 400

    cur.execute("SELECT sku FROM products WHERE id=?", (product_id,))
    p = cur.fetchone()
    if not p:
        c.close()
        return "Brak produktu", 404

    if supabase_enabled():
        created_item = supabase_insert_row("order_items", {
            "order_id": order_id,
            "product_id": product_id,
            "sku": p["sku"],
            "qty": qty,
            "created_at": now_iso(),
        })
        if not created_item or "id" not in created_item:
            c.close()
            return "Nie udaĹ‚o siÄ™ dodaÄ‡ pozycji do Supabase", 500
        cur.execute(
            "INSERT INTO order_items(id, order_id, product_id, sku, qty, created_at) VALUES (?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET order_id=excluded.order_id, product_id=excluded.product_id, sku=excluded.sku, qty=excluded.qty, created_at=excluded.created_at",
            (int(created_item["id"]), order_id, product_id, p["sku"], qty, created_item.get("created_at") or now_iso())
        )
    else:
        cur.execute("""
          INSERT INTO order_items(order_id, product_id, sku, qty, created_at)
          VALUES(?,?,?,?,?)
        """, (order_id, product_id, p["sku"], qty, now_iso()))
    c.commit()
    c.close()
    return redirect(url_for("order_view", order_id=order_id))

@app.post("/orders/<int:order_id>/items/<int:item_id>/update")
def order_item_update(order_id, item_id):
    qty = to_float(request.form.get("qty"), 0.0)
    if qty <= 0:
        return "IloĹ›Ä‡ musi byÄ‡ > 0", 400
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT status, warehouse_issued FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)
    if INVENTORY_AUTOMATION_ENABLED and int(o["warehouse_issued"] or 0) == 1:
        c.close()
        return "ZamĂłwienie wydane z magazynu jest tylko do podglÄ…du", 400
    invoiced_qty = int(invoiced_qty_by_order_item_ids([item_id]).get(int(item_id)) or 0)
    if qty < invoiced_qty:
        c.close()
        return f"Nie moĹĽesz ustawiÄ‡ iloĹ›ci poniĹĽej juĹĽ zafakturowanej ({invoiced_qty} szt.)", 400
    cur.execute("UPDATE order_items SET qty=? WHERE id=? AND order_id=?", (qty, item_id, order_id))
    c.commit()
    c.close()

    if supabase_enabled():
        supabase_update_rows("order_items", {"qty": qty}, {"id": item_id})

    return redirect(url_for("order_view", order_id=order_id))


@app.post("/orders/<int:order_id>/items/<int:item_id>/delete")
def order_item_delete(order_id, item_id):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT status, warehouse_issued FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)
    if int(o["warehouse_issued"] or 0) == 1:
        c.close()
        return "ZamĂłwienie wydane z magazynu jest tylko do podglÄ…du", 400
    invoiced_qty = int(invoiced_qty_by_order_item_ids([item_id]).get(int(item_id)) or 0)
    if invoiced_qty > 0:
        c.close()
        return f"Nie moĹĽesz usunÄ…Ä‡ pozycji, bo jest juĹĽ zafakturowana ({invoiced_qty} szt.)", 400

    if supabase_enabled():
        supabase_delete_rows("order_items", {"id": item_id})

    cur.execute("DELETE FROM order_items WHERE id=? AND order_id=?", (item_id, order_id))
    c.commit()
    c.close()
    return redirect(url_for("order_view", order_id=order_id))


@app.post("/orders/<int:order_id>/delete")
def order_delete(order_id):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id, warehouse_issued FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)

    cur.execute("SELECT product_id, qty FROM order_items WHERE order_id=?", (order_id,))
    items = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT id FROM invoices WHERE order_id=?", (order_id,))
    invoice_ids = [int(r["id"]) for r in cur.fetchall()]

    changed_product_ids = []
    if INVENTORY_AUTOMATION_ENABLED and int(o["warehouse_issued"] or 0) == 1:
        for it in items:
            pid = int(it["product_id"])
            qty = int(it["qty"])
            cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
            cur.execute("UPDATE stock SET qty = qty + ? WHERE product_id=?", (qty, pid))
            changed_product_ids.append(pid)

    if invoice_ids:
        cur.execute("DELETE FROM invoice_allocations WHERE invoice_id IN (" + ",".join(["?"] * len(invoice_ids)) + ")", tuple(invoice_ids))
        cur.execute("DELETE FROM invoice_meta WHERE invoice_id IN (" + ",".join(["?"] * len(invoice_ids)) + ")", tuple(invoice_ids))
    cur.execute("DELETE FROM invoice_allocations WHERE order_id=?", (order_id,))
    cur.execute("DELETE FROM invoices WHERE order_id=?", (order_id,))
    cur.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
    cur.execute("DELETE FROM orders WHERE id=?", (order_id,))
    c.commit()
    c.close()

    if supabase_enabled():
        try:
            supabase_delete_rows("invoice_allocations", {"order_id": order_id})
            for iid in invoice_ids:
                supabase_delete_rows("invoice_allocations", {"invoice_id": iid})
                supabase_delete_rows("invoice_meta", {"invoice_id": iid})
                supabase_delete_rows("invoices", {"id": iid})
            supabase_delete_rows("order_items", {"order_id": order_id})
            supabase_delete_rows("orders", {"id": order_id})
            if changed_product_ids:
                sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
        except Exception:
            pass

    return redirect(url_for("orders"))

@app.post("/orders/<int:order_id>/status")
def order_status_update(order_id):
    new_status = norm(request.form.get("status")).lower()
    allowed = {"new", "confirmed", "in_delivery", "issued"}
    if new_status not in allowed:
        return "NieprawidĹ‚owy status", 400

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id, order_no, status, created_at, warehouse_issued FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)

    changed_product_ids = []
    warehouse_issued = int(o["warehouse_issued"] or 0)

    # Opcjonalny, starszy tryb magazynowy. W Beton Łagów jest domyślnie wyłączony:
    # wydanie betonu dokumentuje WZ/transport, ale nie rozchodowuje automatycznie
    # materiałów produkcyjnych (cementu, żwiru ani piasku).
    if new_status == "in_delivery" and warehouse_issued == 0 and INVENTORY_AUTOMATION_ENABLED:
        cur.execute("""
          SELECT oi.product_id, oi.qty
          FROM order_items oi
          WHERE oi.order_id=?
          ORDER BY oi.id
        """, (order_id,))
        items = cur.fetchall()

        for it in items:
            pid = int(it["product_id"])
            qty = int(it["qty"])
            cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
            cur.execute("UPDATE stock SET qty = qty - ? WHERE product_id=?", (qty, pid))
            changed_product_ids.append(pid)

        warehouse_issued = 1

    # To tylko znacznik przebiegu dokumentu WZ, nie ruch stanu magazynowego.
    elif new_status == "in_delivery" and warehouse_issued == 0:
        warehouse_issued = 1

    cur.execute(
        "UPDATE orders SET status=?, warehouse_issued=? WHERE id=?",
        (new_status, warehouse_issued, order_id)
    )
    c.commit()
    c.close()

    if supabase_enabled():
        try:
            supabase_update_rows(
                "orders",
                {"status": new_status, "warehouse_issued": warehouse_issued},
                {"id": order_id}
            )
        except Exception:
            pass

        if changed_product_ids:
            try:
                sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
            except Exception:
                pass

    return redirect(url_for("order_view", order_id=order_id))


@app.get("/orders/<int:order_id>/issue")
def order_issue(order_id):
    # Stara akcja wyĹ‚Ä…czona. Wydanie dzieje siÄ™ teraz przy zmianie statusu na "W dostawie".
    return redirect(url_for("order_view", order_id=order_id))


@app.route("/orders/<int:order_id>/invoice", methods=["GET", "POST"])
def order_invoice(order_id):
    maybe_pull_shared_from_supabase()
    wz_id = to_int(request.args.get("wz_id") or request.form.get("wz_id"), 0)
    if wz_id:
        wz_check = conn()
        wz_row = wz_check.execute("SELECT id,order_id,status FROM wz_documents WHERE id=? AND deleted_at IS NULL", (wz_id,)).fetchone()
        wz_check.close()
        if not wz_row or int(wz_row["order_id"]) != int(order_id):
            abort(404)
        if wz_row["status"] not in {"ready_invoice", "invoiced"}:
            return "WZ nie jest jeszcze gotowe do fakturowania.", 409
    sent_invoice_id = to_int(request.args.get("invoice_id"), 0) if norm(request.args.get("sent")) == "1" else 0
    if sent_invoice_id:
        meta = load_invoice_meta(sent_invoice_id) or {}
        upsert_invoice_meta(
            sent_invoice_id,
            meta.get("pdf_path", ""),
            meta.get("invoice_items_json", ""),
            sent_to_client=1,
            seen_by_client=int(meta.get("seen_by_client") or 0),
            seen_at=meta.get("seen_at"),
            payment_reminder=int(meta.get("payment_reminder") or 0),
            paid=int(meta.get("paid") or 0),
            paid_at=meta.get("paid_at")
        )
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)

    related_orders = [dict(o)] if norm(o["status"]).lower() in CURRENT_ORDER_STATUSES else []
    customer_email_key = _email_key(o["customer_email"])
    if customer_email_key:
        status_ph = ",".join(["?"] * len(CURRENT_ORDER_STATUSES))
        cur.execute(f"""
          SELECT *
          FROM orders
          WHERE LOWER(COALESCE(customer_email,'')) = ?
            AND LOWER(COALESCE(status,'')) IN ({status_ph})
          ORDER BY created_at DESC, id DESC
        """, (customer_email_key, *sorted(CURRENT_ORDER_STATUSES)))
        related_orders = [dict(r) for r in cur.fetchall()]

    related_order_ids = [int(r["id"]) for r in related_orders] or [-1]
    related_order_by_id = {int(r["id"]): r for r in related_orders}
    order_ph = ",".join(["?"] * len(related_order_ids))

    cur.execute(f"""
      SELECT oi.*, p.model, p.name,
             oo.order_no AS source_order_no,
             oo.created_at AS source_order_created_at,
             oo.note AS source_order_note,
             COALESCE(pr.net_price, 0) AS net_price,
             COALESCE(pr.gross_price, 0) AS gross_price,
             (oi.qty * COALESCE(pr.net_price, 0)) AS line_value_net,
             (oi.qty * COALESCE(pr.gross_price, 0)) AS line_value_gross
      FROM order_items oi
      JOIN orders oo ON oo.id=oi.order_id
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      WHERE oi.order_id IN ({order_ph})
      ORDER BY oo.created_at DESC, oo.id DESC, oi.id
    """, related_order_ids)
    items = [dict(r) for r in cur.fetchall()]
    if wz_id:
        cur.execute("SELECT order_item_id,qty_issued,qty_planned FROM wz_items WHERE wz_id=?", (wz_id,))
        wz_qty = {int(r["order_item_id"]): float(r["qty_issued"] if r["qty_issued"] is not None else r["qty_planned"]) for r in cur.fetchall()}
        items = [it for it in items if int(it["id"]) in wz_qty]
    invoiced_by_item = invoiced_qty_by_order_item_ids([int(it["id"]) for it in items])
    for it in items:
        source_order = related_order_by_id.get(int(it.get("order_id") or 0), {})
        ordered_qty = int(it.get("qty") or 0)
        done_qty = int(invoiced_by_item.get(int(it["id"])) or 0)
        it["source_order_no"] = order_display_no(
            source_order.get("id") or it.get("order_id"),
            source_order.get("created_at") or it.get("source_order_created_at"),
            source_order.get("order_no") or it.get("source_order_no"),
            source_order.get("note") or it.get("source_order_note") or ""
        )
        it["source_order_note"] = source_order.get("note") or it.get("source_order_note") or ""
        it["ordered_qty"] = ordered_qty
        it["invoiced_qty"] = done_qty
        it["remaining_qty"] = max(0, ordered_qty - done_qty)
        if wz_id:
            it["remaining_qty"] = min(it["remaining_qty"], int(wz_qty.get(int(it["id"]), 0)))

    cur.execute("SELECT * FROM company_profile WHERE id=1")
    company = cur.fetchone()

    customer_row = None
    if o["customer_id"]:
        cur.execute("SELECT * FROM customers WHERE id=?", (o["customer_id"],))
        customer_row = cur.fetchone()
    if not customer_row:
        cur.execute("SELECT * FROM customers WHERE name=? ORDER BY id DESC LIMIT 1", (o["customer_name"],))
        customer_row = cur.fetchone()

    cur.execute(f"""
      SELECT
        i.*,
        m.invoice_id AS meta_invoice_id,
        COALESCE(m.pdf_path,'') AS pdf_path,
        COALESCE(m.sent_to_client,0) AS sent_to_client,
        COALESCE(m.seen_by_client,0) AS seen_by_client,
        COALESCE(m.payment_reminder,0) AS payment_reminder,
        COALESCE(m.paid,0) AS paid,
        COALESCE(m.paid_at,'') AS paid_at,
        COALESCE(m.seen_at,'') AS seen_at,
        COALESCE(m.invoice_items_json,'') AS invoice_items_json
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      WHERE i.order_id IN ({order_ph})
      ORDER BY i.id DESC
    """, related_order_ids)
    invoice_rows = [dict(r) for r in cur.fetchall()]
    c.close()

    default_issue = app_now().strftime("%Y-%m-%d")
    buyer_address_source = customer_row["address"] if customer_row and customer_row["address"] else (o["customer_address"] or "")
    st, pc, city = split_address(buyer_address_source)
    buyer_tax_no = customer_row["nip"] if customer_row and customer_row["nip"] else ""
    buyer_address_default = "\n".join([x for x in [st, f"{pc} {city}".strip()] if x]).strip()

    msg = ""
    if request.args.get("generated") == "1":
        msg = "Faktura zostaĹ‚a zapisana."
    if request.args.get("sent") == "1":
        msg = "Faktura zostaĹ‚a udostÄ™pniona klientowi."
    if request.args.get("deleted") == "1":
        msg = "Faktura zostaĹ‚a usuniÄ™ta."
    if request.args.get("deleted") == "1":
        msg = "Faktura zostaĹ‚a usuniÄ™ta."

    if request.method == "GET":
        data = {
            "invoice_no": next_invoice_no(default_issue),
            "place": "KotuszĂłw",
            "issue_date": default_issue,
            "sell_date": default_issue,
            "payment_type": "gotowka",
            "payment_to": (app_now() + timedelta(days=30)).strftime("%Y-%m-%d"),
            "buyer_name": o["customer_name"] or "",
            "buyer_tax_no": buyer_tax_no,
            "buyer_address": buyer_address_default,
            "buyer_country": "PL",
            "buyer_email": o["customer_email"] or "",
            "buyer_phone": o["customer_phone"] or "",
            "discount_percent": "0",
        }
    else:
        data = {k: norm(request.form.get(k)) for k in [
            "invoice_no", "place", "issue_date", "sell_date", "payment_type", "payment_to",
            "buyer_name", "buyer_tax_no", "buyer_address", "buyer_country",
            "buyer_email", "buyer_phone", "discount_percent"
        ]}
        st, pc, city = split_address(data.get("buyer_address", ""))
        data["buyer_street"] = st
        data["buyer_post_code"] = pc
        data["buyer_city"] = city
        if not data["invoice_no"]:
            data["invoice_no"] = next_invoice_no(data["issue_date"] or default_issue)
        if not data["issue_date"]:
            data["issue_date"] = default_issue
        if not data["sell_date"]:
            data["sell_date"] = data["issue_date"]

        invoice_items = prepare_invoice_items(items, request.form)
        wz_mismatch = False
        if wz_id:
            selected = {int(x.get("order_item_id") or x.get("id") or 0): to_float(x.get("qty"), 0.0) for x in invoice_items}
            expected = {int(item_id): to_float(qty, 0.0) for item_id, qty in wz_qty.items()}
            wz_mismatch = selected != expected
        existing_invoice_id = invoice_no_exists(data["invoice_no"])
        if existing_invoice_id:
            msg = f"Faktura o takim numerze już istnieje! Numer: {data['invoice_no']}. Wybierz inny numer faktury."
        elif wz_mismatch:
            msg = "Faktura z WZ musi zawierać wszystkie pozycje i dokładne ilości wydane na tym WZ."
        elif not invoice_items:
            msg = "Faktura musi zawieraÄ‡ co najmniej jednÄ… pozycjÄ™."
        else:
            pdf_path, total_net, total_gross = generate_order_invoice_pdf(o, invoice_items, data)
            c = conn()
            cur = c.cursor()
            cur.execute("""
              INSERT INTO invoices(order_id, invoice_no, issue_date, sell_date, payment_type, payment_to,
                                   buyer_name, buyer_tax_no, buyer_street, buyer_post_code, buyer_city, buyer_country,
                                   buyer_email, buyer_phone, total_net, total_gross, created_at)
              VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                order_id, data["invoice_no"], data["issue_date"], data["sell_date"], data["payment_type"], data["payment_to"],
                data["buyer_name"], data["buyer_tax_no"], data["buyer_street"], data["buyer_post_code"], data["buyer_city"], data["buyer_country"],
                data["buyer_email"], data["buyer_phone"], total_net, total_gross, now_iso()
            ))
            invoice_id = cur.lastrowid
            if not invoice_id:
                cur.execute("SELECT id FROM invoices WHERE invoice_no=? LIMIT 1", (data["invoice_no"],))
                rr = cur.fetchone()
                invoice_id = int(rr["id"]) if rr else 0
            c.commit()
            c.close()
            if wz_id:
                c = conn()
                c.execute("""UPDATE wz_documents SET invoice_id=?,status='invoiced',invoiced_by=?,invoiced_at=?
                             WHERE id=? AND status='ready_invoice'""",
                          (invoice_id, session.get("display_name") or session.get("username") or "Księgowość", now_iso(), wz_id))
                c.execute("UPDATE transports SET invoice_id=?,updated_by=?,updated_at=? WHERE wz_id=? AND deleted_at IS NULL",
                          (invoice_id, session.get("display_name") or session.get("username") or "Księgowość", now_iso(), wz_id))
                c.commit(); c.close()
            stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, data["invoice_no"], pdf_path)
            upsert_invoice_meta(invoice_id, stored_pdf_path, json.dumps(invoice_items, ensure_ascii=False), sent_to_client=None)
            allocation_ids = replace_invoice_allocations(invoice_id, invoice_items)
            touched_order_ids = [int(x.get("source_order_id") or x.get("order_id") or 0) for x in invoice_items]
            completed_order_ids, changed_product_ids = finalize_fully_invoiced_orders(touched_order_ids)
            if supabase_enabled():
                try:
                    sync_local_rows_to_supabase("invoices", "id", [invoice_id])
                except Exception:
                    pass
                try:
                    sync_invoice_meta_to_supabase(invoice_id)
                except Exception:
                    pass
                try:
                    sync_local_rows_to_supabase("invoice_allocations", "id", allocation_ids)
                except Exception:
                    pass
                if completed_order_ids:
                    try:
                        sync_local_rows_to_supabase("orders", "id", completed_order_ids)
                    except Exception:
                        pass
                if changed_product_ids:
                    try:
                        sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
                    except Exception:
                        pass
            return redirect(url_for("invoices", generated="1", invoice_id=invoice_id))

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Faktura z pozycji klienta: {{ o['customer_name'] or o['customer_email'] }}</h1>
          <a class="btn right" href="{{ url_for('order_view', order_id=o['id']) }}">â† SzczegĂłĹ‚y</a>
        </div>
        {% if msg %}
          <div class="hint" style="margin-top:10px;">{{ msg }}</div>
        {% endif %}
      </div>

      <div class="card">
        <form method="post" class="row">
          <div><label class="muted small">Numer faktury</label><input name="invoice_no" value="{{ d['invoice_no'] }}" required></div>
          <div><label class="muted small">Miejsce</label><input name="place" value="{{ d['place'] }}"></div>
          <div><label class="muted small">Data wystawienia</label><input name="issue_date" type="date" value="{{ d['issue_date'] }}"></div>
          <div><label class="muted small">Data sprzedaĹĽy</label><input name="sell_date" type="date" value="{{ d['sell_date'] }}"></div>
          <div><label class="muted small">Forma pĹ‚atnoĹ›ci</label>
            <select name="payment_type">
              <option value="gotowka" {% if d['payment_type'] in ['cash','gotowka'] %}selected{% endif %}>gotĂłwka</option>
              <option value="przelew" {% if d['payment_type'] in ['transfer','przelew'] %}selected{% endif %}>przelew</option>
              <option value="karta" {% if d['payment_type'] in ['card','karta'] %}selected{% endif %}>karta</option>
            </select>
          </div>
          <div><label class="muted small">Termin pĹ‚atnoĹ›ci</label><input name="payment_to" type="date" value="{{ d['payment_to'] }}"></div>
          <div><label class="muted small">Rabat %</label><input name="discount_percent" value="{{ d['discount_percent'] or "0" }}"></div>

          <div><label class="muted small">Nabywca</label><input name="buyer_name" value="{{ d['buyer_name'] }}" required></div>
          <div><label class="muted small">NIP nabywcy</label><input name="buyer_tax_no" value="{{ d['buyer_tax_no'] }}"></div>
          <div><label class="muted small">Adres nabywcy</label><textarea name="buyer_address" placeholder="Ulica&#10;Kod pocztowy Miasto">{{ d['buyer_address'] }}</textarea></div>
          <div><label class="muted small">Kraj</label><input name="buyer_country" value="{{ d['buyer_country'] }}"></div>
          <div><label class="muted small">Email</label><input name="buyer_email" value="{{ d['buyer_email'] }}"></div>
          <div><label class="muted small">Telefon</label><input name="buyer_phone" value="{{ d['buyer_phone'] }}"></div>

          <div style="grid-column:1/-1;">
            <h2>Pozycje faktury — wybierz ilości z zamówień klienta</h2>
            <div class="hint" style="margin-bottom:10px;">
              Wpisz ilość tylko przy pozycjach, które idą na fakturę. Zamówienia klienta zostają jako osobne listy/notatki.
            </div>
            <table>
              <thead><tr><th>Zamówienie</th><th>Notatka klienta</th><th>SKU</th><th>Model / Nazwa</th><th>Zamówiono</th><th>Zafakturowano</th><th>Pozostało</th><th>Ilość na fakturze</th><th>Netto/szt</th><th>Brutto/szt</th></tr></thead>
              <tbody>
                {% for it in items %}
                <tr>
                  <td><b>{{ it['source_order_no'] }}</b></td>
                  <td>{{ it['source_order_note'] or '-' }}</td>
                  <td><b>{{ it['sku'] }}</b></td>
                  <td>{{ it['model'] or '' }}{% if it['name'] %}<div class="muted small">{{ it['name'] }}</div>{% endif %}</td>
                  <td>{{ it['ordered_qty'] }}</td>
                  <td>{{ it['invoiced_qty'] }}</td>
                  <td><b>{{ it['remaining_qty'] }}</b></td>
                  <td>
                    <input type="number" min="0" name="invoice_qty_{{ it['id'] }}" value="0" max="{{ it['remaining_qty'] }}" style="width:110px;" {% if it['remaining_qty'] <= 0 %}disabled{% endif %}>
                  </td>
                  <td>{{ "%.2f"|format(it['net_price']) }}</td>
                  <td>{{ "%.2f"|format(it['gross_price']) }}</td>
                </tr>
                {% endfor %}
              </tbody>
            </table>
          </div>

          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Zapisz fakturÄ™ PDF</button>
          </div>
        </form>
      </div>

      <div class="card">
        <h2>Zapisane faktury</h2>
        <table>
          <thead><tr><th>Numer</th><th>Data</th><th>Netto</th><th>Brutto</th><th>Płatność</th><th>Akcje</th></tr></thead>
          <tbody>
            {% for inv in invoice_rows %}
              <tr>
                <td><b>{{ inv['invoice_no'] }}</b></td>
                <td>{{ inv['issue_date'] }}</td>
                <td>{{ "%.2f"|format(inv['total_net']) }}</td>
                <td>{{ "%.2f"|format(inv['total_gross']) }}</td>
                <td>
                  {% if inv['paid'] %}
                    <span class="badge ok">Opłacona</span>
                    {% if inv['paid_at'] %}<div class="muted small">{{ inv['paid_at'] }}</div>{% endif %}
                  {% else %}
                    <span class="badge danger">Nieopłacona</span>
                    {% if inv['payment_reminder'] %}<span class="badge">Przypomnienie aktywne</span>{% endif %}
                  {% endif %}
                </td>
                <td>
                  <div class="flex">
                    <a class="btn" href="{{ url_for('invoice_download_admin', invoice_id=inv['id']) }}" target="_blank">Pobierz PDF</a>
                    <form method="post" action="{{ url_for('invoice_regenerate_admin', invoice_id=inv['id']) }}">
                      <button class="btn" type="submit">Regeneruj PDF</button>
                    </form>
                    {% if not inv['paid'] %}
                      <form method="post" action="{{ url_for('invoice_paid_admin', invoice_id=inv['id']) }}">
                        <input type="hidden" name="next" value="{{ request.full_path }}">
                        <button class="btn ok" type="submit">Oznacz jako opłaconą</button>
                      </form>
                    {% else %}
                      <form method="post" action="{{ url_for('invoice_unpaid_admin', invoice_id=inv['id']) }}">
                        <input type="hidden" name="next" value="{{ request.full_path }}">
                        <button class="btn" type="submit">Oznacz jako nieopłaconą</button>
                      </form>
                    {% endif %}
                    <form method="post" action="{{ url_for('order_invoice_delete', order_id=o['id'], invoice_id=inv['id']) }}" onsubmit="return confirm('UsunÄ…Ä‡ fakturÄ™?')">
                      <button class="btn danger" type="submit">UsuĹ„ fakturÄ™</button>
                    </form>
                  </div>
                </td>
              </tr>
            {% endfor %}
            {% if not invoice_rows %}
              <tr><td colspan="7" class="muted">Brak wystawionych faktur.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Faktura", base_url=BASE_URL, db_path=DB_PATH, o=o, d=data, company=company, items=items, invoice_rows=invoice_rows, msg=msg, canonical_order_no=canonical_order_no)


@app.get("/orders/<int:order_id>/print")
def order_print(order_id):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (order_id,))
    o = cur.fetchone()
    if not o:
        c.close()
        abort(404)

    cur.execute("""
      SELECT oi.sku, oi.qty, p.model, p.name, COALESCE(s.qty,0) AS stock
      FROM order_items oi
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN stock s ON s.product_id=p.id
      WHERE oi.order_id=?
      ORDER BY oi.id
    """, (order_id,))
    items = cur.fetchall()
    c.close()

    in_stock = []
    missing = []
    total_qty = 0
    total_missing_qty = 0
    for it in items:
        need = int(it["qty"])
        have = int(it["stock"])
        row = {
            "sku": it["sku"],
            "model": it["model"] or "",
            "name": it["name"] or "",
            "qty": need,
            "stock": have,
            "missing": max(0, need - have),
        }
        total_qty += need
        total_missing_qty += row["missing"]
        if have >= need:
            in_stock.append(row)
        else:
            missing.append(row)

    buf = io.BytesIO()
    w = 210 * mm
    h = 297 * mm
    cpdf = canvas.Canvas(buf, pagesize=(w, h))
    pdf_font, pdf_font_bold = get_pdf_font_names()

    y = h - 18 * mm
    cpdf.setFont(pdf_font_bold, 14)
    cpdf.drawString(15 * mm, y, f"Wydruk zamĂłwienia: {order_display_no(o['id'], o['created_at'], o['order_no'], o['note'])}")
    y -= 7 * mm
    cpdf.setFont(pdf_font, 10)
    cpdf.drawString(15 * mm, y, f"Klient: {o['customer_name']}")
    y -= 5 * mm
    cpdf.drawString(15 * mm, y, f"Data: {o['created_at']}")
    y -= 6 * mm
    cpdf.setFont(pdf_font_bold, 10)
    cpdf.drawString(15 * mm, y, f"ĹÄ…czna liczba sztuk w zamĂłwieniu: {total_qty}")
    y -= 5 * mm
    cpdf.setFont(pdf_font, 10)
    cpdf.drawString(15 * mm, y, f"ĹÄ…czny brak na stanie: {total_missing_qty}")

    def draw_section(title, rows, y_pos, show_missing=False):
        cpdf.setFont(pdf_font_bold, 11)
        cpdf.drawString(15 * mm, y_pos, title)
        y_pos -= 6 * mm
        cpdf.setFont(pdf_font_bold, 9)
        cpdf.drawString(15 * mm, y_pos, "SKU")
        cpdf.drawString(55 * mm, y_pos, "Model/Nazwa")
        cpdf.drawString(160 * mm, y_pos, "IloĹ›Ä‡")
        if show_missing:
            cpdf.drawString(176 * mm, y_pos, "Brak")
        y_pos -= 5 * mm
        cpdf.setFont(pdf_font, 9)
        for r in rows:
            label = (r['model'] or r['name'] or "")[:48]
            cpdf.drawString(15 * mm, y_pos, r['sku'])
            cpdf.drawString(55 * mm, y_pos, label)
            cpdf.drawRightString(173 * mm, y_pos, str(r['qty']))
            if show_missing:
                cpdf.drawRightString(195 * mm, y_pos, str(r['missing']))
            y_pos -= 5 * mm
            if y_pos < 20 * mm:
                cpdf.showPage()
                y_pos = h - 20 * mm
                cpdf.setFont(pdf_font, 9)
        return y_pos

    y -= 10 * mm
    y = draw_section("Produkty w magazynie", in_stock, y, show_missing=False)
    y -= 6 * mm
    y = draw_section("Brak na stanie", missing, y, show_missing=True)

    cpdf.showPage()
    cpdf.save()
    buf.seek(0)
    fname = safe_filename(canonical_order_no(o["id"], o["created_at"], o["order_no"])) + "_druk_zamowienia.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname)


def api_client_stock_catalog():
    maybe_pull_shared_from_supabase()

    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT
        p.id AS product_id,
        p.sku,
        p.model,
        p.name,
        COALESCE(s.qty, 0) AS qty_physical,
        COALESCE(r.reserved_qty, 0) AS qty_reserved,
        CASE
          WHEN COALESCE(s.qty, 0) - COALESCE(r.reserved_qty, 0) > 0
          THEN COALESCE(s.qty, 0) - COALESCE(r.reserved_qty, 0)
          ELSE 0
        END AS qty_on_stock,
        COALESCE(pr.net_price, 0) AS net_price,
        COALESCE(pr.gross_price, 0) AS gross_price
      FROM products p
      LEFT JOIN stock s ON s.product_id = p.id
      LEFT JOIN (
        SELECT oi.product_id, SUM(oi.qty) AS reserved_qty
        FROM order_items oi
        JOIN orders o ON o.id = oi.order_id
        WHERE LOWER(COALESCE(o.status,'')) IN ('new','pending','unconfirmed','confirmed','packed','in_delivery','shipped')
          AND COALESCE(o.warehouse_issued,0) = 0
        GROUP BY oi.product_id
      ) r ON r.product_id = p.id
      LEFT JOIN pricing pr
        ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model))
            OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      ORDER BY p.sku
      LIMIT 5000
    """)
    rows = [dict(r) for r in cur.fetchall()]
    c.close()

    return jsonify(ok=True, rows=rows)


@app.route("/api/client_search_log", methods=["POST", "OPTIONS"])
def api_client_search_log():
    if request.method == "OPTIONS":
        return ("", 204)

    data = request.get_json(silent=True) or {}
    query = norm(data.get("query"))[:120]
    if len(query) < 2:
        return jsonify(ok=True, skipped=True)

    # Tożsamość pochodzi wyłącznie ze zweryfikowanego tokenu Supabase,
    # nigdy z danych przesłanych przez przeglądarkę.
    email = norm(g.client_user.get("email")).lower()[:180]
    name = email
    source = norm(data.get("source"))[:40] or "stock"
    results_count = to_int(data.get("results_count"), 0)
    if results_count < 0:
        results_count = 0
    matches = data.get("matches") if isinstance(data.get("matches"), list) else []
    if results_count < 1:
        return jsonify(ok=True, skipped=True, no_results=True)
    rows_to_save = []
    created_at = now_iso()
    seen_products = set()
    for item in matches[:30]:
        if not isinstance(item, dict):
            continue
        product_sku = norm(item.get("sku"))[:120]
        product_model = norm(item.get("model"))[:120] or product_sku
        product_name = norm(item.get("name"))[:180]
        product_key = (product_model.lower(), product_sku.lower())
        if not product_model or product_key in seen_products:
            continue
        seen_products.add(product_key)
        rows_to_save.append({
            "customer_email": email,
            "customer_name": name,
            "query": query,
            "product_sku": product_sku,
            "product_model": product_model,
            "product_name": product_name,
            "results_count": results_count,
            "source": source,
            "created_at": created_at,
        })

    if not rows_to_save:
        rows_to_save.append({
            "customer_email": email,
            "customer_name": name,
            "query": query,
            "product_sku": "",
            "product_model": "",
            "product_name": "",
            "results_count": results_count,
            "source": source,
            "created_at": created_at,
        })

    cutoff = (app_now() - timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M:%S")
    deduped_rows = []
    c = conn()
    cur = c.cursor()
    for row in rows_to_save:
        if row.get("product_sku") or row.get("product_model"):
            cur.execute("""
              SELECT 1
              FROM client_search_logs
              WHERE LOWER(COALESCE(customer_email,''))=?
                AND LOWER(COALESCE(product_sku,''))=?
                AND LOWER(COALESCE(product_model,''))=?
                AND COALESCE(source,'stock')=?
                AND created_at>=?
              LIMIT 1
            """, (
                row.get("customer_email", "").lower(),
                row.get("product_sku", "").lower(),
                row.get("product_model", "").lower(),
                row.get("source", "stock"),
                cutoff,
            ))
        else:
            cur.execute("""
              SELECT 1
              FROM client_search_logs
              WHERE LOWER(COALESCE(customer_email,''))=?
                AND LOWER(COALESCE(query,''))=?
                AND COALESCE(product_sku,'')=''
                AND COALESCE(product_model,'')=''
                AND COALESCE(source,'stock')=?
                AND created_at>=?
              LIMIT 1
            """, (
                row.get("customer_email", "").lower(),
                row.get("query", "").lower(),
                row.get("source", "stock"),
                cutoff,
            ))
        if cur.fetchone():
            continue
        deduped_rows.append(row)
    c.close()

    if not deduped_rows:
        return jsonify(ok=True, skipped=True, duplicate=True)

    cloud_ok = False
    cloud_saved = 0
    for row in deduped_rows:
        try:
            if save_client_search_log_supabase(row):
                cloud_saved += 1
        except Exception:
            pass
        save_client_search_log_local(row)

    cloud_ok = cloud_saved == len(deduped_rows)
    return jsonify(ok=True, cloud=bool(cloud_ok), rows=len(deduped_rows))


def _email_event_already_ok(event_key):
    if not event_key:
        return False
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT ok FROM email_events WHERE event_key=? LIMIT 1", (event_key,))
        row = cur.fetchone()
        return bool(row and to_int(row["ok"], 0) == 1)
    except Exception:
        return False
    finally:
        c.close()


def _record_email_event(event_key, event_type, ref_id, recipient, result):
    if not event_key:
        return
    ok = 1 if isinstance(result, dict) and result.get("ok") else 0
    try:
        payload = json.dumps(result or {}, ensure_ascii=False)[:6000]
    except Exception:
        payload = json.dumps({"raw": str(result)}, ensure_ascii=False)[:6000]
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT id FROM email_events WHERE event_key=? LIMIT 1", (event_key,))
        row = cur.fetchone()
        if row:
            cur.execute("""
                UPDATE email_events
                SET event_type=?, ref_id=?, recipient=?, ok=?, result_json=?, created_at=?
                WHERE event_key=?
            """, (event_type, str(ref_id or ""), recipient or "", ok, payload, now_iso(), event_key))
        else:
            cur.execute("""
                INSERT INTO email_events(event_key,event_type,ref_id,recipient,ok,result_json,created_at)
                VALUES(?,?,?,?,?,?,?)
            """, (event_key, event_type, str(ref_id or ""), recipient or "", ok, payload, now_iso()))
        c.commit()
    except Exception:
        pass
    finally:
        c.close()


def _send_saved_order_confirmation(order_id: int, force: bool = False) -> dict:
    """Send a confirmation using the order saved by the warehouse backend."""
    if not EMAIL_NOTIFICATIONS_ENABLED:
        return {"ok": True, "skipped": True, "reason": "email_notifications_disabled"}
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT * FROM orders WHERE id=? LIMIT 1", (order_id,))
        row = cur.fetchone()
        if not row:
            return {"ok": False, "error": "Nie znaleziono zamówienia"}
        order = dict(row)
        cur.execute("""
          SELECT oi.sku, oi.qty, COALESCE(p.name, pr.name, '') AS name
          FROM order_items oi
          LEFT JOIN products p ON p.id = oi.product_id
          LEFT JOIN products pr ON pr.sku = oi.sku
          WHERE oi.order_id=?
          ORDER BY oi.id
        """, (order_id,))
        items = [dict(x) for x in cur.fetchall()]
    finally:
        c.close()

    try:
        admin_email = norm(email_config_summary().get("admin_email"))
    except Exception:
        admin_email = ""
    recipient = ", ".join([x for x in [norm(order.get("customer_email")), admin_email] if x])
    recipient_hash = hashlib.sha1(recipient.lower().encode("utf-8")).hexdigest()[:12] if recipient else "no-recipient"
    event_key = f"order_confirmation:{order_id}:{recipient_hash}"

    if not force and _email_event_already_ok(event_key):
        return {"ok": True, "duplicate": True, "skipped": True, "to": recipient}
    if not send_order_confirmation:
        result = {"ok": False, "skipped": True, "error": "Brak modułu email_module.py"}
    else:
        try:
            result = send_order_confirmation(order, items, admin_email=admin_email)
        except Exception as exc:
            result = {"ok": False, "error": str(exc)}
    _record_email_event(event_key, "order_confirmation", order_id, recipient, result)
    return result


def _authenticated_client_user() -> dict | None:
    auth = norm(request.headers.get("Authorization"))
    if not auth.lower().startswith("bearer "):
        return None
    token = auth.split(None, 1)[1].strip()
    if not token:
        return None
    req = urllib.request.Request(f"{SUPABASE_URL}/auth/v1/user", method="GET")
    # Endpoint Auth powinien dostać publiczny klucz projektu jako `apikey`.
    # Token użytkownika pozostaje osobno w nagłówku Authorization.
    api_key = SUPABASE_ANON_KEY or SUPABASE_SERVICE_ROLE_KEY
    if not api_key:
        app.logger.error("Weryfikacja klienta niemożliwa: brak SUPABASE_ANON_KEY")
        return None
    req.add_header("apikey", api_key)
    req.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        email = norm(payload.get("email")).lower()
        if not payload.get("id") or not email:
            return None
        metadata = payload.get("user_metadata") if isinstance(payload.get("user_metadata"), dict) else {}
        return {
            "id": str(payload.get("id")),
            "email": email,
            "name": norm(metadata.get("full_name") or metadata.get("name")) or email.split("@")[0],
        }
    except urllib.error.HTTPError as exc:
        app.logger.warning("Supabase odrzucił token klienta: HTTP %s", exc.code)
        return None
    except Exception as exc:
        app.logger.warning("Nie udało się zweryfikować tokenu klienta: %s", type(exc).__name__)
        return None


def _order_by_idempotency_key(idempotency_key: str) -> dict | None:
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("SELECT id, order_no, customer_email FROM orders WHERE idempotency_key=? LIMIT 1", (idempotency_key,))
        row = cur.fetchone()
        if row:
            return dict(row)
    finally:
        c.close()
    if not supabase_enabled():
        return None
    rows = supabase_request(
        "/rest/v1/orders",
        params={"select": "id,order_no,customer_email", "idempotency_key": f"eq.{idempotency_key}", "limit": 1},
    )
    return dict(rows[0]) if isinstance(rows, list) and rows else None


def _client_order_origin_allowed() -> bool:
    origin = norm(request.headers.get("Origin")).rstrip("/")
    return not origin or origin in CLIENT_ALLOWED_ORIGINS


@app.route("/api/client/orders", methods=["POST", "OPTIONS"])
def api_client_orders_create():
    """Create the complete order and send its email in one backend request."""
    if not _client_order_origin_allowed():
        return jsonify(ok=False, error="Niedozwolone źródło żądania"), 403
    if request.method == "OPTIONS":
        return ("", 204)

    if not supabase_enabled():
        app.logger.error("Odrzucono zamówienie klienta: brak konfiguracji Supabase")
        return jsonify(ok=False, error="Brak konfiguracji połączenia z Supabase"), 503

    data = request.get_json(silent=True) or {}
    client_user = _authenticated_client_user()
    if not client_user:
        app.logger.warning("Odrzucono zamówienie klienta: brak lub nieważny token")
        return jsonify(ok=False, error="Brak autoryzacji"), 401

    customer_email = client_user["email"]
    customer_name = client_user["name"]
    note = norm(data.get("note"))
    idempotency_key = norm(request.headers.get("Idempotency-Key"))
    try:
        uuid.UUID(idempotency_key)
    except Exception:
        return jsonify(ok=False, error="Brak lub niepoprawny Idempotency-Key"), 400

    try:
        maybe_pull_shared_from_supabase(force=True)
    except Exception as exc:
        app.logger.error("Nie udało się odświeżyć danych Supabase przed zamówieniem: %s", exc)
        return jsonify(ok=False, error="Nie udało się odświeżyć danych produktów"), 503

    existing = _order_by_idempotency_key(idempotency_key)
    if existing:
        if norm(existing.get("customer_email")).lower() != customer_email:
            return jsonify(ok=False, error="Konflikt Idempotency-Key"), 409
        app.logger.info("Ponowiono request user_id=%s order_id=%s idempotency_key=%s", client_user["id"], existing["id"], idempotency_key)
        return jsonify(ok=True, duplicate=True, order={"id": existing["id"], "order_no": existing["order_no"]}, email=_send_saved_order_confirmation(int(existing["id"])))

    raw_items = data.get("items")
    if not isinstance(raw_items, list) or not raw_items:
        return jsonify(ok=False, error="Zamówienie nie zawiera pozycji"), 400
    if len(raw_items) > 100:
        return jsonify(ok=False, error="Zamówienie może zawierać maksymalnie 100 pozycji"), 400

    items = []
    c = conn()
    try:
        cur = c.cursor()
        for item in raw_items:
            if not isinstance(item, dict):
                return jsonify(ok=False, error="Nieprawidłowa pozycja zamówienia"), 400
            product_id = item.get("product_id")
            qty = item.get("qty")
            if isinstance(product_id, bool) or not isinstance(product_id, int) or product_id <= 0:
                return jsonify(ok=False, error="Nieprawidłowy identyfikator produktu"), 400
            if isinstance(qty, bool) or not isinstance(qty, int) or qty <= 0:
                label = norm(item.get("sku")) or str(product_id)
                return jsonify(ok=False, error=f"Nieprawidłowa ilość dla produktu {label}"), 400
            cur.execute("SELECT id, sku, name FROM products WHERE id=? LIMIT 1", (product_id,))
            product = cur.fetchone()
            if not product:
                return jsonify(ok=False, error=f"Nie istnieje produkt ID {product_id}"), 400
            submitted_sku = norm(item.get("sku"))
            if submitted_sku and submitted_sku.lower() != norm(product["sku"]).lower():
                return jsonify(ok=False, error=f"Produkt ID {product_id} nie odpowiada SKU {submitted_sku}"), 400
            items.append((product_id, qty))
    finally:
        c.close()

    try:
        order_id = remote_first_create_order(None, customer_name, "", "", customer_email, note, items, idempotency_key=idempotency_key)
        email_result = _send_saved_order_confirmation(order_id)
        if not email_result.get("ok"):
            email_result["pending_retry"] = True
        c = conn()
        try:
            cur = c.cursor()
            cur.execute("SELECT order_no FROM orders WHERE id=?", (order_id,))
            row = cur.fetchone()
            order_no = row["order_no"] if row else make_order_no(order_id, now_iso())
        finally:
            c.close()
        app.logger.info("Utworzono zamówienie user_id=%s order_id=%s order_no=%s items=%s email_ok=%s", client_user["id"], order_id, order_no, len(items), bool(email_result.get("ok")))
        return jsonify(ok=True, order={"id": order_id, "order_no": order_no}, email=email_result)
    except Exception as exc:
        existing = _order_by_idempotency_key(idempotency_key)
        if existing:
            if norm(existing.get("customer_email")).lower() != customer_email:
                return jsonify(ok=False, error="Konflikt Idempotency-Key"), 409
            app.logger.info("Konflikt idempotencji user_id=%s order_id=%s", client_user["id"], existing["id"])
            return jsonify(ok=True, duplicate=True, order={"id": existing["id"], "order_no": existing["order_no"]}, email=_send_saved_order_confirmation(int(existing["id"])))
        app.logger.exception("Błąd tworzenia zamówienia user_id=%s items=%s", client_user["id"], len(items))
        return jsonify(ok=False, error=str(exc)), 500


@app.route("/api/client_order_email", methods=["POST", "OPTIONS"])
def api_client_order_email():
    if not _client_order_origin_allowed():
        return jsonify(ok=False, error="Niedozwolone źródło żądania"), 403
    if request.method == "OPTIONS":
        return ("", 204)

    return jsonify(
        ok=False,
        error="Ten endpoint został wyłączony. Zamówienie i potwierdzenie obsługuje /api/client/orders.",
    ), 410

    app.logger.warning("Użyto przestarzałego endpointu /api/client_order_email; zaktualizuj panel do /api/client/orders")

    data = request.get_json(silent=True) or {}
    order_id = to_int(data.get("order_id"), 0)
    order_no = norm(data.get("order_no"))
    fallback_email = norm(data.get("customer_email") or data.get("email")).lower()
    fallback_name = norm(data.get("customer_name")) or (fallback_email.split("@")[0] if fallback_email else "")
    fallback_note = norm(data.get("note"))
    fallback_items = data.get("items") if isinstance(data.get("items"), list) else []

    order = {
        "id": order_id,
        "order_no": order_no,
        "customer_email": fallback_email,
        "customer_name": fallback_name,
        "note": fallback_note,
        "created_at": now_iso(),
    }
    items = []

    try:
        maybe_pull_shared_from_supabase(force=True)
    except Exception:
        pass

    c = conn()
    cur = c.cursor()
    try:
        if order_id:
            cur.execute("SELECT * FROM orders WHERE id=? LIMIT 1", (order_id,))
        elif order_no:
            cur.execute("SELECT * FROM orders WHERE order_no=? LIMIT 1", (order_no,))
        else:
            cur.execute("SELECT * FROM orders WHERE 1=0")
        row = cur.fetchone()
        if row:
            db_order = dict(row)
            # Panel klienta jest źródłem prawdy dla adresu odbiorcy maila.
            # Lokalna baza na Renderze może mieć starszy rekord po synchronizacji,
            # więc nie wolno blokować podmiany, jeśli email już istnieje.
            if fallback_email:
                db_order["customer_email"] = fallback_email
            if fallback_name:
                db_order["customer_name"] = fallback_name
            if fallback_note and not norm(db_order.get("note")):
                db_order["note"] = fallback_note
            order = db_order
            cur.execute("""
              SELECT oi.sku, oi.qty, COALESCE(p.name, pr.name, '') AS name
              FROM order_items oi
              LEFT JOIN products p ON p.id = oi.product_id
              LEFT JOIN products pr ON pr.sku = oi.sku
              WHERE oi.order_id=?
              ORDER BY oi.id
            """, (row["id"],))
            items = [dict(x) for x in cur.fetchall()]
    except Exception:
        items = []
    finally:
        c.close()

    if not items:
        for item in fallback_items[:80]:
            if not isinstance(item, dict):
                continue
            items.append({
                "sku": norm(item.get("sku")),
                "name": norm(item.get("name")),
                "qty": to_int(item.get("qty"), 0),
            })

    if fallback_email:
        order["customer_email"] = fallback_email
    if fallback_name:
        order["customer_name"] = fallback_name

    event_ref = norm(order.get("id")) or norm(order_id) or norm(order.get("order_no")) or order_no
    try:
        admin_email = norm(email_config_summary().get("admin_email"))
    except Exception:
        admin_email = ""
    recipient = ", ".join([x for x in [norm(order.get("customer_email")), admin_email] if x])
    recipient_hash = hashlib.sha1(recipient.lower().encode("utf-8")).hexdigest()[:12] if recipient else "no-recipient"
    event_key = f"order_confirmation:{event_ref}:{recipient_hash}" if event_ref else ""

    if _email_event_already_ok(event_key):
        return jsonify(ok=True, email={"ok": True, "duplicate": True, "skipped": True, "to": recipient, "order_email": norm(order.get("customer_email"))})

    if not send_order_confirmation:
        result = {"ok": False, "skipped": True, "error": "Brak modułu email_module.py"}
        _record_email_event(event_key, "order_confirmation", event_ref, recipient, result)
        return jsonify(ok=True, email=result)

    try:
        result = send_order_confirmation(order, items, admin_email=admin_email)
    except Exception as exc:
        result = {"ok": False, "error": str(exc)}
    _record_email_event(event_key, "order_confirmation", event_ref, recipient, result)
    return jsonify(ok=True, email=result, to=recipient, order_email=norm(order.get("customer_email")))


@app.post("/email/order-confirmations/retry-failed")
def retry_failed_order_confirmations():
    supplied_token = norm(request.headers.get("X-Admin-Token"))
    if not ADMIN_ACTION_TOKEN or supplied_token != ADMIN_ACTION_TOKEN:
        return jsonify(ok=False, error="Brak autoryzacji"), 401
    c = conn()
    try:
        cur = c.cursor()
        cur.execute("""
          SELECT ref_id
          FROM email_events
          WHERE event_type='order_confirmation' AND ok=0
          ORDER BY created_at
          LIMIT 50
        """)
        order_ids = [to_int(row["ref_id"], 0) for row in cur.fetchall()]
    finally:
        c.close()

    retried = 0
    sent = 0
    for order_id in order_ids:
        if order_id <= 0:
            continue
        retried += 1
        result = _send_saved_order_confirmation(order_id)
        if result.get("ok"):
            sent += 1
    app.logger.info("Retry potwierdzeń zamówień retried=%s sent=%s", retried, sent)
    return jsonify(ok=True, retried=retried, sent=sent)


@app.route("/email-test", methods=["GET", "POST"])
def email_test():
    cfg = email_config_summary()
    cfg = dict(cfg or {})
    cfg["module_loaded"] = bool(send_email)
    cfg["import_error"] = _EMAIL_IMPORT_ERROR
    cfg["api_key_set"] = "RESEND_API_KEY" not in (cfg.get("missing") or [])
    result = None
    test_to = norm(request.form.get("to")) or norm(cfg.get("admin_email")) or "biuro@niedzwieccy.com"

    if request.method == "POST":
        if not send_email:
            result = {
                "ok": False,
                "error": "Moduł email_module.py nie jest załadowany. Wgraj email_module.py do repo obok app.py i zrób deploy.",
                "import_error": _EMAIL_IMPORT_ERROR,
            }
        else:
            try:
                result = send_email(
                    test_to,
                    "Test maili z Beton Łagów Orders",
                    "<div style='font-family:Arial,sans-serif'><h2>Test maili działa</h2><p>Jeśli widzisz tę wiadomość, Resend jest poprawnie podpięty do aplikacji.</p></div>",
                    "Test maili działa. Jeśli widzisz tę wiadomość, Resend jest poprawnie podpięty do aplikacji.",
                )
            except Exception as exc:
                result = {"ok": False, "error": str(exc)}

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <h1>Test maili</h1>
        <p class="muted">Ta strona sprawdza konfigurację Resend po stronie Rendera. API key nie jest tutaj wyświetlany.</p>
        <div class="kpi">
          <span class="pill">Moduł: <b>{{ 'załadowany' if cfg.module_loaded else 'brak' }}</b></span>
          <span class="pill">Wysyłka: <b>{{ 'włączona' if cfg.enabled else 'wyłączona' }}</b></span>
          <span class="pill">Konfiguracja: <b>{{ 'OK' if cfg.configured else 'brakuje danych' }}</b></span>
          <span class="pill">API key: <b>{{ 'ustawiony' if cfg.api_key_set else 'brak' }}</b></span>
        </div>
        <div class="line"></div>
        <p><b>EMAIL_FROM:</b> {{ cfg['from'] or '-' }}</p>
        <p><b>ADMIN_EMAIL:</b> {{ cfg.admin_email or '-' }}</p>
        {% if cfg.missing %}
          <p class="hint"><b>Brakuje:</b> {{ cfg.missing|join(', ') }}</p>
        {% endif %}
        {% if cfg.import_error %}
          <p class="hint"><b>Błąd importu:</b> {{ cfg.import_error }}</p>
        {% endif %}
        <form method="post" class="flex" style="margin-top:12px">
          <input name="to" value="{{ test_to }}" placeholder="email do testu" style="max-width:420px">
          <button class="btn primary" type="submit">Wyślij test</button>
        </form>
      </div>

      {% if result %}
        <div class="card">
          <h2>Wynik testu</h2>
          {% if result.ok %}
            <p class="badge" style="background:#dcfce7;border-color:#86efac">Mail wysłany</p>
          {% else %}
            <p class="badge" style="background:#fee2e2;border-color:#fecaca">Mail nie wysłany</p>
          {% endif %}
          <pre style="white-space:pre-wrap;background:#111;color:#fff;padding:12px;border-radius:12px;overflow:auto">{{ result|tojson(indent=2) }}</pre>
        </div>
      {% endif %}
    {% endblock %}
    """
    return render_template_string(tpl, title="Test maili", base_url=BASE_URL, db_path=DB_PATH, cfg=cfg, result=result, test_to=test_to)


@app.get("/api/order_lookup")
def api_order_lookup():
    maybe_pull_shared_from_supabase(force=True)
    token = norm(request.args.get("token"))
    if not token:
        return jsonify(ok=False, error="Brak tokenu"), 400

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE order_no=? LIMIT 1", (token,))
    o = cur.fetchone()
    if not o:
        cur.execute("SELECT * FROM orders ORDER BY id DESC")
        all_orders = cur.fetchall()
        for row in all_orders:
            if canonical_order_no(row["id"], row["created_at"], row["order_no"]) == norm(token):
                o = row
                break
    if not o:
        c.close()
        return jsonify(ok=False, error="Nie znaleziono zamĂłwienia"), 404
    if _email_key(o["customer_email"]) != _email_key(g.client_user["email"]):
        c.close()
        return jsonify(ok=False, error="Brak dostępu"), 403

    cur.execute("""
      SELECT oi.*, p.model, p.ean, p.name,
             COALESCE(s.qty, 0) AS stock_qty,
             COALESCE(s.qty, 0) AS stock,
             COALESCE((
                SELECT SUM(ci.qty)
                FROM material_order_items ci
                JOIN material_orders cp ON cp.id=ci.package_id
                WHERE ci.product_id=oi.product_id
                  AND cp.status IN ('planned', 'ordered', 'shipped')
             ), 0) AS in_delivery,
             COALESCE(pr.net_price, 0) AS net_price,
             COALESCE(pr.gross_price, 0) AS gross_price,
             (oi.qty * COALESCE(pr.net_price, 0)) AS line_value_net,
             (oi.qty * COALESCE(pr.gross_price, 0)) AS line_value_gross
      FROM order_items oi
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN stock s ON s.product_id=p.id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      WHERE oi.order_id=?
      ORDER BY oi.id
    """, (o["id"],))
    items = [dict(r) for r in cur.fetchall()]

    for it in items:
        it["in_delivery_available"] = int(it.get("in_delivery", 0) or 0)
        it["delivery_used"] = 0
        it["line_shortage"] = 0

    order_id = int(o["id"])
    if o["status"] in ("new", "packed", "confirmed", "in_delivery"):
        cur.execute("SELECT id FROM orders WHERE status IN ('new','packed','confirmed','in_delivery') AND id<=? ORDER BY id", (order_id,))
        scoped_order_ids = [int(r["id"]) for r in cur.fetchall()]
        if scoped_order_ids:
            sph = ",".join(["?"] * len(scoped_order_ids))
            cur.execute(f"""
              SELECT oi.id, oi.order_id, oi.product_id, oi.qty
              FROM order_items oi
              WHERE oi.order_id IN ({sph})
              ORDER BY oi.order_id, oi.id
            """, tuple(scoped_order_ids))
            seq_items = cur.fetchall()

            product_ids = {int(r["product_id"]) for r in seq_items}
            pool_stock = {}
            pool_delivery = {}
            if product_ids:
                pph = ",".join(["?"] * len(product_ids))
                cur.execute(f"""
                  SELECT p.id AS product_id,
                         COALESCE(s.qty,0) AS stock_qty,
                         COALESCE((
                           SELECT SUM(ci.qty)
                           FROM material_order_items ci
                           JOIN material_orders cp ON cp.id=ci.package_id
                           WHERE ci.product_id=p.id
                             AND cp.status IN ('planned', 'ordered', 'shipped')
                         ),0) AS in_delivery_qty
                  FROM products p
                  LEFT JOIN stock s ON s.product_id=p.id
                  WHERE p.id IN ({pph})
                """, tuple(product_ids))
                for pr in cur.fetchall():
                    pid = int(pr["product_id"])
                    pool_stock[pid] = int(pr["stock_qty"])
                    pool_delivery[pid] = int(pr["in_delivery_qty"])

            item_alloc = {}
            for sr in seq_items:
                pid = int(sr["product_id"])
                need = int(sr["qty"])

                stock_now = pool_stock.get(pid, 0)
                from_stock = min(stock_now, need)
                pool_stock[pid] = stock_now - from_stock
                need_after_stock = need - from_stock

                delivery_now = pool_delivery.get(pid, 0)
                from_delivery = min(delivery_now, need_after_stock)
                pool_delivery[pid] = delivery_now - from_delivery
                shortage = need_after_stock - from_delivery

                if int(sr["order_id"]) == order_id:
                    item_alloc[int(sr["id"])] = {
                        "in_delivery_available": from_delivery,
                        "delivery_used": from_delivery,
                        "line_shortage": shortage,
                    }

            for it in items:
                al = item_alloc.get(int(it["id"]))
                if al:
                    it.update(al)
    c.close()

    invoiced_by_item = invoiced_qty_by_order_item_ids([int(it["id"]) for it in items])
    for it in items:
        ordered_qty = int(it.get("qty") or 0)
        invoiced_qty = int(invoiced_by_item.get(int(it["id"])) or 0)
        it["ordered_qty"] = ordered_qty
        it["invoiced_qty"] = invoiced_qty
        it["remaining_qty"] = max(0, ordered_qty - invoiced_qty)
        stock_qty = int(it.get("stock_qty") or 0)
        delivery_used = int(it.get("delivery_used") or 0)
        line_shortage = int(it.get("line_shortage") or 0)
        if o["status"] in ("new", "packed", "confirmed", "in_delivery"):
            it["availability_label"] = "dostępne" if stock_qty >= ordered_qty else "10/20 dni"
        else:
            it["availability_label"] = "dostępne" if stock_qty >= ordered_qty else "10/20 dni"
        if ordered_qty > 0 and invoiced_qty >= ordered_qty:
            it["realization_label"] = "w całości"
        elif invoiced_qty > 0:
            it["realization_label"] = f"częściowo: {invoiced_qty}/{ordered_qty} szt."
        else:
            it["realization_label"] = "0 szt."

    total_net = round(sum(float(it.get("line_value_net") or 0) for it in items), 2)
    total_gross = round(sum(float(it.get("line_value_gross") or 0) for it in items), 2)

    return jsonify(
        ok=True,
        order={
            "id": o["id"],
            "order_no": o["order_no"],
            "status": o["status"],
            "created_at": o["created_at"],
            "customer_name": o["customer_name"],
            "customer_address": o["customer_address"],
            "customer_phone": o["customer_phone"],
            "customer_email": o["customer_email"],
            "note": o["note"],
            "warehouse_issued": int(o["warehouse_issued"] or 0),
            "total_net": total_net,
            "total_gross": total_gross,
        },
        items=items
    )


@app.get("/api/client_invoices")
def api_client_invoices():
    maybe_pull_shared_from_supabase()
    email = _email_key(g.client_user["email"])

    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT
        i.*,
        m.invoice_id AS meta_invoice_id,
        COALESCE(m.pdf_path,'') AS pdf_path,
        COALESCE(m.sent_to_client,0) AS sent_to_client,
        COALESCE(m.seen_by_client,0) AS seen_by_client,
        COALESCE(m.payment_reminder,0) AS payment_reminder,
        COALESCE(m.paid,0) AS paid,
        COALESCE(m.paid_at,'') AS paid_at,
        COALESCE(m.seen_at,'') AS seen_at,
        COALESCE(k.status,'draft') AS ksef_status,
        COALESCE(k.ksef_number,'') AS ksef_number,
        COALESCE(k.last_error,'') AS ksef_error,
        COALESCE(k.sent_at,'') AS ksef_sent_at,
        o.id AS source_order_id,
        o.order_no,
        o.created_at AS source_order_created_at,
        o.note AS source_order_note,
        o.customer_email AS order_customer_email
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      LEFT JOIN ksef_documents k ON k.invoice_id = i.id
      LEFT JOIN orders o ON o.id = i.order_id
      WHERE (
          LOWER(COALESCE(i.buyer_email,'')) = ?
          OR LOWER(COALESCE(o.customer_email,'')) = ?
        )
        AND (
          COALESCE(m.sent_to_client,0)=1
          OR m.invoice_id IS NULL
        )
      ORDER BY i.order_id DESC, i.id DESC
    """, (email, email))
    rows = []
    for r in cur.fetchall():
        d = dict(r)
        if d.get("meta_invoice_id") is None:
            d["sent_to_client"] = 1
        d["order_display"] = order_display_no(
            d.get("source_order_id"),
            d.get("source_order_created_at"),
            d.get("order_no"),
            d.get("source_order_note")
        ) if d.get("source_order_id") else (d.get("order_no") or "")
        d["pdf_exists"] = 1 if d.get("pdf_path") else 0
        api_base = request.url_root.rstrip("/")
        d["download_url"] = f"{api_base}/api/invoices/{d.get('id')}/download?email={urllib.parse.quote_plus(email)}"
        rows.append(d)
    c.close()
    rows.sort(key=lambda x: ((x.get("seen_by_client") or 0), (x.get("issue_date") or ""), int(x.get("id") or 0)), reverse=True)
    return jsonify(ok=True, invoices=rows)


@app.get("/invoices")
def invoices():
    maybe_pull_shared_from_supabase()
    q = norm(request.args.get("q"))
    c = conn()
    cur = c.cursor()
    params = []
    where = ""
    if q:
        like = f"%{q.lower()}%"
        where = """
          WHERE LOWER(COALESCE(i.invoice_no,'')) LIKE ?
             OR LOWER(COALESCE(i.buyer_name,'')) LIKE ?
             OR LOWER(COALESCE(o.customer_name,'')) LIKE ?
             OR LOWER(COALESCE(o.order_no,'')) LIKE ?
             OR LOWER(COALESCE(o.note,'')) LIKE ?
        """
        params = [like, like, like, like, like]

    cur.execute(f"""
      SELECT
        i.*,
        COALESCE(m.pdf_path,'') AS pdf_path,
        COALESCE(m.sent_to_client,0) AS sent_to_client,
        COALESCE(m.seen_by_client,0) AS seen_by_client,
        COALESCE(m.payment_reminder,0) AS payment_reminder,
        COALESCE(m.paid,0) AS paid,
        COALESCE(m.paid_at,'') AS paid_at,
        COALESCE(m.seen_at,'') AS seen_at,
        COALESCE(k.status,'draft') AS ksef_status,
        COALESCE(k.ksef_number,'') AS ksef_number,
        COALESCE(k.last_error,'') AS ksef_error,
        COALESCE(k.sent_at,'') AS ksef_sent_at,
        o.id AS source_order_id,
        o.order_no AS source_order_no,
        o.created_at AS source_order_created_at,
        o.note AS source_order_note,
        o.customer_name AS order_customer_name
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      LEFT JOIN ksef_documents k ON k.invoice_id = i.id
      LEFT JOIN orders o ON o.id = i.order_id
      {where}
      ORDER BY LOWER(COALESCE(i.buyer_name, o.customer_name, '')), i.issue_date DESC, i.id DESC
    """, params)
    rows = [dict(r) for r in cur.fetchall()]
    c.close()

    groups = []
    current_key = None
    current = None
    for inv in rows:
        customer_name = inv.get("buyer_name") or inv.get("order_customer_name") or "Bez klienta"
        key = customer_name.strip().lower()
        if key != current_key:
            current = {"customer_name": customer_name, "invoices": [], "months": [], "total_net": 0.0, "total_gross": 0.0}
            groups.append(current)
            current_key = key
        inv["order_display"] = order_display_no(
            inv.get("source_order_id"),
            inv.get("source_order_created_at"),
            inv.get("source_order_no"),
            inv.get("source_order_note")
        ) if inv.get("source_order_id") else "-"
        inv["pdf_ok"] = 1 if (invoice_pdf_exists(inv.get("pdf_path", ""), inv.get("invoice_no", ""))[0] or inv.get("invoice_items_json")) else 0
        current["invoices"].append(inv)
        current["total_net"] += float(inv.get("total_net") or 0)
        current["total_gross"] += float(inv.get("total_gross") or 0)

    for g in groups:
        month_map = {}
        for inv in g["invoices"]:
            issue_date = norm(inv.get("issue_date"))
            month_key = issue_date[:7] if len(issue_date) >= 7 else "bez-daty"
            month_label = month_key if month_key != "bez-daty" else "Bez daty"
            if month_key not in month_map:
                month_map[month_key] = {"month": month_key, "label": month_label, "invoices": [], "total_net": 0.0, "total_gross": 0.0}
                g["months"].append(month_map[month_key])
            month = month_map[month_key]
            month["invoices"].append(inv)
            month["total_net"] += float(inv.get("total_net") or 0)
            month["total_gross"] += float(inv.get("total_gross") or 0)

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Faktury</h1>
        </div>
        <form method="get" class="flex" style="margin-top:12px;">
          <input name="q" value="{{ q }}" placeholder="Szukaj: klient, numer faktury, numer zamówienia, notatka">
          <button class="btn primary" type="submit">Szukaj</button>
          <a class="btn" href="{{ url_for('invoices') }}">Wyczyść</a>
        </form>
      </div>

      {% for g in groups %}
        <div class="card">
          <details {% if q %}open{% endif %}>
            <summary class="flex" style="cursor:pointer; align-items:center;">
              <h2 style="margin:0;">{{ g.customer_name }}</h2>
              <span class="badge">{{ g.invoices|length }} faktur</span>
              <span class="badge">Netto: {{ "%.2f"|format(g.total_net) }} PLN</span>
              <span class="badge">Brutto: {{ "%.2f"|format(g.total_gross) }} PLN</span>
              <span class="btn right">Pokaż faktury</span>
            </summary>

            {% for m in g.months %}
              <details style="margin-top:10px;" {% if q %}open{% endif %}>
                <summary class="flex" style="cursor:pointer; align-items:center;">
                  <b>{{ m.label }}</b>
                  <span class="badge">{{ m.invoices|length }} faktur</span>
                  <span class="badge">Netto: {{ "%.2f"|format(m.total_net) }} PLN</span>
                  <span class="badge">Brutto: {{ "%.2f"|format(m.total_gross) }} PLN</span>
                </summary>

                <table style="margin-top:10px;">
                  <thead>
                    <tr>
                      <th>Faktura</th>
                      <th>Data</th>
                      <th>Zamówienie</th>
                      <th>Netto</th>
                      <th>Brutto</th>
                      <th>Płatność / KSeF</th>
                      <th>Akcje</th>
                    </tr>
                  </thead>
                  <tbody>
                    {% for inv in m.invoices %}
                      <tr>
                        <td><b>{{ inv.invoice_no }}</b></td>
                        <td>{{ inv.issue_date }}</td>
                        <td>{{ inv.order_display }}</td>
                        <td>{{ "%.2f"|format(inv.total_net) }}</td>
                        <td>{{ "%.2f"|format(inv.total_gross) }}</td>
                        <td>
                          {% if not inv.pdf_ok %}
                            <span class="badge danger">Brak PDF</span>
                          {% endif %}
                          {% if inv.paid %}
                            <span class="badge ok">Opłacona</span>
                          {% else %}
                            <span class="badge danger">Nieopłacona</span>
                          {% endif %}
                          {% if inv.ksef_status == 'sent' %}
                            <span class="badge ok">W KSeF</span>
                            {% if inv.ksef_number %}<div class="muted small">{{ inv.ksef_number }}</div>{% endif %}
                          {% elif inv.ksef_status == 'ready' %}
                            <span class="badge ok">KSeF FA(3) OK</span>
                          {% elif inv.ksef_status == 'error' %}
                            <span class="badge danger">KSeF do poprawy</span>
                            {% if inv.ksef_error %}<div class="muted small">{{ inv.ksef_error }}</div>{% endif %}
                          {% else %}
                            <span class="badge">Nie wysłana do KSeF</span>
                          {% endif %}
                        </td>
                        <td>
                          <div class="flex">
                            <a class="btn" href="{{ url_for('invoice_download_admin', invoice_id=inv.id) }}" target="_blank">Faktura PDF</a>
                            {% if inv.source_order_id %}
                              <a class="btn" href="{{ url_for('order_view', order_id=inv.source_order_id) }}">Zamówienie</a>
                            {% endif %}
                            {% if not inv.paid %}
                              <form method="post" action="{{ url_for('invoice_paid_admin', invoice_id=inv.id) }}">
                                <input type="hidden" name="next" value="{{ request.full_path }}">
                                <button class="btn ok" type="submit">Oznacz jako opłaconą</button>
                              </form>
                            {% else %}
                              <form method="post" action="{{ url_for('invoice_unpaid_admin', invoice_id=inv.id) }}">
                                <input type="hidden" name="next" value="{{ request.full_path }}">
                                <button class="btn" type="submit">Oznacz jako nieopłaconą</button>
                              </form>
                            {% endif %}
                            {% if inv.ksef_status != 'sent' %}
                              <a class="btn" href="{{ url_for('invoice_ksef_xml', invoice_id=inv.id) }}">XML KSeF FA(3)</a>
                              <form method="post" action="{{ url_for('invoice_ksef_validate', invoice_id=inv.id) }}">
                                <input type="hidden" name="next" value="{{ request.full_path }}">
                                <button class="btn" type="submit">Sprawdź KSeF</button>
                              </form>
                              <form method="post" action="{{ url_for('invoice_ksef_send', invoice_id=inv.id) }}" onsubmit="return confirm('UWAGA: to jest realna wysyłka faktury do KSeF. Po wysłaniu faktura otrzyma numer KSeF i nie będzie można jej edytować. Kontynuować?');">
                                <input type="hidden" name="next" value="{{ request.full_path }}">
                                <button class="btn primary" type="submit">Wyślij do KSeF</button>
                              </form>
                              <a class="btn" href="{{ url_for('invoice_edit_admin', invoice_id=inv.id) }}">Edytuj</a>
                              <form method="post" action="{{ url_for('invoice_delete_admin', invoice_id=inv.id) }}" onsubmit="return confirm('Usunąć fakturę {{ inv.invoice_no }}? To usunie też PDF i widoczność w panelu klienta.')">
                                <input type="hidden" name="next" value="{{ request.full_path }}">
                                <button class="btn danger" type="submit">Usuń</button>
                              </form>
                            {% else %}
                              <form method="post" action="{{ url_for('invoice_rollback_admin', invoice_id=inv.id) }}" onsubmit="return confirm('AWARYJNIE cofnąć fakturę {{ inv.invoice_no }} w aplikacji? To usunie lokalny zapis faktury, status KSeF, widoczność u klienta i przeliczy zamówienia oraz stany. Używaj tylko przy pomyłce/testach.');">
                                <input type="hidden" name="next" value="{{ request.full_path }}">
                                <button class="btn danger" type="submit">Cofnij fakturę</button>
                              </form>
                            {% endif %}
                          </div>
                        </td>
                      </tr>
                    {% endfor %}
                  </tbody>
                </table>
              </details>
            {% endfor %}
          </details>
        </div>
      {% endfor %}

      {% if not groups %}
        <div class="card muted">Brak faktur.</div>
      {% endif %}
    {% endblock %}
    """
    return render_template_string(tpl, title="Faktury", base_url=BASE_URL, db_path=DB_PATH, groups=groups, q=q)


def load_invoice_with_meta(invoice_id: int):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT i.*, COALESCE(m.pdf_path,'') AS pdf_path, COALESCE(m.sent_to_client,0) AS sent_to_client,
             COALESCE(m.invoice_items_json,'') AS invoice_items_json
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      WHERE i.id=?
      LIMIT 1
    """, (invoice_id,))
    row = cur.fetchone()
    c.close()
    return dict(row) if row else None

def invoice_meta_payload(invoice_row: dict):
    buyer_address = "\n".join([x for x in [
        invoice_row.get("buyer_street") or "",
        f"{invoice_row.get('buyer_post_code') or ''} {invoice_row.get('buyer_city') or ''}".strip()
    ] if x]).strip()
    ksef_number = invoice_row.get("ksef_number") or ""
    if not ksef_number and invoice_row.get("id"):
        try:
            ksef_number = load_ksef_doc(int(invoice_row.get("id") or 0)).get("ksef_number") or ""
        except Exception:
            ksef_number = ""

    return {
        "invoice_no": invoice_row.get("invoice_no") or "",
        "place": "KotuszĂłw",
        "issue_date": invoice_row.get("issue_date") or app_now().strftime("%Y-%m-%d"),
        "sell_date": invoice_row.get("sell_date") or app_now().strftime("%Y-%m-%d"),
        "payment_type": invoice_row.get("payment_type") or "przelew",
        "payment_to": invoice_row.get("payment_to") or "",
        "buyer_name": invoice_row.get("buyer_name") or "",
        "buyer_tax_no": invoice_row.get("buyer_tax_no") or "",
        "buyer_address": buyer_address,
        "buyer_country": invoice_row.get("buyer_country") or "PL",
        "buyer_email": invoice_row.get("buyer_email") or "",
        "buyer_phone": invoice_row.get("buyer_phone") or "",
        "discount_percent": "0",
        "ksef_number": ksef_number,
    }

def invoice_items_from_saved_json(invoice_id: int):
    meta = load_invoice_meta(invoice_id) or {}
    raw = meta.get("invoice_items_json") or ""
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, list) and data:
                return data
        except Exception:
            pass

    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT oi.*, p.model, p.name,
             COALESCE(pr.net_price, 0) AS net_price,
             COALESCE(pr.gross_price, 0) AS gross_price,
             (oi.qty * COALESCE(pr.net_price, 0)) AS line_value_net,
             (oi.qty * COALESCE(pr.gross_price, 0)) AS line_value_gross
      FROM order_items oi
      JOIN products p ON p.id=oi.product_id
      LEFT JOIN pricing pr ON (TRIM(LOWER(pr.model)) = TRIM(LOWER(p.model)) OR TRIM(LOWER(pr.model)) = TRIM(LOWER(p.sku)))
      WHERE oi.order_id=(SELECT order_id FROM invoices WHERE id=?)
      ORDER BY oi.id
    """, (invoice_id,))
    items = [dict(r) for r in cur.fetchall()]
    c.close()
    return items


def load_company_profile() -> dict:
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM company_profile WHERE id=1")
    row = cur.fetchone()
    c.close()
    return dict(row) if row else {}


def ksef_dir() -> str:
    path = os.path.join(DATA_DIR, "ksef")
    os.makedirs(path, exist_ok=True)
    return path


def ksef_xml_path(invoice_id: int, invoice_no: str) -> str:
    return os.path.join(ksef_dir(), f"{int(invoice_id)}_{xml_filename(invoice_no)}")


def ksef_schema_path() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "fa3_schemat.xsd")


def load_ksef_doc(invoice_id: int) -> dict:
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM ksef_documents WHERE invoice_id=?", (invoice_id,))
    row = cur.fetchone()
    c.close()
    return dict(row) if row else {}


def upsert_ksef_doc(invoice_id: int, status: str, xml_path: str = "", last_error: str = "", ksef_number: str = ""):
    current = load_ksef_doc(invoice_id)
    sent_at = current.get("sent_at", "")
    if status == "sent" and not sent_at:
        sent_at = now_iso()
    c = conn()
    cur = c.cursor()
    cur.execute("""
      INSERT INTO ksef_documents(invoice_id, status, ksef_number, xml_path, last_error, validated_at, sent_at, updated_at)
      VALUES(?,?,?,?,?,?,?,?)
      ON CONFLICT(invoice_id) DO UPDATE SET
        status=excluded.status,
        ksef_number=COALESCE(NULLIF(excluded.ksef_number,''), ksef_documents.ksef_number),
        xml_path=COALESCE(NULLIF(excluded.xml_path,''), ksef_documents.xml_path),
        last_error=excluded.last_error,
        validated_at=excluded.validated_at,
        sent_at=COALESCE(ksef_documents.sent_at, excluded.sent_at),
        updated_at=excluded.updated_at
    """, (
        invoice_id,
        status,
        ksef_number or current.get("ksef_number", ""),
        xml_path or current.get("xml_path", ""),
        last_error or "",
        now_iso(),
        sent_at,
        now_iso(),
    ))
    c.commit()
    c.close()
    try:
        sync_local_rows_to_supabase("ksef_documents", "invoice_id", [invoice_id])
    except Exception:
        pass


def regenerate_invoice_pdf_after_ksef_send(invoice_id: int, ksef_number: str) -> bool:
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        return False
    items = invoice_items_from_saved_json(invoice_id)
    if not items:
        return False

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (inv.get("order_id"),))
    order_row = cur.fetchone()
    c.close()

    meta_payload = invoice_meta_payload(inv)
    meta_payload["ksef_number"] = ksef_number
    pdf_path, total_net, total_gross = generate_order_invoice_pdf(order_row, items, meta_payload)
    packing_pdf_path = generate_invoice_packing_list_pdf(order_row, items, meta_payload, pdf_path)
    stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, inv.get("invoice_no") or f"FV_{invoice_id}", pdf_path, packing_pdf_path)

    c = conn()
    cur = c.cursor()
    cur.execute("UPDATE invoices SET total_net=?, total_gross=? WHERE id=?", (total_net, total_gross, invoice_id))
    c.commit()
    c.close()

    current_meta = load_invoice_meta(invoice_id) or {}
    upsert_invoice_meta(
        invoice_id,
        stored_pdf_path,
        current_meta.get("invoice_items_json") or json.dumps(items, ensure_ascii=False),
        sent_to_client=int(current_meta.get("sent_to_client") or 0),
        seen_by_client=int(current_meta.get("seen_by_client") or 0),
        seen_at=current_meta.get("seen_at"),
        payment_reminder=int(current_meta.get("payment_reminder") or 0),
        paid=int(current_meta.get("paid") or 0),
        paid_at=current_meta.get("paid_at"),
    )

    if supabase_enabled():
        try:
            sync_local_rows_to_supabase("invoices", "id", [invoice_id])
        except Exception:
            pass
        try:
            sync_invoice_meta_to_supabase(invoice_id)
        except Exception:
            pass
    return True


def build_invoice_ksef_payload(invoice_id: int):
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        return None, {}, [], ["Nie znaleziono faktury."]
    company = load_company_profile()
    items = invoice_items_from_saved_json(invoice_id)
    problems = validate_ksef_invoice(inv, company, items)
    return inv, company, items, problems


@app.get("/ksef")
def ksef_dashboard():
    maybe_pull_shared_from_supabase()
    ksef_cfg = ksef_config_summary()
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT i.*, COALESCE(k.status,'draft') AS ksef_status,
             COALESCE(k.ksef_number,'') AS ksef_number,
             COALESCE(k.last_error,'') AS ksef_error,
             COALESCE(k.validated_at,'') AS ksef_validated_at,
             COALESCE(k.sent_at,'') AS ksef_sent_at
      FROM invoices i
      LEFT JOIN ksef_documents k ON k.invoice_id=i.id
      ORDER BY i.issue_date DESC, i.id DESC
      LIMIT 200
    """)
    rows = [dict(r) for r in cur.fetchall()]
    c.close()

    counts = {"draft": 0, "ready": 0, "error": 0, "sent": 0}
    for r in rows:
        counts[r.get("ksef_status") or "draft"] = counts.get(r.get("ksef_status") or "draft", 0) + 1

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">KSeF</h1>
          <span class="badge">FA(3)</span>
        </div>
        <div class="hint" style="margin-top:10px;">
          Generator tworzy XML w strukturze FA(3). Przed wysłaniem sprawdź fakturę przyciskiem „Sprawdź” i przetestuj plik w Aplikacji Podatnika KSeF.
        </div>
        {% if not ksef_cfg.configured %}
          <div class="hint" style="margin-top:10px; border-color:#fecaca; background:#fff1f2;">
            Wysyłka bezpośrednia jest gotowa, ale w Render brakuje: <b>{{ ksef_cfg.missing|join(', ') }}</b>.
          </div>
        {% else %}
          <div class="hint" style="margin-top:10px;">
            Wysyłka bezpośrednia aktywna: <b>{{ ksef_cfg.env }}</b>.
          </div>
        {% endif %}
        <div class="kpi" style="margin-top:10px;">
          <div class="pill">Do sprawdzenia: <b>{{ counts.get('draft',0) }}</b></div>
          <div class="pill">FA(3) OK: <b>{{ counts.get('ready',0) }}</b></div>
          <div class="pill">Do poprawy: <b>{{ counts.get('error',0) }}</b></div>
          <div class="pill">Wysłane: <b>{{ counts.get('sent',0) }}</b></div>
        </div>
      </div>

      <div class="card">
        <table>
          <thead>
            <tr><th>Faktura</th><th>Klient</th><th>Data</th><th>Brutto</th><th>Status KSeF</th><th>Akcje</th></tr>
          </thead>
          <tbody>
            {% for inv in rows %}
              <tr>
                <td><b>{{ inv.invoice_no }}</b></td>
                <td>{{ inv.buyer_name or '-' }}</td>
                <td>{{ inv.issue_date }}</td>
                <td>{{ "%.2f"|format(inv.total_gross or 0) }}</td>
                <td>
                  {% if inv.ksef_status == 'ready' %}
                    <span class="badge ok">FA(3) OK</span>
                  {% elif inv.ksef_status == 'error' %}
                    <span class="badge danger">Do poprawy</span>
                  {% elif inv.ksef_status == 'sent' %}
                    <span class="badge ok">Wysłana</span>
                    {% if inv.ksef_number %}<div class="muted">{{ inv.ksef_number }}</div>{% endif %}
                  {% else %}
                    <span class="badge">Do sprawdzenia</span>
                  {% endif %}
                  {% if inv.ksef_error %}<div class="muted">{{ inv.ksef_error }}</div>{% endif %}
                </td>
                <td>
                  <div class="flex">
                    <form method="post" action="{{ url_for('invoice_ksef_validate', invoice_id=inv.id) }}">
                      <button class="btn" type="submit">Sprawdź</button>
                    </form>
                    <a class="btn primary" href="{{ url_for('invoice_ksef_xml', invoice_id=inv.id) }}">Pobierz XML KSeF FA(3)</a>
                    {% if inv.ksef_status != 'sent' %}
                      <form method="post" action="{{ url_for('invoice_ksef_send', invoice_id=inv.id) }}" onsubmit="return confirm('UWAGA: to jest realna wysyłka faktury do KSeF. Po wysłaniu faktura otrzyma numer KSeF i nie będzie można jej edytować. Kontynuować?');">
                        <input type="hidden" name="next" value="{{ request.full_path }}">
                        <button class="btn primary" type="submit">Wyślij do KSeF</button>
                      </form>
                      <form method="post" action="{{ url_for('invoice_ksef_mark_sent', invoice_id=inv.id) }}" onsubmit="return confirm('Oznaczyć fakturę jako wysłaną do KSeF?');" style="display:flex; gap:6px; flex-wrap:wrap; align-items:center;">
                        <input type="hidden" name="next" value="{{ request.full_path }}">
                        <input name="ksef_number" placeholder="Numer KSeF" style="width:220px;">
                        <button class="btn" type="submit">Oznacz wysłaną</button>
                      </form>
                      <a class="btn" href="{{ url_for('invoice_edit_admin', invoice_id=inv.id) }}">Edytuj fakturę</a>
                    {% else %}
                      <span class="badge ok">Wysłana do KSeF — edycja zablokowana</span>
                      <form method="post" action="{{ url_for('invoice_rollback_admin', invoice_id=inv.id) }}" onsubmit="return confirm('AWARYJNIE cofnąć fakturę {{ inv.invoice_no }} w aplikacji? To usunie lokalny zapis faktury, status KSeF, widoczność u klienta i przeliczy zamówienia oraz stany. Nie usuwa faktury z KSeF.');">
                        <input type="hidden" name="next" value="{{ request.full_path }}">
                        <button class="btn danger" type="submit">Cofnij w aplikacji</button>
                      </form>
                    {% endif %}
                  </div>
                </td>
              </tr>
            {% endfor %}
            {% if not rows %}
              <tr><td colspan="6" class="muted">Brak faktur.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="KSeF", base_url=BASE_URL, db_path=DB_PATH, rows=rows, counts=counts, ksef_cfg=ksef_cfg)


@app.post("/invoices/<int:invoice_id>/ksef/validate")
def invoice_ksef_validate(invoice_id):
    inv, company, items, problems = build_invoice_ksef_payload(invoice_id)
    if not inv:
        return "Nie znaleziono faktury", 404
    if problems:
        upsert_ksef_doc(invoice_id, "error", last_error="; ".join(problems[:5]))
    else:
        xml = build_ksef_draft_xml(inv, company, items)
        path = ksef_xml_path(invoice_id, inv.get("invoice_no") or f"FV_{invoice_id}")
        schema = ksef_schema_path()
        schema_errors = validate_fa3_xml(xml, schema) if os.path.exists(schema) else []
        if schema_errors:
            upsert_ksef_doc(invoice_id, "error", last_error="; ".join(schema_errors[:3]))
            return redirect(request.form.get("next") or url_for("ksef_dashboard"))
        with open(path, "w", encoding="utf-8") as f:
            f.write(xml)
        upsert_ksef_doc(invoice_id, "ready", xml_path=path)
    return redirect(request.form.get("next") or url_for("ksef_dashboard"))


@app.post("/invoices/<int:invoice_id>/ksef/mark-sent")
def invoice_ksef_mark_sent(invoice_id):
    next_url = request.form.get("next") or url_for("ksef_dashboard")
    ksef_number = (request.form.get("ksef_number") or "").strip()
    if not ksef_number:
        upsert_ksef_doc(invoice_id, "error", last_error="Wpisz numer KSeF, żeby oznaczyć fakturę jako wysłaną.")
        return redirect(next_url)
    upsert_ksef_doc(invoice_id, "sent", ksef_number=ksef_number, last_error="")
    try:
        regenerate_invoice_pdf_after_ksef_send(invoice_id, ksef_number)
    except Exception as exc:
        upsert_ksef_doc(invoice_id, "sent", ksef_number=ksef_number, last_error=f"Oznaczono jako wysłaną, ale nie udało się odświeżyć PDF: {exc}")
    return redirect(next_url)


@app.post("/invoices/<int:invoice_id>/ksef/send")
def invoice_ksef_send(invoice_id):
    next_url = request.form.get("next") or url_for("ksef_dashboard")
    current_ksef = load_ksef_doc(invoice_id)
    if current_ksef.get("status") == "sent":
        return redirect(next_url)
    inv, company, items, problems = build_invoice_ksef_payload(invoice_id)
    if not inv:
        return "Nie znaleziono faktury", 404
    if problems:
        upsert_ksef_doc(invoice_id, "error", last_error="; ".join(problems[:5]))
        return redirect(next_url)

    xml = build_ksef_draft_xml(inv, company, items)
    schema = ksef_schema_path()
    schema_errors = validate_fa3_xml(xml, schema) if os.path.exists(schema) else []
    if schema_errors:
        upsert_ksef_doc(invoice_id, "error", last_error="; ".join(schema_errors[:3]))
        return redirect(next_url)

    path = ksef_xml_path(invoice_id, inv.get("invoice_no") or f"FV_{invoice_id}")
    with open(path, "w", encoding="utf-8") as f:
        f.write(xml)

    if send_invoice_to_ksef is None:
        upsert_ksef_doc(invoice_id, "error", xml_path=path, last_error="Brak modułu ksef_api.py albo zależności requests/cryptography.")
        return redirect(next_url)

    result = send_invoice_to_ksef(xml)
    if result.get("ok"):
        ksef_number = result.get("ksef_number") or (f"ref: {result.get('invoice_reference_number')}" if result.get("invoice_reference_number") else "")
        upsert_ksef_doc(invoice_id, "sent", xml_path=path, ksef_number=ksef_number)
        try:
            regenerate_invoice_pdf_after_ksef_send(invoice_id, ksef_number)
        except Exception as exc:
            upsert_ksef_doc(invoice_id, "sent", xml_path=path, ksef_number=ksef_number, last_error=f"Wysłano do KSeF, ale nie udało się odświeżyć PDF: {exc}")
    else:
        upsert_ksef_doc(invoice_id, "error", xml_path=path, last_error=result.get("message") or "Nie udało się wysłać faktury do KSeF.")
    return redirect(next_url)


@app.get("/invoices/<int:invoice_id>/ksef/xml")
def invoice_ksef_xml(invoice_id):
    inv, company, items, problems = build_invoice_ksef_payload(invoice_id)
    if not inv:
        return "Nie znaleziono faktury", 404
    if problems:
        upsert_ksef_doc(invoice_id, "error", last_error="; ".join(problems[:5]))
        return "Nie można wygenerować XML KSeF:\n- " + "\n- ".join(problems), 400

    xml = build_ksef_draft_xml(inv, company, items)
    schema = ksef_schema_path()
    schema_errors = validate_fa3_xml(xml, schema) if os.path.exists(schema) else []
    if schema_errors:
        upsert_ksef_doc(invoice_id, "error", last_error="; ".join(schema_errors[:3]))
        return "XML nie przeszedł walidacji FA(3):\n- " + "\n- ".join(schema_errors), 400

    path = ksef_xml_path(invoice_id, inv.get("invoice_no") or f"FV_{invoice_id}")
    with open(path, "w", encoding="utf-8") as f:
        f.write(xml)
    upsert_ksef_doc(invoice_id, "ready", xml_path=path)

    return send_file(path, mimetype="application/xml", as_attachment=True, download_name=xml_filename(inv.get("invoice_no") or f"FV_{invoice_id}"))


@app.get("/invoices/<int:invoice_id>/download")
def invoice_download_admin(invoice_id):
    row = load_invoice_with_meta(invoice_id)
    if not row:
        return "Nie znaleziono faktury", 404

    if parse_supabase_storage_ref(row.get("pdf_path", "")):
        try:
            data, filename = supabase_storage_download_bytes(row.get("pdf_path", ""))
            return send_file(io.BytesIO(data), mimetype="application/pdf", as_attachment=True, download_name=filename)
        except Exception:
            pass

    ok_pdf, abs_path = invoice_pdf_exists(row.get("pdf_path", ""), row.get("invoice_no", ""))
    if not ok_pdf:
        c = conn()
        cur = c.cursor()
        cur.execute("SELECT * FROM orders WHERE id=?", (row["order_id"],))
        o = cur.fetchone()
        c.close()
        if not o:
            return "Brak powiązanego zamówienia", 404

        items = invoice_items_from_saved_json(invoice_id)
        if not items:
            return "Brak pozycji faktury", 400

        meta = invoice_meta_payload(row)
        abs_path, total_net, total_gross = generate_order_invoice_pdf(o, items, meta)
        packing_pdf_path = generate_invoice_packing_list_pdf(o, items, meta, abs_path)
        stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, row.get("invoice_no") or f"FV_{invoice_id}", abs_path, packing_pdf_path)

        c = conn()
        cur = c.cursor()
        cur.execute("UPDATE invoices SET total_net=?, total_gross=? WHERE id=?", (total_net, total_gross, invoice_id))
        c.commit()
        c.close()

        current_meta = load_invoice_meta(invoice_id) or {}
        upsert_invoice_meta(
            invoice_id,
            stored_pdf_path,
            current_meta.get("invoice_items_json") or json.dumps(items, ensure_ascii=False),
            sent_to_client=int(current_meta.get("sent_to_client") or 0),
            seen_by_client=int(current_meta.get("seen_by_client") or 0),
            seen_at=current_meta.get("seen_at")
        )

        if supabase_enabled():
            try:
                sync_local_rows_to_supabase("invoices", "id", [invoice_id])
            except Exception:
                pass
            try:
                sync_invoice_meta_to_supabase(invoice_id)
            except Exception:
                pass
        if parse_supabase_storage_ref(stored_pdf_path):
            data, filename = supabase_storage_download_bytes(stored_pdf_path)
            return send_file(io.BytesIO(data), mimetype="application/pdf", as_attachment=True, download_name=filename)

    if supabase_enabled() and abs_path and os.path.exists(abs_path) and not parse_supabase_storage_ref(row.get("pdf_path", "")):
        try:
            items = invoice_items_from_saved_json(invoice_id)
            packing_pdf_path = ""
            if items:
                pack_candidate = packing_list_pdf_path_for_invoice(abs_path, row.get("invoice_no") or f"FV_{invoice_id}")
                if os.path.exists(pack_candidate):
                    packing_pdf_path = pack_candidate
            stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, row.get("invoice_no") or f"FV_{invoice_id}", abs_path, packing_pdf_path)
            current_meta = load_invoice_meta(invoice_id) or {}
            upsert_invoice_meta(
                invoice_id,
                stored_pdf_path,
                current_meta.get("invoice_items_json") or (json.dumps(items, ensure_ascii=False) if items else ""),
                sent_to_client=int(current_meta.get("sent_to_client") or 0),
                seen_by_client=int(current_meta.get("seen_by_client") or 0),
                seen_at=current_meta.get("seen_at")
            )
            sync_invoice_meta_to_supabase(invoice_id)
            data, filename = supabase_storage_download_bytes(stored_pdf_path)
            return send_file(io.BytesIO(data), mimetype="application/pdf", as_attachment=True, download_name=filename)
        except Exception:
            pass
    return send_file(abs_path, mimetype="application/pdf", as_attachment=True, download_name=os.path.basename(abs_path))


@app.get("/invoices/<int:invoice_id>/packing-list")
def invoice_packing_list_download_admin(invoice_id):
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        return "Nie znaleziono faktury", 404

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (inv["order_id"],))
    o = cur.fetchone()
    c.close()
    if not o:
        return "Brak powiązanego zamówienia", 404

    items = invoice_items_from_saved_json(invoice_id)
    if not items:
        return "Brak pozycji faktury", 400

    ok_pdf, invoice_abs_path = invoice_pdf_exists(inv.get("pdf_path", ""), inv.get("invoice_no", ""))
    pack_path = packing_list_pdf_path_for_invoice(invoice_abs_path if ok_pdf else "", inv.get("invoice_no") or f"FV_{invoice_id}")
    pack_path = generate_invoice_packing_list_pdf(o, items, invoice_meta_payload(inv), invoice_abs_path if ok_pdf else "")
    if supabase_enabled():
        try:
            packing_ref = supabase_storage_upload_file(
                pack_path,
                invoice_packing_storage_object_path(invoice_id, inv.get("invoice_no") or f"FV_{invoice_id}"),
                content_type="application/pdf",
            )
            data, filename = supabase_storage_download_bytes(packing_ref)
            return send_file(io.BytesIO(data), mimetype="application/pdf", as_attachment=True, download_name=filename)
        except Exception:
            pass

    return send_file(pack_path, mimetype="application/pdf", as_attachment=True, download_name=os.path.basename(pack_path))

@app.post("/invoices/<int:invoice_id>/regenerate")
def invoice_regenerate_admin(invoice_id):
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        return "Nie znaleziono faktury", 404

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (inv["order_id"],))
    o = cur.fetchone()
    c.close()
    if not o:
        return "Brak powiÄ…zanego zamĂłwienia", 404

    items = invoice_items_from_saved_json(invoice_id)
    if not items:
        return "Brak pozycji faktury", 400

    meta = invoice_meta_payload(inv)
    pdf_path, total_net, total_gross = generate_order_invoice_pdf(o, items, meta)
    packing_pdf_path = generate_invoice_packing_list_pdf(o, items, meta, pdf_path)
    stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, inv["invoice_no"], pdf_path, packing_pdf_path)

    c = conn()
    cur = c.cursor()
    cur.execute("UPDATE invoices SET total_net=?, total_gross=? WHERE id=?", (total_net, total_gross, invoice_id))
    c.commit()
    c.close()

    current_meta = load_invoice_meta(invoice_id) or {}
    upsert_invoice_meta(
        invoice_id,
        stored_pdf_path,
        current_meta.get("invoice_items_json") or json.dumps(items, ensure_ascii=False),
        sent_to_client=int(current_meta.get("sent_to_client") or 0),
        seen_by_client=int(current_meta.get("seen_by_client") or 0),
        seen_at=current_meta.get("seen_at")
    )

    if supabase_enabled():
        try:
            sync_local_rows_to_supabase("invoices", "id", [invoice_id])
        except Exception:
            pass
        try:
            sync_invoice_meta_to_supabase(invoice_id)
        except Exception:
            pass

    return redirect(request.referrer or url_for("orders"))


def _redirect_after_invoice_action(default_endpoint="invoices"):
    target = norm(request.values.get("next")) or request.referrer or url_for(default_endpoint)
    return redirect(target)


def _set_invoice_payment_state(invoice_id: int, *, reminder: int | None = None, paid: int | None = None):
    meta = load_invoice_meta(invoice_id) or {}
    pdf_path = meta.get("pdf_path", "")
    items_json = meta.get("invoice_items_json", "")
    sent_to_client = int(meta.get("sent_to_client") or 0)
    seen_by_client = int(meta.get("seen_by_client") or 0)
    seen_at = meta.get("seen_at")
    current_reminder = int(meta.get("payment_reminder") or 0)
    current_paid = int(meta.get("paid") or 0)
    current_paid_at = meta.get("paid_at")

    next_paid = current_paid if paid is None else int(paid)
    next_reminder = current_reminder if reminder is None else int(reminder)
    next_paid_at = current_paid_at
    if next_paid:
        next_reminder = 0
        next_paid_at = now_iso()
    elif paid == 0:
        next_paid_at = None

    upsert_invoice_meta(
        invoice_id,
        pdf_path,
        items_json,
        sent_to_client=sent_to_client,
        seen_by_client=seen_by_client,
        seen_at=seen_at,
        payment_reminder=next_reminder,
        paid=next_paid,
        paid_at=next_paid_at
    )
    if supabase_enabled():
        try:
            sync_invoice_meta_to_supabase(invoice_id)
        except Exception:
            pass


@app.post("/invoices/<int:invoice_id>/paid")
def invoice_paid_admin(invoice_id):
    _set_invoice_payment_state(invoice_id, reminder=0, paid=1)
    return _redirect_after_invoice_action()


@app.post("/invoices/<int:invoice_id>/unpaid")
def invoice_unpaid_admin(invoice_id):
    _set_invoice_payment_state(invoice_id, reminder=0, paid=0)
    return _redirect_after_invoice_action()

@app.post("/api/invoices/<int:invoice_id>/seen")
def api_invoice_seen(invoice_id):
    email = _email_key(g.client_user["email"])
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT
        i.id,
        m.invoice_id AS meta_invoice_id,
        i.buyer_email,
        o.customer_email AS order_customer_email,
        COALESCE(m.pdf_path,'') AS pdf_path,
        COALESCE(m.sent_to_client,0) AS sent_to_client,
        i.invoice_no
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      LEFT JOIN orders o ON o.id = i.order_id
      WHERE i.id=?
      LIMIT 1
    """, (invoice_id,))
    row = cur.fetchone()
    c.close()
    if not row:
        return jsonify(ok=False, error="Nie znaleziono faktury"), 404

    if email:
        buyer_ok = _email_key(row["buyer_email"]) == email
        order_ok = _email_key(row["order_customer_email"]) == email
        has_meta = row["meta_invoice_id"] is not None
        if (has_meta and int(row["sent_to_client"] or 0) != 1) or not (buyer_ok or order_ok):
            return jsonify(ok=False, error="Brak dostÄ™pu"), 403

    ok_pdf, _ = invoice_pdf_exists(row["pdf_path"], row["invoice_no"])
    if not ok_pdf:
        return jsonify(ok=False, error="Brak pliku PDF"), 404

    meta = load_invoice_meta(invoice_id) or {}
    ts = now_iso()
    upsert_invoice_meta(
        invoice_id,
        meta.get("pdf_path",""),
        meta.get("invoice_items_json",""),
        sent_to_client=int(meta.get("sent_to_client") or 0),
        seen_by_client=1,
        seen_at=ts
    )

    if supabase_enabled():
        try:
            sync_invoice_meta_to_supabase(invoice_id)
        except Exception:
            pass

    return jsonify(ok=True, seen_at=ts)

@app.get("/api/invoices/<int:invoice_id>/download")
def api_invoice_download(invoice_id):
    maybe_pull_shared_from_supabase()
    email = _email_key(g.client_user["email"])
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT
        i.*,
        m.invoice_id AS meta_invoice_id,
        COALESCE(m.pdf_path,'') AS pdf_path,
        COALESCE(m.sent_to_client,0) AS sent_to_client,
        o.customer_email AS order_customer_email
      FROM invoices i
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      LEFT JOIN orders o ON o.id = i.order_id
      WHERE i.id=?
      LIMIT 1
    """, (invoice_id,))
    row = cur.fetchone()
    c.close()
    if not row:
        return "Nie znaleziono faktury", 404

    if email:
        buyer_ok = _email_key(row["buyer_email"]) == email
        order_ok = _email_key(row["order_customer_email"]) == email
        has_meta = row["meta_invoice_id"] is not None
        if (has_meta and int(row["sent_to_client"] or 0) != 1) or not (buyer_ok or order_ok):
            return "Brak dostÄ™pu", 403

    def mark_downloaded_by_client():
        if not email:
            return
        meta = load_invoice_meta(invoice_id) or {}
        upsert_invoice_meta(
            invoice_id,
            meta.get("pdf_path", ""),
            meta.get("invoice_items_json", ""),
            sent_to_client=int(meta.get("sent_to_client") or 0),
            seen_by_client=1,
            seen_at=now_iso(),
            payment_reminder=int(meta.get("payment_reminder") or 0),
            paid=int(meta.get("paid") or 0),
            paid_at=meta.get("paid_at")
        )
        if supabase_enabled():
            try:
                sync_invoice_meta_to_supabase(invoice_id)
            except Exception:
                pass

    if parse_supabase_storage_ref(row["pdf_path"]):
        try:
            data, filename = supabase_storage_download_bytes(row["pdf_path"])
            mark_downloaded_by_client()
            return send_file(io.BytesIO(data), mimetype="application/pdf", as_attachment=True, download_name=filename)
        except Exception:
            pass

    ok_pdf, abs_path = invoice_pdf_exists(row["pdf_path"], row["invoice_no"])
    if not ok_pdf:
        cur_order = None
        c = conn()
        cur = c.cursor()
        cur.execute("SELECT * FROM orders WHERE id=?", (row["order_id"],))
        cur_order = cur.fetchone()
        c.close()
        if not cur_order:
            return "Brak powiązanego zamówienia", 404
        items = invoice_items_from_saved_json(invoice_id)
        if not items:
            return "Brak pozycji faktury", 400
        meta = invoice_meta_payload(dict(row))
        abs_path, total_net, total_gross = generate_order_invoice_pdf(cur_order, items, meta)
        packing_pdf_path = generate_invoice_packing_list_pdf(cur_order, items, meta, abs_path)
        stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, row["invoice_no"], abs_path, packing_pdf_path)
        current_meta = load_invoice_meta(invoice_id) or {}
        upsert_invoice_meta(
            invoice_id,
            stored_pdf_path,
            current_meta.get("invoice_items_json") or json.dumps(items, ensure_ascii=False),
            sent_to_client=int(current_meta.get("sent_to_client") or 0),
            seen_by_client=int(current_meta.get("seen_by_client") or 0),
            seen_at=current_meta.get("seen_at")
        )
        if supabase_enabled():
            try:
                sync_invoice_meta_to_supabase(invoice_id)
            except Exception:
                pass
        if parse_supabase_storage_ref(stored_pdf_path):
            data, filename = supabase_storage_download_bytes(stored_pdf_path)
            mark_downloaded_by_client()
            return send_file(io.BytesIO(data), mimetype="application/pdf", as_attachment=True, download_name=filename)

    if supabase_enabled() and abs_path and os.path.exists(abs_path) and not parse_supabase_storage_ref(row["pdf_path"]):
        try:
            items = invoice_items_from_saved_json(invoice_id)
            packing_pdf_path = ""
            if items:
                pack_candidate = packing_list_pdf_path_for_invoice(abs_path, row["invoice_no"])
                if os.path.exists(pack_candidate):
                    packing_pdf_path = pack_candidate
            stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, row["invoice_no"], abs_path, packing_pdf_path)
            current_meta = load_invoice_meta(invoice_id) or {}
            upsert_invoice_meta(
                invoice_id,
                stored_pdf_path,
                current_meta.get("invoice_items_json") or (json.dumps(items, ensure_ascii=False) if items else ""),
                sent_to_client=int(current_meta.get("sent_to_client") or 0),
                seen_by_client=int(current_meta.get("seen_by_client") or 0),
                seen_at=current_meta.get("seen_at")
            )
            sync_invoice_meta_to_supabase(invoice_id)
            data, filename = supabase_storage_download_bytes(stored_pdf_path)
            mark_downloaded_by_client()
            return send_file(io.BytesIO(data), mimetype="application/pdf", as_attachment=True, download_name=filename)
        except Exception:
            pass

    try:
        mark_downloaded_by_client()
        return send_file(abs_path, mimetype="application/pdf", as_attachment=True, download_name=os.path.basename(abs_path))
    except Exception as e:
        return f"BĹ‚Ä…d pobierania PDF: {e}", 500

def _delete_invoice_everywhere(invoice_id: int):
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        abort(404)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT DISTINCT order_id FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    touched_order_ids = [int(r["order_id"]) for r in cur.fetchall()]
    meta_items_raw = inv.get("invoice_items_json") or ""
    if meta_items_raw:
        try:
            meta_items = json.loads(meta_items_raw)
            if isinstance(meta_items, list):
                for it in meta_items:
                    oid = int(it.get("source_order_id") or it.get("order_id") or 0)
                    if oid and oid not in touched_order_ids:
                        touched_order_ids.append(oid)
        except Exception:
            pass
    if int(inv.get("order_id") or 0) and int(inv.get("order_id") or 0) not in touched_order_ids:
        touched_order_ids.append(int(inv.get("order_id") or 0))
    c.close()

    ok_pdf, abs_path = invoice_pdf_exists(inv.get("pdf_path", ""), inv.get("invoice_no", ""))
    try:
        pack_path = packing_list_pdf_path_for_invoice(abs_path if ok_pdf else "", inv.get("invoice_no", ""))
        if pack_path and os.path.exists(pack_path):
            os.remove(pack_path)
        if ok_pdf and abs_path and os.path.exists(abs_path):
            os.remove(abs_path)
    except Exception:
        pass

    c = conn()
    cur = c.cursor()
    cur.execute("DELETE FROM invoice_allocations WHERE invoice_id=?", (invoice_id,))
    cur.execute("DELETE FROM invoice_meta WHERE invoice_id=?", (invoice_id,))
    cur.execute("DELETE FROM ksef_documents WHERE invoice_id=?", (invoice_id,))
    cur.execute("DELETE FROM invoices WHERE id=?", (invoice_id,))
    c.commit()
    c.close()

    changed_order_ids, changed_product_ids = reconcile_orders_after_invoice_change(touched_order_ids)

    if supabase_enabled():
        try:
            supabase_delete_rows("invoice_allocations", {"invoice_id": invoice_id})
        except Exception:
            pass
        try:
            supabase_delete_rows("invoice_meta", {"invoice_id": invoice_id})
        except Exception:
            pass
        try:
            supabase_delete_rows("ksef_documents", {"invoice_id": invoice_id})
        except Exception:
            pass
        try:
            supabase_delete_rows("invoices", {"id": invoice_id})
        except Exception:
            pass
        if changed_order_ids:
            try:
                sync_local_rows_to_supabase("orders", "id", changed_order_ids)
            except Exception:
                pass
        if changed_product_ids:
            try:
                sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
            except Exception:
                pass

    return inv


@app.post("/invoices/<int:invoice_id>/delete")
def invoice_delete_admin(invoice_id):
    _delete_invoice_everywhere(invoice_id)
    return _redirect_after_invoice_action()


@app.post("/invoices/<int:invoice_id>/rollback")
def invoice_rollback_admin(invoice_id):
    _delete_invoice_everywhere(invoice_id)
    return _redirect_after_invoice_action()


@app.post("/orders/<int:order_id>/invoice/<int:invoice_id>/delete")
def order_invoice_delete(order_id, invoice_id):
    inv = load_invoice_with_meta(invoice_id)
    if not inv or int(inv.get("order_id") or 0) != int(order_id):
        abort(404)
    _delete_invoice_everywhere(invoice_id)

    return redirect(url_for("order_invoice", order_id=order_id, deleted="1"))


def _invoice_email_context(invoice_id: int):
    c = conn()
    cur = c.cursor()
    cur.execute("""
      SELECT
        i.*,
        o.order_no AS order_no,
        o.customer_email AS customer_email,
        o.customer_name AS customer_name,
        m.pdf_path AS pdf_path,
        m.payment_reminder AS payment_reminder,
        m.paid AS paid,
        m.seen_by_client AS seen_by_client,
        m.seen_at AS seen_at
      FROM invoices i
      LEFT JOIN orders o ON o.id = i.order_id
      LEFT JOIN invoice_meta m ON m.invoice_id = i.id
      WHERE i.id=?
      LIMIT 1
    """, (invoice_id,))
    row = cur.fetchone()
    c.close()
    if not row:
        abort(404)

    invoice = dict(row)
    email = _email_key(invoice.get("buyer_email") or invoice.get("customer_email"))
    if email:
        pdf_url = build_public_url(f"/api/invoices/{invoice_id}/download?email={urllib.parse.quote_plus(email)}")
    else:
        pdf_url = build_public_url(f"/api/invoices/{invoice_id}/download")
    return invoice, pdf_url


@app.route("/invoices/<int:invoice_id>/edit", methods=["GET", "POST"])
def invoice_edit_admin(invoice_id):
    inv = load_invoice_with_meta(invoice_id)
    if not inv:
        return "Nie znaleziono faktury", 404
    ksef_doc = load_ksef_doc(invoice_id)
    if ksef_doc.get("status") == "sent":
        tpl = r"""
        {% extends "base.html" %}
        {% block content %}
          <div class="card">
            <div class="flex">
              <h1 style="margin:0;">Faktura wysłana do KSeF</h1>
              <a class="btn right" href="{{ url_for('invoices') }}">← Faktury</a>
            </div>
            <div class="hint" style="margin-top:10px;">
              Ta faktura ma już numer KSeF i jej edycja została zablokowana, żeby nie powstała różnica między aplikacją a KSeF.
            </div>
            {% if ksef_doc.ksef_number %}
              <p><b>Numer KSeF:</b> {{ ksef_doc.ksef_number }}</p>
            {% endif %}
            <div class="flex" style="margin-top:12px;">
              <a class="btn" href="{{ url_for('invoice_download_admin', invoice_id=inv.id) }}" target="_blank">Faktura PDF</a>
              <a class="btn" href="{{ url_for('invoice_ksef_xml', invoice_id=inv.id) }}">XML KSeF FA(3)</a>
              <a class="btn" href="{{ url_for('invoices') }}">Wróć do faktur</a>
            </div>
          </div>
        {% endblock %}
        """
        return render_template_string(tpl, title="Faktura wysłana do KSeF", base_url=BASE_URL, db_path=DB_PATH, inv=inv, ksef_doc=ksef_doc)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM orders WHERE id=?", (inv["order_id"],))
    order_row = cur.fetchone()
    c.close()

    edit_items = invoice_edit_items(invoice_id, dict(inv))

    msg = ""
    if request.method == "POST":
        data = {k: norm(request.form.get(k)) for k in [
            "invoice_no", "issue_date", "sell_date", "payment_type", "payment_to",
            "buyer_name", "buyer_tax_no", "buyer_address", "buyer_country",
            "buyer_email", "buyer_phone"
        ]}
        invoice_items = prepare_invoice_edit_items(edit_items, request.form)
        existing_invoice_id = invoice_no_exists(data["invoice_no"], invoice_id)
        if not data["invoice_no"]:
            msg = "Numer faktury jest wymagany."
        elif existing_invoice_id:
            msg = f"Faktura o takim numerze już istnieje! Numer: {data['invoice_no']}. Wybierz inny numer faktury."
        elif not invoice_items:
            msg = "Faktura musi zawierać co najmniej jedną pozycję."
        else:
            old_order_ids = sorted({int(x.get("source_order_id") or x.get("order_id") or 0) for x in edit_items if int(x.get("current_invoice_qty") or 0) > 0})
            st, pc, city = split_address(data.get("buyer_address", ""))
            c = conn()
            cur = c.cursor()
            cur.execute("""
              UPDATE invoices
              SET invoice_no=?, issue_date=?, sell_date=?, payment_type=?, payment_to=?,
                  buyer_name=?, buyer_tax_no=?, buyer_street=?, buyer_post_code=?, buyer_city=?,
                  buyer_country=?, buyer_email=?, buyer_phone=?
              WHERE id=?
            """, (
                data["invoice_no"], data["issue_date"], data["sell_date"], data["payment_type"], data["payment_to"],
                data["buyer_name"], data["buyer_tax_no"], st, pc, city,
                data["buyer_country"], data["buyer_email"], data["buyer_phone"], invoice_id
            ))
            c.commit()
            c.close()

            updated = load_invoice_with_meta(invoice_id)
            if invoice_items and updated:
                order_for_pdf = order_row
                if not order_for_pdf:
                    first_order_id = int(invoice_items[0].get("source_order_id") or invoice_items[0].get("order_id") or 0)
                    if first_order_id:
                        c = conn()
                        cur = c.cursor()
                        cur.execute("SELECT * FROM orders WHERE id=?", (first_order_id,))
                        order_for_pdf = cur.fetchone()
                        c.close()

                pdf_path, total_net, total_gross = generate_order_invoice_pdf(order_for_pdf, invoice_items, invoice_meta_payload(updated))
                packing_pdf_path = generate_invoice_packing_list_pdf(order_for_pdf, invoice_items, invoice_meta_payload(updated), pdf_path)
                stored_pdf_path = upload_invoice_pdfs_to_supabase(invoice_id, data["invoice_no"], pdf_path, packing_pdf_path)
                allocation_ids = replace_invoice_allocations(invoice_id, invoice_items)
                new_order_ids = sorted({int(x.get("source_order_id") or x.get("order_id") or 0) for x in invoice_items})
                touched_order_ids = sorted(set(old_order_ids + new_order_ids))
                changed_order_ids, changed_product_ids = reconcile_orders_after_invoice_change(touched_order_ids)

                c = conn()
                cur = c.cursor()
                cur.execute("UPDATE invoices SET total_net=?, total_gross=? WHERE id=?", (total_net, total_gross, invoice_id))
                c.commit()
                c.close()

                meta = load_invoice_meta(invoice_id) or {}
                upsert_invoice_meta(
                    invoice_id,
                    stored_pdf_path,
                    json.dumps(invoice_items, ensure_ascii=False),
                    sent_to_client=int(meta.get("sent_to_client") or 0),
                    seen_by_client=0,
                    seen_at=None,
                    payment_reminder=int(meta.get("payment_reminder") or 0),
                    paid=int(meta.get("paid") or 0),
                    paid_at=meta.get("paid_at")
                )

            if supabase_enabled():
                try:
                    sync_local_rows_to_supabase("invoices", "id", [invoice_id])
                except Exception:
                    pass
                try:
                    sync_invoice_meta_to_supabase(invoice_id)
                except Exception:
                    pass
                try:
                    supabase_delete_rows("invoice_allocations", {"invoice_id": invoice_id})
                except Exception:
                    pass
                if allocation_ids:
                    try:
                        sync_local_rows_to_supabase("invoice_allocations", "id", allocation_ids)
                    except Exception:
                        pass
                if changed_order_ids:
                    try:
                        sync_local_rows_to_supabase("orders", "id", changed_order_ids)
                    except Exception:
                        pass
                if changed_product_ids:
                    try:
                        sync_local_rows_to_supabase("stock", "product_id", changed_product_ids)
                    except Exception:
                        pass

            return redirect(url_for("invoices", edited="1", invoice_id=invoice_id))

    buyer_address = "\n".join([x for x in [
        inv.get("buyer_street") or "",
        " ".join([inv.get("buyer_post_code") or "", inv.get("buyer_city") or ""]).strip()
    ] if x])

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Edytuj fakturę {{ inv.invoice_no }}</h1>
          <a class="btn right" href="{{ url_for('invoices') }}">← Faktury</a>
        </div>
        {% if msg %}<div class="hint" style="margin-top:10px;">{{ msg }}</div>{% endif %}
      </div>

      <div class="card">
        <form method="post" class="row">
          <div><label class="muted small">Numer faktury</label><input name="invoice_no" value="{{ inv.invoice_no }}" required></div>
          <div><label class="muted small">Data wystawienia</label><input name="issue_date" type="date" value="{{ inv.issue_date }}"></div>
          <div><label class="muted small">Data sprzedaży</label><input name="sell_date" type="date" value="{{ inv.sell_date }}"></div>
          <div><label class="muted small">Forma płatności</label>
            <select name="payment_type">
              <option value="gotowka" {% if inv.payment_type in ['cash','gotowka'] %}selected{% endif %}>gotówka</option>
              <option value="przelew" {% if inv.payment_type in ['transfer','przelew'] %}selected{% endif %}>przelew</option>
              <option value="karta" {% if inv.payment_type in ['card','karta'] %}selected{% endif %}>karta</option>
            </select>
          </div>
          <div><label class="muted small">Termin płatności</label><input name="payment_to" type="date" value="{{ inv.payment_to }}"></div>
          <div><label class="muted small">Nabywca</label><input name="buyer_name" value="{{ inv.buyer_name }}"></div>
          <div><label class="muted small">NIP nabywcy</label><input name="buyer_tax_no" value="{{ inv.buyer_tax_no }}"></div>
          <div><label class="muted small">Adres nabywcy</label><textarea name="buyer_address" placeholder="Ulica&#10;Kod pocztowy Miasto">{{ buyer_address }}</textarea></div>
          <div><label class="muted small">Kraj</label><input name="buyer_country" value="{{ inv.buyer_country or 'PL' }}"></div>
          <div><label class="muted small">Email</label><input name="buyer_email" value="{{ inv.buyer_email }}"></div>
          <div><label class="muted small">Telefon</label><input name="buyer_phone" value="{{ inv.buyer_phone }}"></div>
          <div style="grid-column:1/-1;">
            <h2>Pozycje faktury</h2>
            <div class="hint" style="margin-bottom:10px;">
              Zmień ilości pozycji na tej fakturze. Wpisanie 0 usuwa pozycję z faktury.
            </div>
            <table>
              <thead>
                <tr>
                  <th>Zamówienie</th>
                  <th>Notatka</th>
                  <th>SKU</th>
                  <th>Model / Nazwa</th>
                  <th>Zamówiono</th>
                  <th>Na innych fakturach</th>
                  <th>Maks. na tej fakturze</th>
                  <th>Ilość na fakturze</th>
                  <th>Netto/szt</th>
                  <th>Brutto/szt</th>
                </tr>
              </thead>
              <tbody>
                {% for it in edit_items %}
                  <tr>
                    <td><b>{{ it.source_order_no }}</b></td>
                    <td>{{ it.source_order_note or '-' }}</td>
                    <td>{{ it.sku }}</td>
                    <td>{{ it.model or '' }}{% if it.name %}<div class="muted small">{{ it.name }}</div>{% endif %}</td>
                    <td>{{ it.ordered_qty }}</td>
                    <td>{{ it.invoiced_other_qty }}</td>
                    <td><b>{{ it.remaining_qty }}</b></td>
                    <td><input type="number" min="0" max="{{ it.remaining_qty }}" name="invoice_qty_{{ it.id }}" value="{{ it.current_invoice_qty }}" style="width:110px;"></td>
                    <td>{{ "%.2f"|format(it.net_price) }}</td>
                    <td>{{ "%.2f"|format(it.gross_price) }}</td>
                  </tr>
                {% endfor %}
              </tbody>
            </table>
          </div>
          <div style="grid-column:1/-1;" class="flex">
            <button class="btn primary" type="submit">Zapisz i regeneruj PDF</button>
            <a class="btn" href="{{ url_for('invoice_download_admin', invoice_id=inv.id) }}" target="_blank">Podgląd PDF</a>
          </div>
        </form>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Edytuj fakturę", base_url=BASE_URL, db_path=DB_PATH, inv=inv, buyer_address=buyer_address, msg=msg, edit_items=edit_items)

# -------------------------
# ZAMÓWIENIA MATERIAŁÓW
# -------------------------

@app.get("/material-orders")
def material_orders():
    # WyĹ‚Ä…czony pull z Supabase tylko dla moduĹ‚u Zamówienia materiałów.
    # Tu pracujemy na lokalnej bazie, ĹĽeby POST -> redirect nie cofaĹ‚ zmian.
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM material_orders ORDER BY id DESC LIMIT 200")
    packs = cur.fetchall()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Zamówienia materiałów</h1>
        </div>
        <div class="muted">ZarzÄ…dzaj przesyĹ‚kami: status, tracking i zawartoĹ›Ä‡ paczki. Tracking otwiera 17TRACK.</div>
      </div>

      <div class="card">
        <h2>Nowe zamówienie materiałów</h2>
        <form method="post" action="{{ url_for('material_order_create') }}" class="row">
          <div>
            <label class="muted small">Numer zamówienia materiałów</label>
            <input name="package_no" placeholder="np. PO-2026-02-01" required>
          </div>
          <div>
            <label class="muted small">Notatka</label>
            <input name="note">
          </div>
          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Zapisz</button>
          </div>
        </form>
      </div>

      <div class="card">
        <h2>Zamówienia materiałów (max 200)</h2>
        <table>
          <thead>
            <tr><th>Nr</th><th>Status</th><th>Notatka</th><th>Data</th><th>Akcje</th></tr>
          </thead>
          <tbody>
            {% for p in packs %}
              <tr>
                <td><b>{{ p['package_no'] }}</b></td>
                <td>
                  <form method="post" action="{{ url_for('material_order_status', package_id=p['id']) }}" class="flex">
                    <select name="status" style="width:140px;">
                      <option value="planned" {% if p['status']=='planned' %}selected{% endif %}>planned</option>
                      <option value="ordered" {% if p['status']=='ordered' %}selected{% endif %}>ordered</option>
                      <option value="shipped" {% if p['status']=='shipped' %}selected{% endif %}>shipped</option>
                      <option value="arrived" {% if p['status']=='arrived' %}selected{% endif %}>arrived</option>
                    </select>
                    <button class="btn" type="submit">ZmieĹ„</button>
                  </form>
                </td>
                <td>{{ p['note'] or "-" }}</td>
                <td class="muted">{{ p['created_at'] }}</td>
                <td class="flex">
                  <a class="btn primary" href="{{ url_for('material_order_detail', package_id=p['id']) }}">Pozycje</a>
                  <a class="btn" target="_blank" href="{{ url_for('material_order_print', package_id=p['id']) }}">Drukuj</a>
                  <form method="post" action="{{ url_for('material_order_delete', package_id=p['id']) }}" onsubmit="return confirm('UsunÄ…Ä‡ paczkÄ™?')">
                    <button class="btn danger" type="submit">UsuĹ„</button>
                  </form>
                </td>
              </tr>
            {% endfor %}
            {% if not packs %}
              <tr><td colspan="6" class="muted">Brak zamówień materiałów.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title="Zamówienia materiałów", base_url=BASE_URL, db_path=DB_PATH, packs=packs)

@app.post("/material-orders/create")
def material_order_create():
    package_no = norm(request.form.get("package_no"))
    status = "draft"
    note = norm(request.form.get("note"))

    if not package_no:
        return "Brak numeru zamówienia materiałów", 400

    c = conn()
    cur = c.cursor()
    try:
        cur.execute("""
          INSERT INTO material_orders(package_no, status, tracking, note, created_at)
          VALUES(?,?,?,?,?)
        """, (package_no, status, "", note, now_iso()))
        c.commit()
    except sqlite3.IntegrityError:
        pass
    finally:
        c.close()

    return redirect(url_for("material_orders"))

@app.post("/material-orders/<int:package_id>/status")
def material_order_status(package_id):
    status = norm(request.form.get("status"))
    if status not in {"planned", "ordered", "shipped", "arrived"}:
        return "NieprawidĹ‚owy status", 400

    c = conn()
    cur = c.cursor()

    cur.execute("SELECT status FROM material_orders WHERE id=?", (package_id,))
    pack = cur.fetchone()
    if not pack:
        c.close()
        abort(404)

    old_status = pack["status"]

    cur.execute("SELECT product_id, qty FROM material_order_items WHERE package_id=?", (package_id,))
    items = cur.fetchall()

    # PrzejĹ›cie NA arrived: fizycznie przyjÄ™to towar -> dodaj na stan.
    if INVENTORY_AUTOMATION_ENABLED and old_status != "arrived" and status == "arrived":
        for it in items:
            pid = it["product_id"]
            qty = int(it["qty"])
            cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
            cur.execute("UPDATE stock SET qty = qty + ? WHERE product_id=?", (qty, pid))

    # CofniÄ™cie Z arrived na inny status: towar wraca jako "w drodze" -> odejmij ze stanu.
    elif INVENTORY_AUTOMATION_ENABLED and old_status == "arrived" and status != "arrived":
        for it in items:
            pid = it["product_id"]
            qty = int(it["qty"])
            cur.execute("INSERT OR IGNORE INTO stock(product_id, qty) VALUES (?, 0)", (pid,))
            cur.execute("UPDATE stock SET qty = qty - ? WHERE product_id=?", (qty, pid))

    cur.execute("UPDATE material_orders SET status=? WHERE id=?", (status, package_id))
    c.commit()
    c.close()
    return redirect(url_for("material_orders"))

@app.post("/material-orders/<int:package_id>/tracking")
def material_order_tracking(package_id):
    tracking = norm(request.form.get("tracking"))

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT id FROM material_orders WHERE id=?", (package_id,))
    if not cur.fetchone():
        c.close()
        abort(404)

    cur.execute("UPDATE material_orders SET tracking=? WHERE id=?", (tracking, package_id))
    c.commit()
    c.close()

    ref = request.referrer or ""
    if ref.endswith(f"/material-orders/{package_id}"):
        return redirect(url_for("material_order_detail", package_id=package_id))
    return redirect(url_for("material_orders"))

@app.get("/material-orders/<int:package_id>/print")
def material_order_print(package_id):
    c = conn(); cur = c.cursor()
    pack = cur.execute("SELECT * FROM material_orders WHERE id=?", (package_id,)).fetchone()
    if not pack:
        c.close(); abort(404)
    items = cur.execute("""SELECT mi.sku,mi.qty,p.name,p.unit FROM material_order_items mi
        LEFT JOIN products p ON p.id=mi.product_id WHERE mi.package_id=? ORDER BY mi.id""", (package_id,)).fetchall()
    c.close()
    return render_template_string('''<!doctype html><html lang="pl"><meta charset="utf-8"><title>{{p.package_no}}</title><style>body{font:14px Arial,sans-serif;max-width:900px;margin:35px auto;color:#111}table{width:100%;border-collapse:collapse;margin-top:24px}th,td{border:1px solid #222;padding:9px;text-align:left}button{padding:8px 12px}@media print{button{display:none}}</style><button onclick="print()">Drukuj</button><h1>Zamówienie materiałów {{p.package_no}}</h1><p>Data utworzenia: {{p.created_at[:10]}}</p><p><b>Dostawca / uwagi:</b> {{p.note or '—'}}</p><table><thead><tr><th>Materiał</th><th>Jednostka</th><th>Ilość</th></tr></thead><tbody>{% for i in items %}<tr><td>{{i.name or i.sku}}</td><td>{{i.unit or 'szt.'}}</td><td>{{i.qty}}</td></tr>{% else %}<tr><td colspan="3">Brak pozycji.</td></tr>{% endfor %}</tbody></table><div style="margin-top:80px;display:flex;justify-content:space-between"><span>____________________________<br>Osoba zamawiająca</span><span>____________________________<br>Akceptacja dostawcy</span></div></html>''', p=pack, items=items)

@app.get("/material-orders/<int:package_id>")
def material_order_detail(package_id):
    # WyĹ‚Ä…czony pull z Supabase tylko dla moduĹ‚u Zamówienia materiałów.
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT * FROM material_orders WHERE id=?", (package_id,))
    pack = cur.fetchone()
    if not pack:
        c.close()
        abort(404)

    cur.execute("SELECT id, sku, model, name FROM products ORDER BY sku LIMIT 5000")
    products_rows = cur.fetchall()

    cur.execute("""
      SELECT ci.*, p.model, p.name
      FROM material_order_items ci
      JOIN products p ON p.id=ci.product_id
      WHERE ci.package_id=?
      ORDER BY ci.id DESC
    """, (package_id,))
    items = cur.fetchall()
    c.close()

    tpl = r"""
    {% extends "base.html" %}
    {% block content %}
      <div class="card">
        <div class="flex">
          <h1 style="margin:0;">Paczka {{ pack['package_no'] }}</h1>
          <span class="badge">{{ pack['status'] }}</span>
          <a class="btn right" href="{{ url_for('material_orders') }}">â† Lista paczek</a>
        </div>
        <div class="muted">Tracking: {{ pack['tracking'] or '-' }}</div>
        <form method="post" action="{{ url_for('material_order_tracking', package_id=pack['id']) }}" class="flex" style="margin-top:10px;">
          <input name="tracking" value="{{ pack['tracking'] or '' }}" placeholder="nr trackingu" style="width:260px;">
          <button class="btn" type="submit">ZmieĹ„ tracking</button>
          {% if pack['tracking'] %}
            <a class="btn" target="_blank" href="https://t.17track.net/en#nums={{ pack['tracking']|urlencode }}">OtwĂłrz 17TRACK</a>
          {% endif %}
        </form>
      </div>

      <div class="card">
        <h2>Dodaj zawartoĹ›Ä‡ paczki</h2>
        <form method="post" action="{{ url_for('material_order_item_add', package_id=pack['id']) }}" class="items-row">
          <div>
            <label class="muted small">Produkt</label>
            <select name="product_id" required>
              <option value="">-- wybierz --</option>
              {% for p in products %}
                <option value="{{ p['id'] }}">{{ p['sku'] }}{% if p['model'] %} â€˘ {{ p['model'] }}{% endif %}{% if p['name'] %} â€˘ {{ p['name'] }}{% endif %}</option>
              {% endfor %}
            </select>
          </div>
          <div>
            <label class="muted small">IloĹ›Ä‡</label>
            <input name="qty" value="1" required>
          </div>
          <div class="flex" style="align-items:flex-end;">
            <button class="btn primary" type="submit">Dodaj</button>
          </div>
        </form>
      </div>

      <div class="card">
        <h2>ZawartoĹ›Ä‡ paczki</h2>
        <table>
          <thead>
            <tr><th>SKU</th><th>Model / Nazwa</th><th>IloĹ›Ä‡</th><th>Data</th><th>Akcje</th></tr>
          </thead>
          <tbody>
            {% for it in items %}
              <tr>
                <td><b>{{ it['sku'] }}</b></td>
                <td>{{ it['model'] or '' }}{% if it['name'] %}<div class="muted">{{ it['name'] }}</div>{% endif %}</td>
                <td><span class="badge">{{ it['qty'] }}</span></td>
                <td class="muted">{{ it['created_at'] }}</td>
                <td>
                  <form method="post" action="{{ url_for('material_order_item_delete', package_id=pack['id'], item_id=it['id']) }}" onsubmit="return confirm('UsunÄ…Ä‡ pozycjÄ™?')">
                    <button class="btn danger" type="submit">UsuĹ„</button>
                  </form>
                </td>
              </tr>
            {% endfor %}
            {% if not items %}
              <tr><td colspan="5" class="muted">Brak pozycji w paczce.</td></tr>
            {% endif %}
          </tbody>
        </table>
      </div>
    {% endblock %}
    """
    return render_template_string(tpl, title=f"Paczka {pack['package_no']}", base_url=BASE_URL, db_path=DB_PATH,
                                  pack=pack, products=products_rows, items=items)


@app.post("/material-orders/<int:package_id>/delete")
def material_order_delete(package_id):
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT status FROM material_orders WHERE id=?", (package_id,))
    pack = cur.fetchone()
    if not pack:
        c.close()
        abort(404)

    if norm(pack["status"]).lower() == "arrived":
        c.close()
        return "Nie moĹĽna usunÄ…Ä‡ paczki ARRIVED", 400

    if supabase_enabled():
        try:
            cur.execute("SELECT id FROM material_order_items WHERE package_id=?", (package_id,))
            item_ids = [int(r["id"]) for r in cur.fetchall()]
            for iid in item_ids:
                supabase_delete_rows("material_order_items", {"id": iid})
            supabase_delete_rows("material_orders", {"id": package_id})
        except Exception:
            pass

    cur.execute("DELETE FROM material_order_items WHERE package_id=?", (package_id,))
    cur.execute("DELETE FROM material_orders WHERE id=?", (package_id,))
    c.commit()
    c.close()
    return redirect(url_for("material_orders"))

@app.post("/material-orders/<int:package_id>/items/add")
def material_order_item_add(package_id):
    product_id = to_int(request.form.get("product_id"), 0)
    qty = to_int(request.form.get("qty"), 0)
    if product_id <= 0 or qty <= 0:
        return "NieprawidĹ‚owy produkt lub iloĹ›Ä‡", 400

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT sku FROM products WHERE id=?", (product_id,))
    p = cur.fetchone()
    if not p:
        c.close()
        return "Produkt nie istnieje", 404

    cur.execute("SELECT id FROM material_orders WHERE id=?", (package_id,))
    if not cur.fetchone():
        c.close()
        return "Paczka nie istnieje", 404

    cur.execute(
        "INSERT INTO material_order_items(package_id, product_id, sku, qty, created_at) VALUES (?,?,?,?,?)",
        (package_id, product_id, p["sku"], qty, now_iso())
    )
    c.commit()
    c.close()
    return redirect(url_for("material_order_detail", package_id=package_id))

@app.post("/material-orders/<int:package_id>/items/<int:item_id>/delete")
def material_order_item_delete(package_id, item_id):
    if supabase_enabled():
        supabase_delete_rows("material_order_items", {"id": item_id})

    c = conn()
    cur = c.cursor()
    cur.execute("DELETE FROM material_order_items WHERE id=? AND package_id=?", (item_id, package_id))
    c.commit()
    c.close()
    return redirect(url_for("material_order_detail", package_id=package_id))


# =========================
# RUN
# =========================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", "5000")), debug=os.environ.get("FLASK_DEBUG") == "1")
